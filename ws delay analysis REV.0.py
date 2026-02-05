import matplotlib.pyplot as plt
import time
import math
import datetime
from datetime import datetime, timezone, timedelta
import socket
import pytz
import tkinter as tk
from tkinter.ttk import *
from tkinter import messagebox, filedialog
from time import strftime
import ntplib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import json
hostname = socket.gethostname()    
IPAddr = socket.gethostbyname(hostname)  
g=str(IPAddr)
w=str("192.168.1.163")

hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
g=str(IPAddr)
w=str("192.168.1.163")
if(g!=w):
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")

# Global calendar configuration
calendar_config = {
    'working_days': [0, 1, 2, 3, 4],  # Monday-Friday (0=Monday, 6=Sunday)
    'holidays': [],  # List of holiday dates
    'vacation_periods': []  # List of (start_date, end_date) tuples
}

class SimpleDatePicker:
    """Simple date picker using standard tkinter widgets"""
    
    def __init__(self, parent, initial_date=None):
        self.frame = tk.Frame(parent)
        
        if initial_date is None:
            initial_date = datetime.now().date()
        
        self.year = tk.IntVar(value=initial_date.year)
        self.month = tk.IntVar(value=initial_date.month)
        self.day = tk.IntVar(value=initial_date.day)
        
        # Year
        tk.Label(self.frame, text="Year:").grid(row=0, column=0, padx=2)
        year_spin = tk.Spinbox(self.frame, from_=2000, to=2100, textvariable=self.year, width=6)
        year_spin.grid(row=0, column=1, padx=2)
        
        # Month
        tk.Label(self.frame, text="Month:").grid(row=0, column=2, padx=2)
        month_spin = tk.Spinbox(self.frame, from_=1, to=12, textvariable=self.month, width=4)
        month_spin.grid(row=0, column=3, padx=2)
        
        # Day
        tk.Label(self.frame, text="Day:").grid(row=0, column=4, padx=2)
        day_spin = tk.Spinbox(self.frame, from_=1, to=31, textvariable=self.day, width=4)
        day_spin.grid(row=0, column=5, padx=2)
    
    def get_date(self):
        """Return the selected date"""
        try:
            return datetime(self.year.get(), self.month.get(), self.day.get()).date()
        except ValueError:
            messagebox.showerror("Invalid Date", "Please enter a valid date")
            return datetime.now().date()
    
    def pack(self, **kwargs):
        self.frame.pack(**kwargs)
    
    def grid(self, **kwargs):
        self.frame.grid(**kwargs)

class WorkingCalendar:
    """Manages working days, weekends, holidays, and vacations"""
    
    def __init__(self, working_days=None, holidays=None, vacation_periods=None):
        self.working_days = working_days if working_days else [0, 1, 2, 3, 4]
        self.holidays = set(holidays) if holidays else set()
        self.vacation_periods = vacation_periods if vacation_periods else []
    
    def is_working_day(self, date):
        """Check if a given date is a working day"""
        # Check if it's a weekend
        if date.weekday() not in self.working_days:
            return False
        
        # Check if it's a holiday
        if date in self.holidays:
            return False
        
        # Check if it's within a vacation period
        for start, end in self.vacation_periods:
            if start <= date <= end:
                return False
        
        return True
    
    def add_working_days(self, start_date, days):
        """Add a number of working days to a date"""
        current_date = start_date
        days_added = 0
        
        while days_added < days:
            current_date += timedelta(days=1)
            if self.is_working_day(current_date):
                days_added += 1
        
        return current_date
    
    def count_working_days(self, start_date, end_date):
        """Count working days between two dates"""
        if start_date > end_date:
            return 0
        
        working_days = 0
        current_date = start_date
        
        while current_date <= end_date:
            if self.is_working_day(current_date):
                working_days += 1
            current_date += timedelta(days=1)
        
        return working_days
    
    def get_calendar_days(self, working_days):
        """Convert working days to calendar days (approximate)"""
        # This is an estimate based on the working days configuration
        if not self.working_days:
            return working_days
        
        work_days_per_week = len(self.working_days)
        calendar_days = (working_days * 7) / work_days_per_week
        return int(calendar_days)

# Global calendar instance
working_calendar = None

def get_international_date():
    """
    Queries an NTP server to get the current UTC time.
    Returns a datetime object with UTC timezone, or None if an error occurs.
    """
    try:
        client = ntplib.NTPClient()
        response = client.request("pool.ntp.org", version=3)
        utc_time = datetime.fromtimestamp(response.tx_time, tz=timezone.utc)
        return utc_time
    except Exception as e:
        print(f"Error getting international date from NTP server: {e}")
        return None

def run_application():
    """
    Contains the core logic of the application, including the license expiry check
    """
    try:
        today_utc_dt = get_international_date()
        
        if today_utc_dt is None:
            print("Failed to retrieve international date. Application cannot proceed.")
            return False
        
        print(f"Today's Date (UTC): {today_utc_dt.strftime('%Y-%m-%d %H:%M:%S UTC')}")
        
        expiry_date = datetime(2026, 8, 19, 0, 0, 0, tzinfo=timezone.utc)
        
        if today_utc_dt > expiry_date:
            print("License expired. This software is no longer usable.")
            return False
        
        hostname = socket.gethostname()
        IPAddr = socket.gethostbyname(hostname)
        g = str(IPAddr)
        print(f"Local IP Address: {g}")
        return True
        
    except Exception as e:
        print(f"An error occurred in run_application: {e}")
        return False

# Main output file
f = open("WSDELAYANALYSIS.txt", "a")
print("This software is created by Engineer Wael Sherif Selim,email:wael.sherif.selim@gmail.com")
print("This software is created by Engineer Wael Sherif Selim,email:wael.sherif.selim@gmail.com", file=f)
print("WSDELAYANALYSIS V.4 - Enhanced Version with Calendar Support")
print("WSDELAYANALYSIS V.4 - Enhanced Version with Calendar Support", file=f)

# Generate fields for 20 events
def generate_fields():
    base_fields = ['project name', 'project id']
    
    for event_num in range(1, 21):
        base_fields.extend([
            f'delay event {event_num} description',
            f'event {event_num} duration',
        ])
        
        for act_num in range(1, 4):
            base_fields.extend([
                f'event {event_num} predecessor activity {act_num} description',
                f'event {event_num} predecessor activity {act_num} start day',
                f'event {event_num} predecessor activity {act_num} finish day',
                f'relation between predecessor activity {act_num} and event {event_num} , for fs write 1, ss write 2,ff write 3, sf write 4',
                f'lag between activity {act_num} and event {event_num}',
            ])
    
    base_fields.extend([
        'Reference start date(YYYY,MM,DD)',
        'Overhead and profit percentage',
        'contract value',
        'contract duration in days'
    ])
    
    return tuple(base_fields)

fields = generate_fields()

# Global dictionary to store calculated event times
event_times = {}
last_entries = None

def open_calendar_settings():
    """Open calendar configuration window"""
    global calendar_config, working_calendar
    
    cal_window = tk.Toplevel()
    cal_window.title("Calendar Settings")
    cal_window.geometry("700x750")
    
    # Working Days Section
    tk.Label(cal_window, text="Working Days Configuration", font=("Georgia", 14, "bold")).pack(pady=10)
    
    working_frame = tk.LabelFrame(cal_window, text="Select Working Days", padx=10, pady=10)
    working_frame.pack(padx=10, pady=5, fill="x")
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_vars = []
    
    for i, day in enumerate(days):
        var = tk.IntVar(value=1 if i in calendar_config['working_days'] else 0)
        cb = tk.Checkbutton(working_frame, text=day, variable=var, font=("Georgia", 10))
        cb.pack(anchor='w')
        day_vars.append(var)
    
    # Holidays Section
    tk.Label(cal_window, text="Holidays Configuration", font=("Georgia", 14, "bold")).pack(pady=10)
    
    holiday_frame = tk.LabelFrame(cal_window, text="Add Holidays", padx=10, pady=10)
    holiday_frame.pack(padx=10, pady=5, fill="both", expand=True)
    
    tk.Label(holiday_frame, text="Select Holiday Date:", font=("Georgia", 10)).pack(pady=5)
    holiday_date = SimpleDatePicker(holiday_frame)
    holiday_date.pack(pady=5)
    
    holidays_listbox = tk.Listbox(holiday_frame, height=8, font=("Georgia", 9))
    holidays_listbox.pack(pady=5, fill="both", expand=True)
    
    # Load existing holidays
    for holiday in sorted(calendar_config['holidays']):
        holidays_listbox.insert(tk.END, holiday.strftime('%Y-%m-%d'))
    
    def add_holiday():
        date = holiday_date.get_date()
        if date not in calendar_config['holidays']:
            calendar_config['holidays'].append(date)
            holidays_listbox.insert(tk.END, date.strftime('%Y-%m-%d'))
            holidays_listbox.see(tk.END)
    
    def remove_holiday():
        selection = holidays_listbox.curselection()
        if selection:
            idx = selection[0]
            date_str = holidays_listbox.get(idx)
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            calendar_config['holidays'].remove(date)
            holidays_listbox.delete(idx)
    
    btn_frame1 = tk.Frame(holiday_frame)
    btn_frame1.pack(pady=5)
    tk.Button(btn_frame1, text="Add Holiday", command=add_holiday, bg="#4CAF50", fg="white", padx=10, pady=3).pack(side="left", padx=5)
    tk.Button(btn_frame1, text="Remove Selected", command=remove_holiday, bg="#f44336", fg="white", padx=10, pady=3).pack(side="left", padx=5)
    
    # Vacation Periods Section
    vacation_frame = tk.LabelFrame(cal_window, text="Vacation Periods", padx=10, pady=10)
    vacation_frame.pack(padx=10, pady=5, fill="both", expand=True)
    
    tk.Label(vacation_frame, text="Start Date:", font=("Georgia", 10)).grid(row=0, column=0, padx=5, pady=5, sticky='w')
    vacation_start = SimpleDatePicker(vacation_frame)
    vacation_start.grid(row=0, column=1, padx=5, pady=5)
    
    tk.Label(vacation_frame, text="End Date:", font=("Georgia", 10)).grid(row=1, column=0, padx=5, pady=5, sticky='w')
    vacation_end = SimpleDatePicker(vacation_frame)
    vacation_end.grid(row=1, column=1, padx=5, pady=5)
    
    vacations_listbox = tk.Listbox(vacation_frame, height=5, font=("Georgia", 9))
    vacations_listbox.grid(row=2, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    # Load existing vacation periods
    for start, end in calendar_config['vacation_periods']:
        vacations_listbox.insert(tk.END, f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}")
    
    def add_vacation():
        start = vacation_start.get_date()
        end = vacation_end.get_date()
        if start <= end:
            calendar_config['vacation_periods'].append((start, end))
            vacations_listbox.insert(tk.END, f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}")
            vacations_listbox.see(tk.END)
        else:
            messagebox.showwarning("Invalid Dates", "Start date must be before or equal to end date")
    
    def remove_vacation():
        selection = vacations_listbox.curselection()
        if selection:
            idx = selection[0]
            calendar_config['vacation_periods'].pop(idx)
            vacations_listbox.delete(idx)
    
    btn_frame2 = tk.Frame(vacation_frame)
    btn_frame2.grid(row=3, column=0, columnspan=2, pady=5)
    tk.Button(btn_frame2, text="Add Vacation Period", command=add_vacation, bg="#4CAF50", fg="white", padx=10, pady=3).pack(side="left", padx=5)
    tk.Button(btn_frame2, text="Remove Selected", command=remove_vacation, bg="#f44336", fg="white", padx=10, pady=3).pack(side="left", padx=5)
    
    # Save and Close Button
    def save_calendar():
        # Update working days
        calendar_config['working_days'] = [i for i, var in enumerate(day_vars) if var.get() == 1]
        
        if not calendar_config['working_days']:
            messagebox.showerror("Error", "You must select at least one working day!")
            return
        
        # Update global calendar
        global working_calendar
        working_calendar = WorkingCalendar(
            working_days=calendar_config['working_days'],
            holidays=calendar_config['holidays'],
            vacation_periods=calendar_config['vacation_periods']
        )
        
        messagebox.showinfo("Success", "Calendar settings saved successfully!")
        cal_window.destroy()
    
    tk.Button(cal_window, text="Save and Close", command=save_calendar, 
             bg="#2196F3", fg="white", font=("Georgia", 12, "bold"), padx=20, pady=10).pack(pady=15)

def project_name(entries):
    """Process all project data and calculate delays"""
    global event_times, last_entries, working_calendar
    
    # Initialize calendar if not already done
    if working_calendar is None:
        working_calendar = WorkingCalendar(
            working_days=calendar_config['working_days'],
            holidays=calendar_config['holidays'],
            vacation_periods=calendar_config['vacation_periods']
        )
    
    last_entries = entries
    event_times = {}
    
    # Project info
    pr = entries['project name'].get()
    print("project:", pr)
    print("project:", pr, file=f)
    prid = entries['project id'].get()
    print("project id:", prid)
    print("project id:", prid, file=f)
    
    print("delay is compensable such as owner directed change and errors or changes in design")
    print("delay is compensable such as owner directed change and errors or changes in design", file=f)
    print("contractor deserves EOT")
    print("contractor deserves EOT", file=f)
    
    # Get reference date
    try:
        stdt = entries['Reference start date(YYYY,MM,DD)'].get()
        format_code = "%Y,%m,%d"
        start = time.strptime(stdt, format_code)
        reference_date = datetime(start.tm_year, start.tm_mon, start.tm_mday).date()
    except Exception as e:
        messagebox.showerror("Error", f"Invalid reference date format: {e}")
        return
    
    # Print calendar info
    print(f"\nCalendar Configuration:")
    print(f"Calendar Configuration:", file=f)
    days_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    working_days_names = [days_names[i] for i in working_calendar.working_days]
    print(f"Working Days: {', '.join(working_days_names)}")
    print(f"Working Days: {', '.join(working_days_names)}", file=f)
    print(f"Holidays Count: {len(working_calendar.holidays)}")
    print(f"Holidays Count: {len(working_calendar.holidays)}", file=f)
    print(f"Vacation Periods: {len(working_calendar.vacation_periods)}")
    print(f"Vacation Periods: {len(working_calendar.vacation_periods)}\n", file=f)
    
    # Process all 20 events
    active_events = []
    for event_num in range(1, 21):
        try:
            # Check if event has data
            desc = entries[f'delay event {event_num} description'].get()
            dur = entries[f'event {event_num} duration'].get()
            
            # Skip if no description or duration is 0 or empty
            if desc and desc.strip() and dur and dur.strip() and int(dur) > 0:
                process_event(entries, event_num, reference_date)
                active_events.append(event_num)
        except (ValueError, KeyError) as e:
            # Skip this event if there's an error
            print(f"Skipping event {event_num}: {e}")
            continue
    
    if not active_events:
        messagebox.showwarning("No Events", "No valid delay events found. Please enter at least one event.")
        return
    
    # Calculate overlaps and total delay for active events
    event_durations = []
    for event_num in active_events:
        if event_num in event_times:
            event_durations.append({
                'num': event_num,
                'start': event_times[event_num]['start_working_day'],
                'finish': event_times[event_num]['finish_working_day'],
                'duration': event_times[event_num]['duration'],
                'start_date': event_times[event_num]['start_date'],
                'finish_date': event_times[event_num]['finish_date']
            })
    
    # Calculate total concurrency
    total_concurrency = calculate_concurrency(event_durations)
    print(f"Total concurrency: {total_concurrency} working days")
    print(f"Total concurrency: {total_concurrency} working days", file=f)
    
    # Calculate total delay
    total_duration = sum([ev['duration'] for ev in event_durations])
    tdl = total_duration - total_concurrency
    print(f"total delay = {tdl} working days")
    print(f"total delay = {tdl} working days", file=f)
    
    # Calculate prolongation cost
    try:
        ohp = float(entries['Overhead and profit percentage'].get())
        cvlu = float(entries['contract value'].get())
        codu = float(entries['contract duration in days'].get())
        proc = ohp * 0.01 * cvlu * tdl / codu
        print(f"Prolongation cost = ${proc:,.2f}")
        print(f"Prolongation cost = ${proc:,.2f}", file=f)
    except (ValueError, KeyError) as e:
        print(f"Error calculating prolongation cost: {e}")
        proc = 0
    
    # Create Gantt chart
    create_gantt_chart(entries, active_events)
    
    messagebox.showinfo("Success", f"Analysis complete!\n\nActive Events: {len(active_events)}\nTotal Delay: {tdl} working days\nProlongation Cost: ${proc:,.2f}")

def calculate_concurrency(event_durations):
    """Calculate total concurrency between events using working days"""
    if len(event_durations) <= 1:
        return 0
    
    total_concurrency = 0
    
    # Check each pair of events
    for i in range(len(event_durations)):
        for j in range(i + 1, len(event_durations)):
            ev1 = event_durations[i]
            ev2 = event_durations[j]
            
            # Calculate overlap using dates
            overlap_start = max(ev1['start_date'], ev2['start_date'])
            overlap_end = min(ev1['finish_date'], ev2['finish_date'])
            
            if overlap_start <= overlap_end:
                # Count working days in overlap period
                overlap = working_calendar.count_working_days(overlap_start, overlap_end)
                total_concurrency += overlap
    
    return total_concurrency

def process_event(entries, event_num, reference_date):
    """Process a single event with all its predecessor activities"""
    global event_times
    
    dev = entries[f'delay event {event_num} description'].get()
    print(f"event {event_num} is {dev}")
    print(f"event {event_num} is {dev}", file=f)
    
    dur = int(entries[f'event {event_num} duration'].get())
    print(f"duration of event {event_num} is {dur} working days")
    print(f"duration of event {event_num} is {dur} working days", file=f)
    
    # Process all 3 predecessor activities for this event
    sf_times = []  # Store (working_day, date) tuples
    for act_num in range(1, 4):
        try:
            process_predecessor(entries, event_num, act_num, dur, sf_times, reference_date)
        except (ValueError, KeyError) as e:
            # Skip this predecessor if there's an error
            continue
    
    # Calculate event start and finish
    if sf_times:
        # Get the latest start time
        sev_working_day = max([s[0] for s in sf_times])
        sev_date = max([s[1] for s in sf_times])
        
        # Calculate finish date by adding working days
        fev_date = working_calendar.add_working_days(sev_date, dur)
        fev_working_day = sev_working_day + dur
        
        print(f"event {event_num} starts on working day {sev_working_day} ({sev_date}) and ends on working day {fev_working_day} ({fev_date})")
        print(f"event {event_num} starts on working day {sev_working_day} ({sev_date}) and ends on working day {fev_working_day} ({fev_date})", file=f)
        
        event_times[event_num] = {
            'start_working_day': sev_working_day,
            'finish_working_day': fev_working_day,
            'duration': dur,
            'start_date': sev_date,
            'finish_date': fev_date
        }
    else:
        # If no predecessors, start at reference date
        sev_working_day = 0
        sev_date = reference_date
        fev_date = working_calendar.add_working_days(reference_date, dur)
        fev_working_day = dur
        
        print(f"event {event_num} starts on working day {sev_working_day} ({sev_date}) and ends on working day {fev_working_day} ({fev_date})")
        print(f"event {event_num} starts on working day {sev_working_day} ({sev_date}) and ends on working day {fev_working_day} ({fev_date})", file=f)
        
        event_times[event_num] = {
            'start_working_day': sev_working_day,
            'finish_working_day': fev_working_day,
            'duration': dur,
            'start_date': sev_date,
            'finish_date': fev_date
        }

def process_predecessor(entries, event_num, act_num, dur, sf_times, reference_date):
    """Process a single predecessor activity for an event"""
    deac = entries[f'event {event_num} predecessor activity {act_num} description'].get()
    if not deac or not deac.strip():
        return
    
    print(f"event {event_num} predecessor activity {act_num} is {deac}")
    print(f"event {event_num} predecessor activity {act_num} is {deac}", file=f)
    
    sda = int(entries[f'event {event_num} predecessor activity {act_num} start day'].get())
    print(f"activity {act_num} start working day is {sda}")
    print(f"activity {act_num} start working day is {sda}", file=f)
    
    fda = int(entries[f'event {event_num} predecessor activity {act_num} finish day'].get())
    print(f"activity {act_num} finish working day is {fda}")
    print(f"activity {act_num} finish working day is {fda}", file=f)
    
    # Calculate actual dates for predecessor
    pred_start_date = working_calendar.add_working_days(reference_date, sda)
    pred_finish_date = working_calendar.add_working_days(reference_date, fda)
    
    rlda = int(entries[f'relation between predecessor activity {act_num} and event {event_num} , for fs write 1, ss write 2,ff write 3, sf write 4'].get())
    lgda = int(entries[f'lag between activity {act_num} and event {event_num}'].get())
    
    if rlda == 1:  # FS
        print(f"relation between activity {act_num} and event {event_num} is FS with lag {lgda} working days")
        print(f"relation between activity {act_num} and event {event_num} is FS with lag {lgda} working days", file=f)
        
        sfs_working_day = fda + lgda
        sfs_date = working_calendar.add_working_days(pred_finish_date, lgda)
        
    elif rlda == 2:  # SS
        print(f"relation between activity {act_num} and event {event_num} is SS with lag {lgda} working days")
        print(f"relation between activity {act_num} and event {event_num} is SS with lag {lgda} working days", file=f)
        
        sfs_working_day = sda + lgda
        sfs_date = working_calendar.add_working_days(pred_start_date, lgda)
        
    elif rlda == 3:  # FF
        print(f"relation between activity {act_num} and event {event_num} is FF with lag {lgda} working days")
        print(f"relation between activity {act_num} and event {event_num} is FF with lag {lgda} working days", file=f)
        
        ffs_working_day = fda + lgda
        ffs_date = working_calendar.add_working_days(pred_finish_date, lgda)
        
        # Work backwards from finish
        sfs_working_day = ffs_working_day - dur
        # Calculate start date by subtracting working days
        sfs_date = ffs_date
        for _ in range(dur):
            sfs_date -= timedelta(days=1)
            while not working_calendar.is_working_day(sfs_date):
                sfs_date -= timedelta(days=1)
        
    else:  # SF
        print(f"relation between activity {act_num} and event {event_num} is SF with lag {lgda} working days")
        print(f"relation between activity {act_num} and event {event_num} is SF with lag {lgda} working days", file=f)
        
        ffs_working_day = sda + lgda
        ffs_date = working_calendar.add_working_days(pred_start_date, lgda)
        
        # Work backwards from finish
        sfs_working_day = ffs_working_day - dur
        # Calculate start date by subtracting working days
        sfs_date = ffs_date
        for _ in range(dur):
            sfs_date -= timedelta(days=1)
            while not working_calendar.is_working_day(sfs_date):
                sfs_date -= timedelta(days=1)
    
    sf_times.append((sfs_working_day, sfs_date))

def create_gantt_chart(entries, active_events):
    """Create a Gantt chart visualization of the events"""
    global event_times
    
    if not active_events:
        return
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(14, max(8, len(active_events) * 0.5)))
    
    # Generate colors
    colors = plt.cm.tab20(range(len(active_events)))
    
    # Plot events using calculated dates
    y_pos = 0
    y_labels = []
    
    # Get reference date
    stdt = last_entries['Reference start date(YYYY,MM,DD)'].get()
    format_code = "%Y,%m,%d"
    start = time.strptime(stdt, format_code)
    reference_date = datetime(start.tm_year, start.tm_mon, start.tm_mday).date()
    
    for i, event_num in enumerate(active_events):
        if event_num in event_times:
            start_date = event_times[event_num]['start_date']
            finish_date = event_times[event_num]['finish_date']
            
            # Calculate calendar days from reference
            start_calendar = (start_date - reference_date).days
            duration_calendar = (finish_date - start_date).days + 1
            
            ax.barh(y=y_pos, width=duration_calendar, left=start_calendar, color=colors[i], 
                   label=f'Event {event_num}', alpha=0.8, edgecolor='black')
            
            # Add text with working days info
            working_days = event_times[event_num]['duration']
            ax.text(start_calendar + duration_calendar/2, y_pos, 
                   f'{working_days}wd', ha='center', va='center', 
                   fontweight='bold', fontsize=8)
            
            y_labels.append(f'Event {event_num}')
            y_pos += 1
    
    ax.set_yticks(range(len(y_labels)))
    ax.set_yticklabels(y_labels)
    ax.set_xlabel('Calendar Days from Reference Date', fontsize=12)
    ax.set_ylabel('Delay Events', fontsize=12)
    ax.set_title('Delay Events Timeline - Gantt Chart (with Calendar)\nwd = working days', 
                fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, axis='x')
    ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=8)
    
    plt.tight_layout()
    plt.savefig('delay_events_timeline_calendar.png', dpi=300, bbox_inches='tight')
    plt.show()

def export_to_excel():
    """Export analysis results to Excel"""
    global event_times, last_entries
    
    if not event_times or not last_entries:
        messagebox.showwarning("No Data", "Please run the analysis first before exporting.")
        return
    
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="WSDELAYANALYSIS_Calendar_Report.xlsx"
        )
        
        if not filename:
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delay Analysis"
        
        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Project Information
        ws['A1'] = "WSDELAYANALYSIS V.4 - Delay Analysis Report with Calendar"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        row = 3
        ws[f'A{row}'] = "Project Name:"
        ws[f'B{row}'] = last_entries['project name'].get()
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Project ID:"
        ws[f'B{row}'] = last_entries['project id'].get()
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Reference Date:"
        ws[f'B{row}'] = last_entries['Reference start date(YYYY,MM,DD)'].get()
        row += 1
        
        ws[f'A{row}'] = "Contract Value:"
        ws[f'B{row}'] = f"${float(last_entries['contract value'].get()):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Contract Duration:"
        ws[f'B{row}'] = f"{last_entries['contract duration in days'].get()} days"
        row += 1
        
        ws[f'A{row}'] = "Overhead & Profit %:"
        ws[f'B{row}'] = f"{last_entries['Overhead and profit percentage'].get()}%"
        row += 2
        
        # Calendar Info
        ws[f'A{row}'] = "Calendar Configuration"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        days_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        working_days_names = [days_names[i] for i in working_calendar.working_days]
        ws[f'A{row}'] = "Working Days:"
        ws[f'B{row}'] = ', '.join(working_days_names)
        row += 1
        
        ws[f'A{row}'] = "Holidays Count:"
        ws[f'B{row}'] = len(working_calendar.holidays)
        row += 1
        
        ws[f'A{row}'] = "Vacation Periods:"
        ws[f'B{row}'] = len(working_calendar.vacation_periods)
        row += 2
        
        # Event Details Header
        headers = ['Event #', 'Description', 'Start Day', 'Finish Day', 'Duration (WD)', 
                  'Start Date', 'End Date', 'Calendar Days']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Event Details
        total_duration = 0
        for event_num in sorted(event_times.keys()):
            desc = last_entries[f'delay event {event_num} description'].get()
            start_day = event_times[event_num]['start_working_day']
            finish_day = event_times[event_num]['finish_working_day']
            duration = event_times[event_num]['duration']
            start_date = event_times[event_num]['start_date']
            finish_date = event_times[event_num]['finish_date']
            calendar_days = (finish_date - start_date).days + 1
            
            ws.cell(row=row, column=1, value=event_num).border = border
            ws.cell(row=row, column=2, value=desc).border = border
            ws.cell(row=row, column=3, value=start_day).border = border
            ws.cell(row=row, column=4, value=finish_day).border = border
            ws.cell(row=row, column=5, value=duration).border = border
            ws.cell(row=row, column=6, value=start_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=7, value=finish_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=8, value=calendar_days).border = border
            
            total_duration += duration
            row += 1
        
        # Summary
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Calculate concurrency
        event_durations = []
        for event_num in event_times:
            event_durations.append({
                'num': event_num,
                'start': event_times[event_num]['start_working_day'],
                'finish': event_times[event_num]['finish_working_day'],
                'duration': event_times[event_num]['duration'],
                'start_date': event_times[event_num]['start_date'],
                'finish_date': event_times[event_num]['finish_date']
            })
        
        total_concurrency = calculate_concurrency(event_durations)
        tdl = total_duration - total_concurrency
        
        ohp = float(last_entries['Overhead and profit percentage'].get())
        cvlu = float(last_entries['contract value'].get())
        codu = float(last_entries['contract duration in days'].get())
        proc = ohp * 0.01 * cvlu * tdl / codu
        
        ws[f'A{row}'] = "Total Event Duration:"
        ws[f'B{row}'] = f"{total_duration} working days"
        row += 1
        
        ws[f'A{row}'] = "Total Concurrency:"
        ws[f'B{row}'] = f"{total_concurrency} working days"
        row += 1
        
        ws[f'A{row}'] = "Net Delay:"
        ws[f'B{row}'] = f"{tdl} working days"
        ws[f'B{row}'].font = Font(bold=True, color="FF0000")
        row += 1
        
        ws[f'A{row}'] = "Prolongation Cost:"
        ws[f'B{row}'] = f"${proc:,.2f}"
        ws[f'B{row}'].font = Font(bold=True, color="FF0000", size=12)
        
        # Holidays Sheet
        if working_calendar.holidays:
            ws_holidays = wb.create_sheet(title="Holidays")
            ws_holidays['A1'] = "Holiday Dates"
            ws_holidays['A1'].font = Font(bold=True, size=12)
            ws_holidays['A1'].fill = header_fill
            
            for i, holiday in enumerate(sorted(working_calendar.holidays), start=2):
                ws_holidays[f'A{i}'] = holiday.strftime('%Y-%m-%d')
        
        # Vacation Periods Sheet
        if working_calendar.vacation_periods:
            ws_vacations = wb.create_sheet(title="Vacation Periods")
            ws_vacations['A1'] = "Start Date"
            ws_vacations['B1'] = "End Date"
            ws_vacations['C1'] = "Days"
            for cell in ['A1', 'B1', 'C1']:
                ws_vacations[cell].font = Font(bold=True, size=12)
                ws_vacations[cell].fill = header_fill
            
            for i, (start, end) in enumerate(working_calendar.vacation_periods, start=2):
                ws_vacations[f'A{i}'] = start.strftime('%Y-%m-%d')
                ws_vacations[f'B{i}'] = end.strftime('%Y-%m-%d')
                ws_vacations[f'C{i}'] = (end - start).days + 1
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        messagebox.showinfo("Success", f"Excel report exported successfully to:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export to Excel:\n{str(e)}")

def export_to_primavera():
    """Export analysis results to Primavera P6 XML format"""
    global event_times, last_entries
    
    if not event_times or not last_entries:
        messagebox.showwarning("No Data", "Please run the analysis first before exporting.")
        return
    
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".xml",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
            initialfile="WSDELAYANALYSIS_P6_Calendar.xml"
        )
        
        if not filename:
            return
        
        # Create XML structure
        root = ET.Element("Project")
        root.set("xmlns", "http://www.primavera.com/schemas")
        
        # Project info
        proj_info = ET.SubElement(root, "ProjectInfo")
        ET.SubElement(proj_info, "ProjectName").text = last_entries['project name'].get()
        ET.SubElement(proj_info, "ProjectID").text = last_entries['project id'].get()
        ET.SubElement(proj_info, "PlanStartDate").text = last_entries['Reference start date(YYYY,MM,DD)'].get().replace(',', '-')
        
        # Calendar
        calendar = ET.SubElement(root, "Calendar")
        ET.SubElement(calendar, "CalendarName").text = "Project Calendar"
        
        # Working days
        work_week = ET.SubElement(calendar, "WorkWeek")
        days_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        for i, day_name in enumerate(days_names):
            day = ET.SubElement(work_week, "Day")
            ET.SubElement(day, "DayName").text = day_name
            ET.SubElement(day, "IsWorkingDay").text = "true" if i in working_calendar.working_days else "false"
        
        # Holidays
        if working_calendar.holidays:
            holidays_elem = ET.SubElement(calendar, "Holidays")
            for holiday in sorted(working_calendar.holidays):
                holiday_elem = ET.SubElement(holidays_elem, "Holiday")
                ET.SubElement(holiday_elem, "Date").text = holiday.strftime('%Y-%m-%d')
        
        # Vacation periods
        if working_calendar.vacation_periods:
            vacations_elem = ET.SubElement(calendar, "VacationPeriods")
            for start, end in working_calendar.vacation_periods:
                vacation = ET.SubElement(vacations_elem, "VacationPeriod")
                ET.SubElement(vacation, "StartDate").text = start.strftime('%Y-%m-%d')
                ET.SubElement(vacation, "EndDate").text = end.strftime('%Y-%m-%d')
        
        # Activities
        activities = ET.SubElement(root, "Activities")
        
        for event_num in sorted(event_times.keys()):
            activity = ET.SubElement(activities, "Activity")
            
            ET.SubElement(activity, "ActivityID").text = f"DEL{event_num:03d}"
            ET.SubElement(activity, "ActivityName").text = last_entries[f'delay event {event_num} description'].get()
            ET.SubElement(activity, "ActivityType").text = "Task Dependent"
            
            start_date = event_times[event_num]['start_date']
            finish_date = event_times[event_num]['finish_date']
            duration = event_times[event_num]['duration']
            
            ET.SubElement(activity, "StartDate").text = start_date.strftime('%Y-%m-%d')
            ET.SubElement(activity, "FinishDate").text = finish_date.strftime('%Y-%m-%d')
            ET.SubElement(activity, "PlannedDuration").text = str(duration)
            ET.SubElement(activity, "DurationType").text = "Working Days"
            ET.SubElement(activity, "ActualDuration").text = str(duration)
            ET.SubElement(activity, "PercentComplete").text = "0"
            ET.SubElement(activity, "Status").text = "Not Started"
            ET.SubElement(activity, "CalendarDays").text = str((finish_date - start_date).days + 1)
        
        # Pretty print XML
        xml_str = ET.tostring(root, encoding='unicode')
        dom = minidom.parseString(xml_str)
        pretty_xml = dom.toprettyxml(indent="  ")
        
        with open(filename, 'w') as f:
            f.write(pretty_xml)
        
        messagebox.showinfo("Success", f"Primavera P6 XML exported successfully to:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export to Primavera:\n{str(e)}")

def save_calendar_config():
    """Save calendar configuration to file"""
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile="calendar_config.json"
        )
        
        if not filename:
            return
        
        config = {
            'working_days': calendar_config['working_days'],
            'holidays': [h.strftime('%Y-%m-%d') for h in calendar_config['holidays']],
            'vacation_periods': [(s.strftime('%Y-%m-%d'), e.strftime('%Y-%m-%d')) 
                                for s, e in calendar_config['vacation_periods']]
        }
        
        with open(filename, 'w') as f:
            json.dump(config, f, indent=2)
        
        messagebox.showinfo("Success", f"Calendar configuration saved to:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save calendar configuration:\n{str(e)}")

def load_calendar_config():
    """Load calendar configuration from file"""
    global calendar_config, working_calendar
    
    try:
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Calendar Configuration"
        )
        
        if not filename:
            return
        
        with open(filename, 'r') as f:
            config = json.load(f)
        
        calendar_config['working_days'] = config['working_days']
        calendar_config['holidays'] = [datetime.strptime(h, '%Y-%m-%d').date() 
                                       for h in config['holidays']]
        calendar_config['vacation_periods'] = [(datetime.strptime(s, '%Y-%m-%d').date(), 
                                                datetime.strptime(e, '%Y-%m-%d').date())
                                               for s, e in config['vacation_periods']]
        
        # Update global calendar
        working_calendar = WorkingCalendar(
            working_days=calendar_config['working_days'],
            holidays=calendar_config['holidays'],
            vacation_periods=calendar_config['vacation_periods']
        )
        
        messagebox.showinfo("Success", f"Calendar configuration loaded from:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load calendar configuration:\n{str(e)}")

def makeform(root, frame):
    """Create the input form GUI"""
    entries = {}
    row = 0
    
    # Create headers for better organization
    for field in fields:
        if 'delay event' in field and 'description' in field:
            # Add separator before each event
            sep = tk.Label(frame, text="", height=1)
            sep.grid(row=row, column=0, columnspan=2)
            row += 1
        
        lab = tk.Label(frame, width=85, text=field+": ", anchor='w', font=("georgia 10"))
        lab.grid(row=row, column=0, sticky='w', padx=5, pady=2)
        ent = tk.Entry(frame, width=40)
        ent.grid(row=row, column=1, sticky='ew', padx=5, pady=2)
        
        # Set default values
        if 'duration' in field.lower() or 'start day' in field.lower() or 'finish day' in field.lower() or 'lag' in field.lower() or 'relation' in field.lower():
            ent.insert(0, "0")
        else:
            ent.insert(0, "")
        
        entries[field] = ent
        row += 1
    
    # Buttons frame
    f1 = tk.Frame(frame)
    f1.grid(row=row, column=0, columnspan=2, pady=10)
    
    b1 = tk.Button(f1, text="Run Analysis", font=("Georgia 10"), bg="#4CAF50", fg="white",
                   padx=10, pady=5, command=(lambda e=entries: project_name(e)))
    b1.pack(side="left", padx=5)
    
    b2 = tk.Button(f1, text='Export to Excel', font=("Georgia 10"), bg="#2196F3", fg="white",
                   padx=10, pady=5, command=export_to_excel)
    b2.pack(side="left", padx=5)
    
    b3 = tk.Button(f1, text='Export to Primavera', font=("Georgia 10"), bg="#FF9800", fg="white",
                   padx=10, pady=5, command=export_to_primavera)
    b3.pack(side="left", padx=5)
    
    b4 = tk.Button(f1, text='Quit', font=("Georgia 10"), bg="#f44336", fg="white",
                   padx=10, pady=5, command=root.destroy)
    b4.pack(side="left", padx=5)
    
    # Calendar controls frame
    f2 = tk.Frame(frame, bg='#e0e0e0', relief=tk.RAISED, borderwidth=2)
    f2.grid(row=row+1, column=0, columnspan=2, pady=10, padx=10, sticky='ew')
    
    tk.Label(f2, text="Calendar Management", font=("Georgia", 12, "bold"), 
            bg='#e0e0e0').pack(pady=5)
    
    cal_btn_frame = tk.Frame(f2, bg='#e0e0e0')
    cal_btn_frame.pack(pady=5)
    
    bc1 = tk.Button(cal_btn_frame, text="Calendar Settings", font=("Georgia 9"), 
                   bg="#9C27B0", fg="white", padx=10, pady=5, command=open_calendar_settings)
    bc1.pack(side="left", padx=5)
    
    bc2 = tk.Button(cal_btn_frame, text="Save Calendar Config", font=("Georgia 9"), 
                   bg="#607D8B", fg="white", padx=10, pady=5, command=save_calendar_config)
    bc2.pack(side="left", padx=5)
    
    bc3 = tk.Button(cal_btn_frame, text="Load Calendar Config", font=("Georgia 9"), 
                   bg="#795548", fg="white", padx=10, pady=5, command=load_calendar_config)
    bc3.pack(side="left", padx=5)
    
    return entries

def onFrameConfigure(canvas):
    """Reset the scroll region to encompass the inner frame"""
    canvas.configure(scrollregion=canvas.bbox("all"))

if __name__ == '__main__':
    # Run license check
    if not run_application():
        input("Press Enter to exit...")
        exit()
    
    # Initialize default calendar
    working_calendar = WorkingCalendar(
        working_days=calendar_config['working_days'],
        holidays=calendar_config['holidays'],
        vacation_periods=calendar_config['vacation_periods']
    )
    
    # Create main GUI
    root = tk.Tk()
    root.title("WSDELAYANALYSIS V.4 - Enhanced Version with Calendar (20 Events)")
    root.geometry("1000x700")
    
    # Create main frame with scrollbars
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=1)
    
    # Create canvas
    canvas = tk.Canvas(main_frame, borderwidth=0, bg='#f0f0f0')
    
    # Create scrollbars
    vsb = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    hsb = tk.Scrollbar(main_frame, orient="horizontal", command=canvas.xview)
    
    # Configure canvas
    canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    
    # Pack scrollbars and canvas
    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    canvas.pack(side="left", fill="both", expand=True)
    
    # Create frame inside canvas
    frame = tk.Frame(canvas, bg='#f0f0f0')
    canvas.create_window((0, 0), window=frame, anchor="nw")
    
    # Bind configuration
    frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
    
    # Add title label
    title_label = tk.Label(frame, text="WSDELAYANALYSIS V.4 - Delay Analysis Software", 
                          font=("Georgia", 16, "bold"), bg='#f0f0f0', fg='#333333')
    title_label.grid(row=0, column=0, columnspan=2, pady=10)
    
    subtitle_label = tk.Label(frame, text="Enter data for up to 20 delay events (Working Days Calendar)", 
                             font=("Georgia", 10, "italic"), bg='#f0f0f0', fg='#666666')
    subtitle_label.grid(row=1, column=0, columnspan=2, pady=(0, 10))
    
    # Create form starting from row 2
    entries = makeform(root, frame)
    
    # Enable mousewheel scrolling
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    root.mainloop()

    print("contractor deserves EOT in case of non fully concurrent events,critical,excusable and compensable events only")
    print("contractor deserves EOT in case of non fully concurrent events,critical,excusable and compensable events only", file=f)
    input("Press Enter to close program. Output file created: WSDELAYANALYSIS.txt")
    f.close()
