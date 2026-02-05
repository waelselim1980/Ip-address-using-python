import matplotlib.pyplot as plt # pyright: ignore[reportMissingModuleSource]
import matplotlib # pyright: ignore[reportMissingModuleSource]
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg # pyright: ignore[reportMissingModuleSource]
import datetime
import socket
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import ntplib # pyright: ignore[reportMissingImports]
from socket import timeout
import pandas as pd # pyright: ignore[reportMissingModuleSource]
from openpyxl import Workbook # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]
import os
from typing import Dict, Tuple, List
hostname = socket.gethostname()    
IPAddr = socket.gethostbyname(hostname)  
g=str(IPAddr)
w=str("192.168.1.163")
if(g!=w): # pyright: ignore[reportUndefinedVariable]
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")

# License configuration
LICENSE_EXPIRATION = datetime.datetime(2026, 8, 19)
NTP_SERVERS = ['pool.ntp.org', 'time.nist.gov', 'time.google.com']

class LicenseManager:
    """Manages software licensing"""
    
    @staticmethod
    def get_network_time() -> datetime.datetime:
        """Get current time from NTP server"""
        for server in NTP_SERVERS:
            try:
                client = ntplib.NTPClient()
                response = client.request(server, timeout=5)
                return datetime.datetime.fromtimestamp(response.tx_time)
            except (ntplib.NTPException, timeout, Exception):
                continue
        return None
    
    @staticmethod
    def check_license() -> bool:
        """Check if the software license is still valid"""
        try:
            network_time = LicenseManager.get_network_time()
            
            if network_time:
                current_date = network_time
            else:
                current_date = datetime.datetime.now()
                messagebox.showwarning(
                    "License Warning", 
                    "Could not verify time with NTP server. Using system time."
                )
            
            if current_date > LICENSE_EXPIRATION:
                messagebox.showerror(
                    "License Expired", 
                    f"This software license expired on {LICENSE_EXPIRATION.strftime('%B %d, %Y')}.\n"
                    "Please contact the developer for license renewal."
                )
                return False
            
            days_remaining = (LICENSE_EXPIRATION - current_date).days
            if days_remaining <= 7:
                messagebox.showwarning(
                    "License Warning", 
                    f"Software license expires in {days_remaining} days on "
                    f"{LICENSE_EXPIRATION.strftime('%B %d, %Y')}."
                )
            
            return True
            
        except Exception as e:
            messagebox.showerror("License Error", f"License verification failed: {str(e)}")
            return False


class PropertyAnalyzer:
    """Handles all property analysis calculations"""
    
    @staticmethod
    def calculate_rental_income_schedule(monthly_rent: float, years: int, 
                                        annual_increase: float, vacancy_rate: float) -> Tuple[float, List[Dict]]:
        """Calculate detailed rental income schedule with year-by-year breakdown"""
        schedule = []
        total_income = 0
        current_monthly_rent = monthly_rent
        
        for year in range(1, int(years) + 1):
            gross_annual = current_monthly_rent * 12
            vacancy_loss = gross_annual * (vacancy_rate / 100)
            net_annual = gross_annual - vacancy_loss
            total_income += net_annual
            
            schedule.append({
                'year': year,
                'monthly_rent': current_monthly_rent,
                'gross_annual': gross_annual,
                'vacancy_loss': vacancy_loss,
                'net_annual': net_annual,
                'cumulative': total_income
            })
            
            current_monthly_rent *= (1 + annual_increase / 100)
        
        return total_income, schedule
    
    @staticmethod
    def calculate_expense_schedule(annual_tax: float, annual_maintenance: float, 
                                  annual_management: float, monthly_insurance: float, 
                                  years: int) -> Tuple[float, List[Dict]]:
        """Calculate detailed expense schedule"""
        schedule = []
        annual_insurance = monthly_insurance * 12
        annual_total = annual_tax + annual_maintenance + annual_management + annual_insurance
        
        for year in range(1, int(years) + 1):
            schedule.append({
                'year': year,
                'property_tax': annual_tax,
                'maintenance': annual_maintenance,
                'management': annual_management,
                'insurance': annual_insurance,
                'total_expenses': annual_total,
                'cumulative': annual_total * year
            })
        
        return annual_total * years, schedule
    
    @staticmethod
    def calculate_selling_scenario(selling_price: float, closing_costs: float, 
                                  commission_rate: float, purchase_price: float, 
                                  capital_gains_rate: float) -> Dict:
        """Calculate net proceeds from selling property now"""
        commission = selling_price * (commission_rate / 100)
        total_selling_costs = closing_costs + commission
        gross_proceeds = selling_price - total_selling_costs
        
        capital_gains = max(0, selling_price - purchase_price)
        capital_gains_tax = capital_gains * (capital_gains_rate / 100)
        
        net_proceeds = gross_proceeds - capital_gains_tax
        
        return {
            'selling_price': selling_price,
            'commission': commission,
            'closing_costs': closing_costs,
            'total_costs': total_selling_costs,
            'capital_gains': capital_gains,
            'capital_gains_tax': capital_gains_tax,
            'gross_proceeds': gross_proceeds,
            'net_proceeds': net_proceeds
        }
    
    @staticmethod
    def calculate_rental_scenario(params: Dict) -> Dict:
        """Calculate complete rental + future sale scenario"""
        # Unpack parameters
        monthly_rent = params['monthly_rent']
        years = params['years']
        annual_rent_increase = params['annual_rent_increase']
        vacancy_rate = params['vacancy_rate']
        annual_tax = params['annual_tax']
        annual_maintenance = params['annual_maintenance']
        annual_management = params['annual_management']
        monthly_insurance = params['monthly_insurance']
        income_tax_rate = params['income_tax_rate']
        depreciation_deduction = params['depreciation_deduction']
        current_value = params['current_value']
        appreciation_rate = params['appreciation_rate']
        commission_rate = params['commission_rate']
        closing_costs = params['closing_costs']
        capital_gains_rate = params['capital_gains_rate']
        
        # Calculate rental income
        total_rental_income, income_schedule = PropertyAnalyzer.calculate_rental_income_schedule(
            monthly_rent, years, annual_rent_increase, vacancy_rate
        )
        
        # Calculate expenses
        total_expenses, expense_schedule = PropertyAnalyzer.calculate_expense_schedule(
            annual_tax, annual_maintenance, annual_management, monthly_insurance, years
        )
        
        # Calculate depreciation
        total_depreciation = depreciation_deduction * years
        depreciation_tax_benefit = total_depreciation * (income_tax_rate / 100)
        
        # Calculate net rental income
        net_rental_income = total_rental_income - total_expenses
        rental_income_tax = max(0, net_rental_income * (income_tax_rate / 100))
        net_rental_after_tax = net_rental_income - rental_income_tax + depreciation_tax_benefit
        
        # Calculate future property value
        future_value = current_value * ((1 + appreciation_rate / 100) ** years)
        property_appreciation = future_value - current_value
        
        # Calculate future sale
        future_commission = future_value * (commission_rate / 100)
        future_selling_costs = future_commission + closing_costs
        future_capital_gains = max(0, future_value - current_value)
        future_capital_gains_tax = future_capital_gains * (capital_gains_rate / 100)
        future_net_proceeds = future_value - future_selling_costs - future_capital_gains_tax
        
        # Total value
        total_value = net_rental_after_tax + future_net_proceeds
        
        return {
            'total_rental_income': total_rental_income,
            'income_schedule': income_schedule,
            'total_expenses': total_expenses,
            'expense_schedule': expense_schedule,
            'net_rental_income': net_rental_income,
            'rental_income_tax': rental_income_tax,
            'total_depreciation': total_depreciation,
            'depreciation_tax_benefit': depreciation_tax_benefit,
            'net_rental_after_tax': net_rental_after_tax,
            'future_value': future_value,
            'property_appreciation': property_appreciation,
            'future_commission': future_commission,
            'future_selling_costs': future_selling_costs,
            'future_capital_gains': future_capital_gains,
            'future_capital_gains_tax': future_capital_gains_tax,
            'future_net_proceeds': future_net_proceeds,
            'total_value': total_value,
            'avg_annual_cash_flow': net_rental_after_tax / years,
            'cash_on_cash_return': (net_rental_after_tax / years / current_value) * 100
        }

class ExcelExporter:
    """Handles Excel export with professional formatting"""
    
    @staticmethod
    def create_styled_workbook(input_data: Dict, sell_result: Dict, 
                              rent_result: Dict, currency: str) -> str:
        """Create professionally formatted Excel workbook"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        ExcelExporter._create_summary_sheet(wb, input_data, sell_result, rent_result, currency)
        ExcelExporter._create_input_sheet(wb, input_data, currency)
        ExcelExporter._create_rental_income_sheet(wb, rent_result, currency)
        ExcelExporter._create_expense_sheet(wb, rent_result, currency)
        ExcelExporter._create_comparison_sheet(wb, sell_result, rent_result, currency)
        
        # Save file
        filename = f"property_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
    
    @staticmethod
    def _apply_header_style(cell):
        """Apply consistent header styling"""
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def _create_summary_sheet(wb, input_data, sell_result, rent_result, currency):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "PROPERTY ANALYSIS - EXECUTIVE SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        # Analysis info
        ws['A3'] = "Project:"
        ws['B3'] = input_data.get('Project Name', 'N/A')
        ws['A4'] = "Analysis Date:"
        ws['B4'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "Currency:"
        ws['B5'] = currency
        ws['A6'] = "Analysis Period:"
        ws['B6'] = f"{input_data['Analysis Period (Years)']} years"
        
        # Key Results
        ws['A8'] = "FINANCIAL COMPARISON"
        ws['A8'].font = Font(bold=True, size=12)
        
        headers = ['Scenario', f'Net Value ({currency})', 'Difference', 'ROI']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            ExcelExporter._apply_header_style(cell)
        
        # Sell scenario
        ws['A10'] = "Sell Property Now"
        ws['B10'] = sell_result['net_proceeds']
        ws['B10'].number_format = '#,##0.00'
        ws['C10'] = "-"
        ws['D10'] = "-"
        
        # Rent scenario
        ws['A11'] = "Rent & Sell Later"
        ws['B11'] = rent_result['total_value']
        ws['B11'].number_format = '#,##0.00'
        
        difference = rent_result['total_value'] - sell_result['net_proceeds']
        ws['C11'] = difference
        ws['C11'].number_format = '#,##0.00'
        
        roi = (difference / input_data['Current Property Value']) * 100
        ws['D11'] = f"{roi:.2f}%"
        
        # Recommendation
        ws['A13'] = "RECOMMENDATION:"
        ws['A13'].font = Font(bold=True, size=12)
        
        if rent_result['total_value'] > sell_result['net_proceeds']:
            recommendation = "RENT THE PROPERTY"
            ws['B13'] = recommendation
            ws['B13'].font = Font(bold=True, color="008000", size=12)
        else:
            recommendation = "SELL THE PROPERTY"
            ws['B13'] = recommendation
            ws['B13'].font = Font(bold=True, color="FF0000", size=12)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @staticmethod
    def _create_input_sheet(wb, input_data, currency):
        """Create input parameters sheet"""
        ws = wb.create_sheet("Input Parameters")
        
        ws['A1'] = "INPUT PARAMETERS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        for key, value in input_data.items():
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)) and key != 'Analysis Period (Years)':
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def _create_rental_income_sheet(wb, rent_result, currency):
        """Create detailed rental income breakdown"""
        ws = wb.create_sheet("Rental Income Detail")
        
        ws['A1'] = "RENTAL INCOME SCHEDULE"
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ['Year', 'Monthly Rent', 'Gross Annual', 'Vacancy Loss', 'Net Annual', 'Cumulative']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            ExcelExporter._apply_header_style(cell)
        
        row = 4
        for item in rent_result['income_schedule']:
            ws.cell(row=row, column=1, value=item['year'])
            ws.cell(row=row, column=2, value=item['monthly_rent']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=item['gross_annual']).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=item['vacancy_loss']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=item['net_annual']).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=item['cumulative']).number_format = '#,##0.00'
            row += 1
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _create_expense_sheet(wb, rent_result, currency):
        """Create detailed expense breakdown"""
        ws = wb.create_sheet("Operating Expenses")
        
        ws['A1'] = "OPERATING EXPENSE SCHEDULE"
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ['Year', 'Property Tax', 'Maintenance', 'Management', 'Insurance', 'Total', 'Cumulative']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            ExcelExporter._apply_header_style(cell)
        
        row = 4
        for item in rent_result['expense_schedule']:
            ws.cell(row=row, column=1, value=item['year'])
            ws.cell(row=row, column=2, value=item['property_tax']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=item['maintenance']).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=item['management']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=item['insurance']).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=item['total_expenses']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=item['cumulative']).number_format = '#,##0.00'
            row += 1
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _create_comparison_sheet(wb, sell_result, rent_result, currency):
        """Create detailed comparison sheet"""
        ws = wb.create_sheet("Detailed Comparison")
        
        ws['A1'] = "DETAILED FINANCIAL COMPARISON"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Sell Now Section
        ws['A3'] = "SELL PROPERTY NOW"
        ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        sell_items = [
            ('Selling Price', sell_result['selling_price']),
            ('Realtor Commission', -sell_result['commission']),
            ('Closing Costs', -sell_result['closing_costs']),
            ('Capital Gains Tax', -sell_result['capital_gains_tax']),
            ('NET PROCEEDS', sell_result['net_proceeds'])
        ]
        
        row = 4
        for label, value in sell_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0.00'
            if label.startswith('NET'):
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Rent & Sell Later Section
        row += 1
        ws[f'A{row}'] = "RENT & SELL LATER"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row += 1
        rent_items = [
            ('Total Rental Income', rent_result['total_rental_income']),
            ('Operating Expenses', -rent_result['total_expenses']),
            ('Income Tax on Rental', -rent_result['rental_income_tax']),
            ('Depreciation Tax Benefit', rent_result['depreciation_tax_benefit']),
            ('Net Rental Cash Flow', rent_result['net_rental_after_tax']),
            ('', 0),
            ('Future Property Value', rent_result['future_value']),
            ('Future Selling Costs', -rent_result['future_selling_costs']),
            ('Future Capital Gains Tax', -rent_result['future_capital_gains_tax']),
            ('Future Net Proceeds', rent_result['future_net_proceeds']),
            ('', 0),
            ('TOTAL VALUE', rent_result['total_value'])
        ]
        
        for label, value in rent_items:
            if label:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value if value != 0 else ''
                if value != 0:
                    ws[f'B{row}'].number_format = '#,##0.00'
                if label.startswith('TOTAL') or label.startswith('Net Rental'):
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

class PropertyCalculatorGUI:
    """Main GUI application"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sell vs Rent Property Calculator - Professional Edition")
        self.root.geometry("900x800")
        
        # Check license
        if not LicenseManager.check_license():
            sys.exit()
        
        self.setup_gui()
        
    def setup_gui(self):
        """Setup the GUI components"""
        # License info
        license_frame = tk.Frame(self.root, bg="#366092", height=30)
        license_frame.pack(fill="x")
        
        license_label = tk.Label(
            license_frame,
            text=f"Licensed until: {LICENSE_EXPIRATION.strftime('%B %d, %Y')} | "
                 f"Created by Engineer Wael Sherif Selim | wael.sherif.selim@gmail.com",
            font=("Georgia", 9),
            fg="white",
            bg="#366092"
        )
        license_label.pack(pady=5)
        
        # Main container with notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Input tab
        self.input_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.input_frame, text="Input Parameters")
        self.setup_input_tab()
        
        # Results tab (initially disabled)
        self.results_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.results_frame, text="Results & Charts")
        
        # Button frame
        button_frame = tk.Frame(self.root, bg="#f0f0f0")
        button_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            button_frame,
            text="Calculate Analysis",
            font=("Georgia", 12, "bold"),
            command=self.run_calculation,
            bg="#4CAF50",
            fg="white",
            padx=30,
            pady=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            button_frame,
            text="Clear All",
            font=("Georgia", 12),
            command=self.clear_inputs,
            bg="#FF9800",
            fg="white",
            padx=30,
            pady=10
        ).pack(side="left", padx=5)
        
        tk.Button(
            button_frame,
            text="Exit",
            font=("Georgia", 12),
            command=self.root.destroy,
            bg="#f44336",
            fg="white",
            padx=30,
            pady=10
        ).pack(side="right", padx=5)
    
    def setup_input_tab(self):
        """Setup input fields"""
        # Create scrollable canvas
        canvas = tk.Canvas(self.input_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.input_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Field definitions with tooltips
        field_groups = {
            "Project Information": [
                ('Project Name', 'Property Analysis', 'Name of your analysis project'),
                ('Currency (USD, EUR, etc.)', 'USD', 'Currency for all amounts')
            ],
            "Property Values": [
                ('Current Property Value', '500000', 'Current market value of property'),
                ('Expected Selling Price', '520000', 'Price if selling now'),
                ('Selling Closing Costs', '5000', 'Fixed costs to sell (legal, etc.)'),
                ('Realtor Commission (%)', '6.0', 'Real estate agent commission %')
            ],
            "Taxes": [
                ('Capital Gains Tax Rate (%)', '15.0', 'Tax on property appreciation'),
                ('Income Tax Rate (%)', '25.0', 'Tax rate on rental income'),
                ('Annual Property Tax', '6000', 'Property tax per year')
            ],
            "Rental Income": [
                ('Monthly Rental Income', '3000', 'Expected monthly rent'),
                ('Annual Rent Increase (%)', '3.0', 'Annual rent growth %'),
                ('Vacancy Rate (%)', '5.0', 'Expected vacancy %')
            ],
            "Operating Expenses": [
                ('Monthly Property Insurance', '150', 'Insurance cost per month'),
                ('Annual Maintenance Cost', '3000', 'Repairs & maintenance per year'),
                ('Annual Management Cost', '1800', 'Property management fees per year')
            ],
            "Analysis Parameters": [
                ('Number of Years Analysis', '10', 'Analysis time horizon (1-30 years)'),
                ('Property Appreciation Rate (%)', '3.0', 'Annual property value growth %'),
                ('Depreciation Deduction', '10000', 'Annual depreciation for tax purposes')
            ]
        }
        
        self.entries = {}
        row = 0
        
        for group_name, fields in field_groups.items():
            # Group header
            group_label = tk.Label(
                scrollable_frame,
                text=group_name,
                font=("Georgia", 11, "bold"),
                bg="#366092",
                fg="white",
                padx=5,
                pady=5
            )
            group_label.grid(row=row, column=0, columnspan=3, sticky="ew", padx=5, pady=(10, 2))
            row += 1
            
            # Fields in group
            for field_name, default_value, tooltip in fields:
                label = tk.Label(
                    scrollable_frame,
                    text=field_name + ":",
                    font=("Georgia", 10),
                    anchor='w'
                )
                label.grid(row=row, column=0, sticky="w", padx=10, pady=3)
                
                entry = tk.Entry(scrollable_frame, width=25, font=("Georgia", 10))
                entry.insert(0, default_value)
                entry.grid(row=row, column=1, sticky="w", padx=5, pady=3)
                
                # Tooltip label
                tooltip_label = tk.Label(
                    scrollable_frame,
                    text="ⓘ",
                    font=("Georgia", 10),
                    fg="blue",
                    cursor="question_arrow"
                )
                tooltip_label.grid(row=row, column=2, sticky="w", padx=2, pady=3)
                self.create_tooltip(tooltip_label, tooltip)
                
                self.entries[field_name] = entry
                row += 1
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_tooltip(self, widget, text):
        """Create tooltip for widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(
                tooltip,
                text=text,
                background="lightyellow",
                relief="solid",
                borderwidth=1,
                font=("Georgia", 9)
            )
            label.pack()
            widget.tooltip = tooltip
        
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
        
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
    
    def clear_inputs(self):
        """Clear all input fields"""
        if messagebox.askyesno("Confirm", "Clear all input fields?"):
            for entry in self.entries.values():
                entry.delete(0, tk.END)
    
    def run_calculation(self):
        """Execute the property analysis calculation"""
        try:
            # Gather inputs
            input_data = self.gather_inputs()
            
            # Validate inputs
            self.validate_inputs(input_data)
            
            # Calculate sell scenario
            sell_result = PropertyAnalyzer.calculate_selling_scenario(
                input_data['Expected Selling Price'],
                input_data['Selling Closing Costs'],
                input_data['Realtor Commission (%)'],
                input_data['Current Property Value'],
                input_data['Capital Gains Tax Rate (%)']
            )
            
            # Calculate rental scenario
            rental_params = {
                'monthly_rent': input_data['Monthly Rental Income'],
                'years': input_data['Analysis Period (Years)'],
                'annual_rent_increase': input_data['Annual Rent Increase (%)'],
                'vacancy_rate': input_data['Vacancy Rate (%)'],
                'annual_tax': input_data['Annual Property Tax'],
                'annual_maintenance': input_data['Annual Maintenance Cost'],
                'annual_management': input_data['Annual Management Cost'],
                'monthly_insurance': input_data['Monthly Property Insurance'],
                'income_tax_rate': input_data['Income Tax Rate (%)'],
                'depreciation_deduction': input_data['Depreciation Deduction'],
                'current_value': input_data['Current Property Value'],
                'appreciation_rate': input_data['Property Appreciation Rate (%)'],
                'commission_rate': input_data['Realtor Commission (%)'],
                'closing_costs': input_data['Selling Closing Costs'],
                'capital_gains_rate': input_data['Capital Gains Tax Rate (%)']
            }
            
            rent_result = PropertyAnalyzer.calculate_rental_scenario(rental_params)
            
            # Generate outputs
            self.generate_text_report(input_data, sell_result, rent_result)
            self.display_results(input_data, sell_result, rent_result)
            
            # Export to Excel
            excel_file = ExcelExporter.create_styled_workbook(
                input_data, sell_result, rent_result, input_data['Currency']
            )
            
            # Show success message
            difference = abs(rent_result['total_value'] - sell_result['net_proceeds'])
            if rent_result['total_value'] > sell_result['net_proceeds']:
                recommendation = "RENT THE PROPERTY"
                message = f"Renting generates {difference:,.2f} {input_data['Currency']} more value"
            else:
                recommendation = "SELL THE PROPERTY"
                message = f"Selling generates {difference:,.2f} {input_data['Currency']} more value"
            
            messagebox.showinfo(
                "Analysis Complete",
                f"RECOMMENDATION: {recommendation}\n\n"
                f"{message}\n\n"
                f"Reports saved:\n"
                f"- sell_vs_rent_analysis.txt\n"
                f"- {excel_file}"
            )
            
            # Switch to results tab
            self.notebook.select(1)
            
        except ValueError as e:
            messagebox.showerror("Input Error", f"Invalid input: {str(e)}")
        except Exception as e:
            messagebox.showerror("Calculation Error", f"Error during calculation: {str(e)}")
    
    def gather_inputs(self) -> Dict:
        """Gather and convert all inputs"""
        data = {}
        
        # String fields
        data['Project Name'] = self.entries['Project Name'].get().strip()
        data['Currency'] = self.entries['Currency (USD, EUR, etc.)'].get().strip().upper()
        
        # Numeric fields
        numeric_fields = {
            'Current Property Value': float,
            'Expected Selling Price': float,
            'Selling Closing Costs': float,
            'Realtor Commission (%)': float,
            'Capital Gains Tax Rate (%)': float,
            'Monthly Rental Income': float,
            'Monthly Property Insurance': float,
            'Number of Years Analysis': int,
            'Annual Rent Increase (%)': float,
            'Annual Property Tax': float,
            'Annual Maintenance Cost': float,
            'Annual Management Cost': float,
            'Vacancy Rate (%)': float,
            'Property Appreciation Rate (%)': float,
            'Income Tax Rate (%)': float,
            'Depreciation Deduction': float
        }
        
        for field, convert_func in numeric_fields.items():
            value_str = self.entries[field].get().strip()
            if not value_str:
                raise ValueError(f"{field} cannot be empty")
            data[field] = convert_func(value_str)
        
        # Rename for consistency
        data['Analysis Period (Years)'] = data.pop('Number of Years Analysis')
        
        return data
    
    def validate_inputs(self, data: Dict):
        """Validate input data"""
        if not data['Project Name']:
            raise ValueError("Project Name is required")
        
        if data['Analysis Period (Years)'] < 1 or data['Analysis Period (Years)'] > 30:
            raise ValueError("Analysis period must be between 1 and 30 years")
        
        if data['Current Property Value'] <= 0:
            raise ValueError("Current Property Value must be positive")
        
        if data['Expected Selling Price'] <= 0:
            raise ValueError("Expected Selling Price must be positive")
        
        if data['Monthly Rental Income'] < 0:
            raise ValueError("Monthly Rental Income cannot be negative")
        
        if data['Vacancy Rate (%)'] < 0 or data['Vacancy Rate (%)'] > 100:
            raise ValueError("Vacancy Rate must be between 0 and 100")
        
        if data['Realtor Commission (%)'] < 0 or data['Realtor Commission (%)'] > 100:
            raise ValueError("Realtor Commission must be between 0 and 100")
    
    def generate_text_report(self, input_data: Dict, sell_result: Dict, rent_result: Dict):
        """Generate comprehensive text report"""
        currency = input_data['Currency']
        
        with open("sell_vs_rent_analysis.txt", "w", encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("COMPREHENSIVE PROPERTY ANALYSIS: SELL vs RENT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Created by: Engineer Wael Sherif Selim\n")
            f.write(f"Email: wael.sherif.selim@gmail.com\n")
            f.write(f"Analysis Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Software License Valid Until: {LICENSE_EXPIRATION.strftime('%Y-%m-%d')}\n")
            f.write("=" * 80 + "\n\n")
            
            # Input Summary
            f.write("INPUT PARAMETERS\n")
            f.write("-" * 80 + "\n")
            f.write(f"Project Name: {input_data['Project Name']}\n")
            f.write(f"Currency: {currency}\n")
            f.write(f"Analysis Period: {input_data['Analysis Period (Years)']} years\n")
            f.write(f"Current Property Value: {input_data['Current Property Value']:,.2f} {currency}\n")
            f.write(f"Expected Selling Price: {input_data['Expected Selling Price']:,.2f} {currency}\n")
            f.write(f"Monthly Rental Income: {input_data['Monthly Rental Income']:,.2f} {currency}\n")
            f.write(f"Annual Rent Increase: {input_data['Annual Rent Increase (%)']:.2f}%\n")
            f.write(f"Vacancy Rate: {input_data['Vacancy Rate (%)']:.2f}%\n")
            f.write(f"Property Appreciation Rate: {input_data['Property Appreciation Rate (%)']:.2f}%\n")
            f.write(f"Income Tax Rate: {input_data['Income Tax Rate (%)']:.2f}%\n")
            f.write("\n" + "=" * 80 + "\n\n")
            
            # Scenario 1: Sell Now
            f.write("SCENARIO 1: SELL PROPERTY NOW\n")
            f.write("-" * 80 + "\n")
            f.write(f"Gross Selling Price: {sell_result['selling_price']:,.2f} {currency}\n")
            f.write(f"Less: Realtor Commission: ({sell_result['commission']:,.2f}) {currency}\n")
            f.write(f"Less: Closing Costs: ({sell_result['closing_costs']:,.2f}) {currency}\n")
            f.write(f"Less: Capital Gains Tax: ({sell_result['capital_gains_tax']:,.2f}) {currency}\n")
            f.write(f"{'':->50}\n")
            f.write(f"NET PROCEEDS FROM SALE: {sell_result['net_proceeds']:,.2f} {currency}\n")
            f.write("\n" + "=" * 80 + "\n\n")
            
            # Scenario 2: Rent and Sell Later
            f.write(f"SCENARIO 2: RENT FOR {input_data['Analysis Period (Years)']} YEARS, THEN SELL\n")
            f.write("-" * 80 + "\n")
            f.write("\nRENTAL INCOME ANALYSIS:\n")
            f.write(f"Total Rental Income (before vacancy): {rent_result['total_rental_income']:,.2f} {currency}\n")
            f.write(f"Less: Vacancy Loss: Included in net calculation\n")
            f.write(f"Net Rental Income: {rent_result['total_rental_income']:,.2f} {currency}\n")
            
            f.write("\nOPERATING EXPENSES:\n")
            f.write(f"Total Operating Expenses: ({rent_result['total_expenses']:,.2f}) {currency}\n")
            f.write(f"  - Property Tax: {input_data['Annual Property Tax'] * input_data['Analysis Period (Years)']:,.2f} {currency}\n")
            f.write(f"  - Maintenance: {input_data['Annual Maintenance Cost'] * input_data['Analysis Period (Years)']:,.2f} {currency}\n")
            f.write(f"  - Management: {input_data['Annual Management Cost'] * input_data['Analysis Period (Years)']:,.2f} {currency}\n")
            f.write(f"  - Insurance: {input_data['Monthly Property Insurance'] * 12 * input_data['Analysis Period (Years)']:,.2f} {currency}\n")
            
            f.write("\nTAX CONSIDERATIONS:\n")
            f.write(f"Income Tax on Rental Income: ({rent_result['rental_income_tax']:,.2f}) {currency}\n")
            f.write(f"Depreciation Tax Benefit: {rent_result['depreciation_tax_benefit']:,.2f} {currency}\n")
            f.write(f"{'':->50}\n")
            f.write(f"NET RENTAL CASH FLOW (after tax): {rent_result['net_rental_after_tax']:,.2f} {currency}\n")
            
            f.write(f"\nPROPERTY VALUE PROJECTION:\n")
            f.write(f"Current Property Value: {input_data['Current Property Value']:,.2f} {currency}\n")
            f.write(f"Future Property Value (Year {input_data['Analysis Period (Years)']}): {rent_result['future_value']:,.2f} {currency}\n")
            f.write(f"Property Appreciation: {rent_result['property_appreciation']:,.2f} {currency}\n")
            
            f.write(f"\nFUTURE SALE ANALYSIS:\n")
            f.write(f"Future Property Value: {rent_result['future_value']:,.2f} {currency}\n")
            f.write(f"Less: Future Selling Costs: ({rent_result['future_selling_costs']:,.2f}) {currency}\n")
            f.write(f"Less: Future Capital Gains Tax: ({rent_result['future_capital_gains_tax']:,.2f}) {currency}\n")
            f.write(f"{'':->50}\n")
            f.write(f"Future Net Proceeds: {rent_result['future_net_proceeds']:,.2f} {currency}\n")
            
            f.write(f"\nTOTAL VALUE FROM RENTING:\n")
            f.write(f"Net Rental Cash Flow: {rent_result['net_rental_after_tax']:,.2f} {currency}\n")
            f.write(f"Plus: Future Net Proceeds: {rent_result['future_net_proceeds']:,.2f} {currency}\n")
            f.write(f"{'':->50}\n")
            f.write(f"TOTAL VALUE: {rent_result['total_value']:,.2f} {currency}\n")
            
            f.write("\n" + "=" * 80 + "\n\n")
            
            # Comparison & Recommendation
            f.write("FINANCIAL COMPARISON & RECOMMENDATION\n")
            f.write("-" * 80 + "\n")
            f.write(f"Option 1 - Sell Now: {sell_result['net_proceeds']:,.2f} {currency}\n")
            f.write(f"Option 2 - Rent & Sell Later: {rent_result['total_value']:,.2f} {currency}\n")
            f.write(f"{'':->50}\n")
            
            difference = rent_result['total_value'] - sell_result['net_proceeds']
            
            if difference > 0:
                recommendation = "RENT THE PROPERTY"
                f.write(f"RECOMMENDATION: {recommendation}\n")
                f.write(f"Additional Value from Renting: {difference:,.2f} {currency}\n")
                f.write(f"Return on Investment: {(difference / input_data['Current Property Value'] * 100):.2f}%\n")
            else:
                recommendation = "SELL THE PROPERTY"
                f.write(f"RECOMMENDATION: {recommendation}\n")
                f.write(f"Additional Value from Selling: {abs(difference):,.2f} {currency}\n")
                f.write(f"Immediate Liquidity: {sell_result['net_proceeds']:,.2f} {currency}\n")
            
            # Additional Metrics
            f.write("\n" + "=" * 80 + "\n\n")
            f.write("ADDITIONAL PERFORMANCE METRICS\n")
            f.write("-" * 80 + "\n")
            f.write(f"Average Annual Cash Flow: {rent_result['avg_annual_cash_flow']:,.2f} {currency}\n")
            f.write(f"Cash-on-Cash Return: {rent_result['cash_on_cash_return']:.2f}%\n")
            f.write(f"Total Return (if renting): {(rent_result['total_value'] / input_data['Current Property Value'] - 1) * 100:.2f}%\n")
            f.write(f"Annualized Return (if renting): {((rent_result['total_value'] / input_data['Current Property Value']) ** (1/input_data['Analysis Period (Years)']) - 1) * 100:.2f}%\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("END OF REPORT\n")
            f.write("=" * 80 + "\n")
    
    def display_results(self, input_data: Dict, sell_result: Dict, rent_result: Dict):
        """Display results with charts in the results tab"""
        # Clear previous results
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        currency = input_data['Currency']
        
        # Create canvas for scrolling
        canvas = tk.Canvas(self.results_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.results_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Summary box
        summary_frame = tk.LabelFrame(
            scrollable_frame,
            text="Executive Summary",
            font=("Georgia", 12, "bold"),
            padx=10,
            pady=10
        )
        summary_frame.pack(fill="x", padx=10, pady=10)
        
        difference = rent_result['total_value'] - sell_result['net_proceeds']
        if difference > 0:
            recommendation = "RENT THE PROPERTY"
            color = "green"
        else:
            recommendation = "SELL THE PROPERTY"
            color = "red"
        
        tk.Label(
            summary_frame,
            text=f"RECOMMENDATION: {recommendation}",
            font=("Georgia", 14, "bold"),
            fg=color
        ).pack()
        
        tk.Label(
            summary_frame,
            text=f"Value Difference: {abs(difference):,.2f} {currency}",
            font=("Georgia", 12)
        ).pack()
        
        # Comparison table
        table_frame = tk.LabelFrame(
            scrollable_frame,
            text="Financial Comparison",
            font=("Georgia", 12, "bold"),
            padx=10,
            pady=10
        )
        table_frame.pack(fill="x", padx=10, pady=10)
        
        headers = ["Scenario", f"Net Value ({currency})", "ROI"]
        for col, header in enumerate(headers):
            tk.Label(
                table_frame,
                text=header,
                font=("Georgia", 10, "bold"),
                bg="#366092",
                fg="white",
                padx=10,
                pady=5
            ).grid(row=0, column=col, sticky="ew", padx=1, pady=1)
        
        scenarios = [
            ("Sell Now", f"{sell_result['net_proceeds']:,.2f}", "-"),
            ("Rent & Sell Later", f"{rent_result['total_value']:,.2f}", 
             f"{(rent_result['total_value'] / input_data['Current Property Value'] - 1) * 100:.2f}%")
        ]
        
        for row, (scenario, value, roi) in enumerate(scenarios, 1):
            tk.Label(table_frame, text=scenario, font=("Georgia", 10), padx=10, pady=5).grid(row=row, column=0, sticky="w")
            tk.Label(table_frame, text=value, font=("Georgia", 10), padx=10, pady=5).grid(row=row, column=1, sticky="e")
            tk.Label(table_frame, text=roi, font=("Georgia", 10), padx=10, pady=5).grid(row=row, column=2, sticky="e")
        
        # Charts
        self.create_charts(scrollable_frame, input_data, sell_result, rent_result)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_charts(self, parent, input_data, sell_result, rent_result):
        """Create visualization charts"""
        chart_frame = tk.LabelFrame(
            parent,
            text="Visual Analysis",
            font=("Georgia", 12, "bold"),
            padx=10,
            pady=10
        )
        chart_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create matplotlib figure with subplots
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 10))
        fig.suptitle('Property Analysis Charts', fontsize=16, fontweight='bold')
        
        currency = input_data['Currency']
        
        # Chart 1: Comparison Bar Chart
        scenarios = ['Sell Now', 'Rent & Sell Later']
        values = [sell_result['net_proceeds'], rent_result['total_value']]
        colors = ['#ff6b6b', '#51cf66']
        
        ax1.bar(scenarios, values, color=colors, alpha=0.7, edgecolor='black')
        ax1.set_ylabel(f'Net Value ({currency})', fontweight='bold')
        ax1.set_title('Scenario Comparison', fontweight='bold')
        ax1.grid(axis='y', alpha=0.3)
        
        for i, v in enumerate(values):
            ax1.text(i, v, f'{v:,.0f}', ha='center', va='bottom', fontweight='bold')
        
        # Chart 2: Rental Income vs Expenses
        years = range(1, input_data['Analysis Period (Years)'] + 1)
        cumulative_income = [item['cumulative'] for item in rent_result['income_schedule']]
        cumulative_expenses = [item['cumulative'] for item in rent_result['expense_schedule']]
        
        ax2.plot(years, cumulative_income, marker='o', label='Income', linewidth=2, color='green')
        ax2.plot(years, cumulative_expenses, marker='s', label='Expenses', linewidth=2, color='red')
        ax2.fill_between(years, cumulative_income, cumulative_expenses, alpha=0.3, color='green')
        ax2.set_xlabel('Year', fontweight='bold')
        ax2.set_ylabel(f'Cumulative Amount ({currency})', fontweight='bold')
        ax2.set_title('Cumulative Income vs Expenses', fontweight='bold')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        # Chart 3: Property Value Projection
        property_values = [input_data['Current Property Value'] * 
                          ((1 + input_data['Property Appreciation Rate (%)'] / 100) ** year)
                          for year in range(input_data['Analysis Period (Years)'] + 1)]
        
        ax3.plot(range(input_data['Analysis Period (Years)'] + 1), property_values, 
                marker='o', linewidth=2, color='blue')
        ax3.set_xlabel('Year', fontweight='bold')
        ax3.set_ylabel(f'Property Value ({currency})', fontweight='bold')
        ax3.set_title('Property Value Appreciation', fontweight='bold')
        ax3.grid(True, alpha=0.3)
        
        # Chart 4: Cash Flow Breakdown (Pie Chart)
        if rent_result['total_value'] > sell_result['net_proceeds']:
            labels = ['Rental Income', 'Operating Expenses', 'Taxes', 'Future Sale Proceeds']
            sizes = [
                rent_result['total_rental_income'],
                rent_result['total_expenses'],
                rent_result['rental_income_tax'] + rent_result['future_capital_gains_tax'],
                rent_result['future_net_proceeds']
            ]
            colors_pie = ['#51cf66', '#ff6b6b', '#ffd43b', '#4dabf7']
            
            ax4.pie(sizes, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90)
            ax4.set_title('Rental Scenario Breakdown', fontweight='bold')
        else:
            labels = ['Net Proceeds', 'Selling Costs', 'Capital Gains Tax']
            sizes = [
                sell_result['net_proceeds'],
                sell_result['total_costs'],
                sell_result['capital_gains_tax']
            ]
            colors_pie = ['#51cf66', '#ff6b6b', '#ffd43b']
            
            ax4.pie(sizes, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90)
            ax4.set_title('Selling Scenario Breakdown', fontweight='bold')
        
        plt.tight_layout()
        
        # Embed in tkinter
        canvas_widget = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas_widget.draw()
        canvas_widget.get_tk_widget().pack(fill="both", expand=True)

def main():
    """Main entry point"""
    root = tk.Tk()
    app = PropertyCalculatorGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()