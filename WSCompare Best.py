import datetime
import socket
import tkinter as tk
from tkinter.ttk import *
from tkinter import messagebox, filedialog
import ntplib # type: ignore
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Hard License System Configuration
# Set expiry date to August 19, 2025
TARGET_EXPIRY_DATE = datetime.datetime(2026, 8, 19, 23, 59, 59)
NTP_SERVERS = [
    "pool.ntp.org",
    "time.google.com",
    "time.cloudflare.com",
    "time.windows.com",
    "time.apple.com"
]

class HardLicenseManager:
    def __init__(self):
        pass

    def _get_ntp_time(self):
        """
        Get accurate time from NTP servers.
        Tries multiple servers and falls back to system time if all fail.
        """
        for server in NTP_SERVERS:
            try:
                client = ntplib.NTPClient()
                # Request NTP response with a timeout
                response = client.request(server, version=3, timeout=5)
                # Convert NTP timestamp to datetime object
                ntp_time = datetime.datetime.fromtimestamp(response.tx_time)
                print(f"NTP time retrieved from {server}: {ntp_time}")
                return ntp_time
            except Exception as e:
                print(f"Failed to get time from {server}: {e}")
                continue

        # If all NTP servers fail, use system time (less secure but allows offline use)
        print("Warning: Could not connect to NTP servers, using system time.")
        return datetime.datetime.now()

    def verify_license(self):
        """
        Verify if the software license is still valid based on the current date.
        Compares current time (preferably NTP) against a hardcoded expiry date.
        """
        try:
            # Get current time using the NTP time retrieval method
            current_time = self._get_ntp_time()

            # Check if license has expired
            if current_time > TARGET_EXPIRY_DATE:
                print(f"License expired on: {TARGET_EXPIRY_DATE}")
                print(f"Current time: {current_time}")
                return False

            # Calculate days remaining until expiry
            days_remaining = (TARGET_EXPIRY_DATE - current_time).days
            print(f"License is valid. Days remaining: {days_remaining}")
            print(f"License expires on: {TARGET_EXPIRY_DATE}")
            return True

        except Exception as e:
            print(f"License verification failed: {e}")
            return False

# Initialize license manager
license_manager = HardLicenseManager()

# Verify license before running the main program
if not license_manager.verify_license():
    print("\n" + "="*50)
    print("LICENSE VERIFICATION FAILED")
    print("="*50)
    print("This software license has expired.")
    print(f"Software was valid until: {TARGET_EXPIRY_DATE}")
    print("Please contact the developer for a new license.")
    print("="*50)
    input("Press Enter to exit...")
    sys.exit()

# Original code continues here...
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
local_ip_address = str(IPAddr)
print(local_ip_address)
current_datetime = datetime.datetime.now()
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
current_date = datetime.datetime.now()
g=str(IPAddr)
w=str('192.168.1.163')
if(g!=w):
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")
# Open the output file in append mode
f = open("ws which property to buy or rent.txt","a")
print("This software is created by Engineer Wael Sherif Selim, email:wael.sherif.selim@gmail.com")
print("This software is created by Engineer Wael Sherif Selim, email:wael.sherif.selim@gmail.com", file=f)
print("Ws which property to buy or rent")
print("Ws which property to buy or rent", file=f)

# Global variable to store last calculation results
last_results = {}

# Define the fields for the property evaluation form with simplified names
fields = (
    'project name',
    # Property 1 fields (30 fields)
    'price',
    'address',
    'area',
    'area units',
    'price per unit',
    'FHA mortgage',
    'taxes',
    'down payment',
    'insurance',
    'mortgage',
    'utilities',
    'other fees',
    'year built',
    'floor level',
    'room count',
    'bedroom count',
    'bathroom count',
    'basement',
    'finish',
    'air conditioning',
    'heating',
    'compound',
    'landscaped',
    'quietness',
    'transportation',
    'garbage service',
    'street lighting',
    'shopping',
    'hospital',
    'school',
    'entertainment',
    'view',

    # Property 2 fields (30 fields)
    'price',
    'address',
    'area',
    'area units',
    'price per unit',
    'FHA mortgage',
    'taxes',
    'down payment',
    'insurance',
    'mortgage',
    'utilities',
    'other fees',
    'year built',
    'floor level',
    'room count',
    'bedroom count',
    'bathroom count',
    'basement',
    'finish',
    'air conditioning',
    'heating',
    'compound',
    'landscaped',
    'quietness',
    'transportation',
    'garbage service',
    'street lighting',
    'shopping',
    'hospital',
    'school',
    'entertainment',
    'view',

    # Property 3 fields (30 fields)
    'price',
    'address',
    'area',
    'area units',
    'price per unit',
    'FHA mortgage',
    'taxes',
    'down payment',
    'insurance',
    'mortgage',
    'utilities',
    'other fees',
    'year built',
    'floor level',
    'room count',
    'bedroom count',
    'bathroom count',
    'basement',
    'finish',
    'air conditioning',
    'heating',
    'compound',
    'landscaped',
    'quietness',
    'transportation',
    'garbage service',
    'street lighting',
    'shopping',
    'hospital',
    'school',
    'entertainment',
    'view'
)

# Mapping from the simplified field names to display names
short_display_names = {
    'project name': 'Project Name',
    'price': 'Price',
    'address': 'Address',
    'area': 'Area',
    'area units': 'Area Units',
    'price per unit': 'Price Per Unit',
    'FHA mortgage': 'FHA Mortgage',
    'taxes': 'Taxes',
    'down payment': 'Down Payment',
    'insurance': 'Insurance',
    'mortgage': 'Mortgage',
    'utilities': 'Utilities',
    'other fees': 'Other Fees',
    'year built': 'Year Built',
    'floor level': 'Floor Level',
    'room count': 'Room Count',
    'bedroom count': 'Bedroom Count',
    'bathroom count': 'Bathroom Count',
    'basement': 'Basement',
    'finish': 'Finish',
    'air conditioning': 'Air Conditioning',
    'heating': 'Heating',
    'compound': 'Compound',
    'landscaped': 'Landscaped',
    'quietness': 'Quietness',
    'transportation': 'Transportation',
    'garbage service': 'Garbage Service',
    'street lighting': 'Street Lighting',
    'shopping': 'Shopping',
    'hospital': 'Hospital',
    'school': 'School',
    'entertainment': 'Entertainment',
    'view': 'View'
}

def get_short_display_name(full_field_string):
    """
    Returns the display-friendly name for the given field.
    """
    return short_display_names.get(full_field_string, full_field_string)

def export_to_excel():
    """
    Exports the last calculation results to an Excel file with formatted styling.
    """
    global last_results
    
    if not last_results:
        messagebox.showwarning("No Data", "Please run the calculation first before exporting to Excel.")
        return
    
    try:
        # Ask user for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Property_Comparison_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return  # User cancelled
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Comparison"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(name='Arial', size=14, bold=True)
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header information
        ws['A1'] = "Property Comparison Report"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = center_alignment
        
        ws['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:D2')
        ws['A2'].alignment = center_alignment
        
        ws['A3'] = f"Project: {last_results['project_name']}"
        ws['A3'].font = Font(name='Arial', size=12, bold=True)
        ws.merge_cells('A3:D3')
        
        # Add empty row
        current_row = 5
        
        # Write column headers
        ws[f'A{current_row}'] = "Criteria"
        ws[f'B{current_row}'] = "Property 1"
        ws[f'C{current_row}'] = "Property 2"
        ws[f'D{current_row}'] = "Property 3"
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Write property data
        # Create list of unique field names (30 criteria)
        criteria_fields = []
        seen = set()
        for i in range(1, len(fields)):
            field_key = fields[i]
            if field_key not in seen:
                criteria_fields.append(field_key)
                seen.add(field_key)
                if len(criteria_fields) == 30:
                    break
        
        # Now write data for each criterion across all three properties
        for idx, field_key in enumerate(criteria_fields):
            short_name = get_short_display_name(field_key)
            
            ws[f'A{current_row}'] = short_name
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = thin_border
            
            # Property 1 - field index: 1 + idx
            prop1_field = fields[1 + idx]
            value1 = last_results['property_1'].get(prop1_field, "N/A")
            ws[f'B{current_row}'] = str(value1)
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = left_alignment
            
            # Property 2 - field index: 31 + idx
            prop2_field = fields[31 + idx]
            value2 = last_results['property_2'].get(prop2_field, "N/A")
            ws[f'C{current_row}'] = str(value2)
            ws[f'C{current_row}'].border = thin_border
            ws[f'C{current_row}'].alignment = left_alignment
            
            # Property 3 - field index: 61 + idx
            prop3_field = fields[61 + idx]
            value3 = last_results['property_3'].get(prop3_field, "N/A")
            ws[f'D{current_row}'] = str(value3)
            ws[f'D{current_row}'].border = thin_border
            ws[f'D{current_row}'].alignment = left_alignment
            
            current_row += 1
        
        # Add summary section
        current_row += 1
        ws[f'A{current_row}'] = "SUMMARY"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        current_row += 1
        
        # Write scores
        ws[f'A{current_row}'] = "Total Score"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = last_results['property_1']['score']
        ws[f'C{current_row}'] = last_results['property_2']['score']
        ws[f'D{current_row}'] = last_results['property_3']['score']
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
            ws[f'{col}{current_row}'].alignment = center_alignment
        
        current_row += 1
        
        # Write percentages
        ws[f'A{current_row}'] = "Percentage"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{last_results['property_1']['percentage']:.2f}%"
        ws[f'C{current_row}'] = f"{last_results['property_2']['percentage']:.2f}%"
        ws[f'D{current_row}'] = f"{last_results['property_3']['percentage']:.2f}%"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
            ws[f'{col}{current_row}'].alignment = center_alignment
        
        current_row += 2
        
        # Write recommendation
        ws[f'A{current_row}'] = "RECOMMENDATION"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 1
        ws[f'A{current_row}'] = last_results['recommendation']
        ws[f'A{current_row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'].alignment = center_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Save the workbook
        wb.save(file_path)
        
        messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
        print(f"Excel file exported to: {file_path}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel:\n{str(e)}")
        print(f"Excel export error: {e}")
        import traceback
        traceback.print_exc()

def project_name(entries):
    """
    Calculates and determines the scores for up to three properties based on user inputs.
    Prints results to console and to the 'ws which property to buy or rent.txt' file.
    Includes robust input validation and error handling.
    """
    global last_results
    last_results = {'property_1': {}, 'property_2': {}, 'property_3': {}}
    
    try:
        pr = entries['project name'].get()
        last_results['project_name'] = pr
        print(f"Project: {pr}")
        print(f"Project: {pr}", file=f)

        # --- Property No. 1 Calculations ---
        print("\n--- Property No. 1 ---")
        print("\n--- Property No. 1 ---", file=f)

        prop1_score_components = []
        for i in range(1, 31): # Fields 1 to 30 for property 1
            field_key = fields[i]
            short_name = get_short_display_name(field_key)
            value = entries[field_key].get()
            last_results['property_1'][field_key] = value

            if field_key == 'address':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'area units':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'FHA mortgage':
                avt = int(value)
                if avt == 5:
                    print(f"{short_name}: Yes (Score: {avt})")
                    print(f"{short_name}: Yes (Score: {avt})", file=f)
                else:
                    print(f"{short_name}: No (Score: {avt})")
                    print(f"{short_name}: No (Score: {avt})", file=f)
                prop1_score_components.append(avt)
            else:
                try:
                    float_value = float(value)
                    print(f"{short_name}: {float_value}")
                    print(f"{short_name}: {float_value}", file=f)
                    prop1_score_components.append(float_value)
                except ValueError:
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.")
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.", file=f)

        prop1 = sum(prop1_score_components)
        last_results['property_1']['score'] = prop1
        print(f"Score of Property 1: {prop1}")
        print(f"Score of Property 1: {prop1}", file=f)

        evp1 = 100 * prop1 / 150
        last_results['property_1']['percentage'] = evp1
        print(f"Percentage of Property 1: {evp1:.2f}%")
        print(f"Percentage of Property 1: {evp1:.2f}%", file=f)

        # --- Property No. 2 Calculations ---
        print("\n--- Property No. 2 ---")
        print("\n--- Property No. 2 ---", file=f)

        prop2_score_components = []
        for i in range(31, 61): # Fields 31 to 60 for property 2
            field_key = fields[i]
            short_name = get_short_display_name(field_key)
            value = entries[field_key].get()
            last_results['property_2'][field_key] = value

            if field_key == 'address':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'area units':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'FHA mortgage':
                avt2 = int(value)
                if avt2 == 5:
                    print(f"{short_name}: Yes (Score: {avt2})")
                    print(f"{short_name}: Yes (Score: {avt2})", file=f)
                else:
                    print(f"{short_name}: No (Score: {avt2})")
                    print(f"{short_name}: No (Score: {avt2})", file=f)
                prop2_score_components.append(avt2)
            else:
                try:
                    float_value = float(value)
                    print(f"{short_name}: {float_value}")
                    print(f"{short_name}: {float_value}", file=f)
                    prop2_score_components.append(float_value)
                except ValueError:
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.")
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.", file=f)

        prop2 = sum(prop2_score_components)
        last_results['property_2']['score'] = prop2
        print(f"Score of Property 2: {prop2}")
        print(f"Score of Property 2: {prop2}", file=f)

        evp2 = 100 * prop2 / 150
        last_results['property_2']['percentage'] = evp2
        print(f"Percentage of Property 2: {evp2:.2f}%")
        print(f"Percentage of Property 2: {evp2:.2f}%", file=f)

        # --- Property No. 3 Calculations ---
        print("\n--- Property No. 3 ---")
        print("\n--- Property No. 3 ---", file=f)

        prop3_score_components = []
        for i in range(61, 91): # Fields 61 to 90 for property 3
            field_key = fields[i]
            short_name = get_short_display_name(field_key)
            value = entries[field_key].get()
            last_results['property_3'][field_key] = value

            if field_key == 'address':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'area units':
                print(f"{short_name}: {value}")
                print(f"{short_name}: {value}", file=f)
            elif field_key == 'FHA mortgage':
                avt3 = int(value)
                if avt3 == 5:
                    print(f"{short_name}: Yes (Score: {avt3})")
                    print(f"{short_name}: Yes (Score: {avt3})", file=f)
                else:
                    print(f"{short_name}: No (Score: {avt3})")
                    print(f"{short_name}: No (Score: {avt3})", file=f)
                prop3_score_components.append(avt3)
            else:
                try:
                    float_value = float(value)
                    print(f"{short_name}: {float_value}")
                    print(f"{short_name}: {float_value}", file=f)
                    prop3_score_components.append(float_value)
                except ValueError:
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.")
                    print(f"Warning: Non-numeric input for {short_name}: '{value}'. Skipping for score calculation.", file=f)

        prop3 = sum(prop3_score_components)
        last_results['property_3']['score'] = prop3
        print(f"Score of Property 3: {prop3}")
        print(f"Score of Property 3: {prop3}", file=f)

        evp3 = 100 * prop3 / 150
        last_results['property_3']['percentage'] = evp3
        print(f"Percentage of Property 3: {evp3:.2f}%")
        print(f"Percentage of Property 3: {evp3:.2f}%", file=f)

        # --- Comparison and Recommendation ---
        print("\n--- Recommendation ---")
        print("\n--- Recommendation ---", file=f)

        scores = {
            "Property 1": evp1,
            "Property 2": evp2,
            "Property 3": evp3
        }

        best_property = max(scores, key=scores.get)
        max_score = scores[best_property]
        recommendation = f"Based on the inputs, the recommended property is: {best_property} with a score of {max_score:.2f}%"
        last_results['recommendation'] = recommendation

        print(f"\n{recommendation}")
        print(f"\n{recommendation}", file=f)
        
        messagebox.showinfo("Calculation Complete", f"{recommendation}\n\nYou can now export to Excel.")

    except ValueError as e:
        error_msg = f"Error: Invalid input. Please ensure all numerical fields contain valid numbers. Details: {e}"
        print(error_msg)
        print(error_msg, file=f)
        messagebox.showerror("Input Error", error_msg)
    except Exception as e:
        error_msg = f"An unexpected error occurred: {e}"
        print(error_msg)
        print(error_msg, file=f)
        messagebox.showerror("Error", error_msg)

def makeform(root_window, parent_frame):
    """
    Creates the Tkinter form with input fields and buttons in a columnar layout
    for three properties.
    """
    entries = {}
    # Global project name field
    lab_proj = tk.Label(parent_frame, width=80, text=get_short_display_name(fields[0]) + ": ", anchor='w', font=("georgia", 10))
    lab_proj.grid(row=0, column=0, columnspan=2, pady=5, sticky="w")
    ent_proj = tk.Entry(parent_frame, width=30, font=("georgia", 10))
    ent_proj.grid(row=0, column=2, columnspan=4, pady=5, sticky="ew")
    ent_proj.insert(0, "My Project")
    entries[fields[0]] = ent_proj

    # Headers for columns
    tk.Label(parent_frame, text="Property 1 Criteria", font=("georgia", 12, "bold")).grid(row=1, column=0, pady=5)
    tk.Label(parent_frame, text="Points", font=("georgia", 12, "bold")).grid(row=1, column=1, pady=5)
    tk.Label(parent_frame, text="Property 2 Criteria", font=("georgia", 12, "bold")).grid(row=1, column=2, pady=5)
    tk.Label(parent_frame, text="Points", font=("georgia", 12, "bold")).grid(row=1, column=3, pady=5)
    tk.Label(parent_frame, text="Property 3 Criteria", font=("georgia", 12, "bold")).grid(row=1, column=4, pady=5)
    tk.Label(parent_frame, text="Points", font=("georgia", 12, "bold")).grid(row=1, column=5, pady=5)

    # Number of criteria per property (excluding 'project name')
    num_criteria_per_property = (len(fields) - 1) // 3
    
    # Start row for property-specific fields
    current_row = 2

    for i in range(num_criteria_per_property):
        # Property 1
        field_p1 = fields[1 + i]
        short_name_p1 = get_short_display_name(field_p1)
        lab_p1 = tk.Label(parent_frame, width=40, text=short_name_p1 + ": ", anchor='w', font=("georgia", 10))
        lab_p1.grid(row=current_row, column=0, pady=1, sticky="w")
        ent_p1 = tk.Entry(parent_frame, width=10, font=("georgia", 10))
        ent_p1.grid(row=current_row, column=1, pady=1, sticky="ew")
        ent_p1.insert(0, "1")
        entries[field_p1] = ent_p1

        # Property 2
        field_p2 = fields[1 + num_criteria_per_property + i]
        short_name_p2 = get_short_display_name(field_p2)
        lab_p2 = tk.Label(parent_frame, width=40, text=short_name_p2 + ": ", anchor='w', font=("georgia", 10))
        lab_p2.grid(row=current_row, column=2, pady=1, sticky="w")
        ent_p2 = tk.Entry(parent_frame, width=10, font=("georgia", 10))
        ent_p2.grid(row=current_row, column=3, pady=1, sticky="ew")
        ent_p2.insert(0, "1")
        entries[field_p2] = ent_p2

        # Property 3
        field_p3 = fields[1 + 2 * num_criteria_per_property + i]
        short_name_p3 = get_short_display_name(field_p3)
        lab_p3 = tk.Label(parent_frame, width=40, text=short_name_p3 + ": ", anchor='w', font=("georgia", 10))
        lab_p3.grid(row=current_row, column=4, pady=1, sticky="w")
        ent_p3 = tk.Entry(parent_frame, width=10, font=("georgia", 10))
        ent_p3.grid(row=current_row, column=5, pady=1, sticky="ew")
        ent_p3.insert(0, "1")
        entries[field_p3] = ent_p3

        current_row += 1

    # Create a frame for buttons
    button_frame = tk.Frame(parent_frame)
    run_button = tk.Button(button_frame, padx=10, pady=5, text="Run Calculation", font=("georgia", 10))
    run_button.config(command=(lambda e=entries: project_name(e)))
    
    export_button = tk.Button(button_frame, padx=10, pady=5, text="Export to Excel", font=("georgia", 10),
                             command=export_to_excel, bg="#4CAF50", fg="white")
    
    quit_button = tk.Button(button_frame, padx=10, pady=5, text='Quit', font=("georgia", 10),
                            command=root_window.destroy)

    button_frame.grid(row=current_row, column=0, columnspan=6, sticky="nsew", pady=10)
    run_button.pack(side="left", padx=5, pady=5)
    export_button.pack(side="left", padx=5, pady=5)
    quit_button.pack(side="left", padx=5, pady=5)

def onFrameConfigure(canvas_widget):
    """
    Reset the scroll region to encompass the inner frame's content.
    """
    canvas_widget.configure(scrollregion=canvas_widget.bbox("all"))

if __name__ == '__main__':
    # Initialize the main Tkinter window
    root = tk.Tk()
    root.title("Ws which property to buy or rent")

    # Create a canvas to hold the scrollable frame
    canvas = tk.Canvas(root, width=1000, borderwidth=0)
    frame = tk.Frame(canvas)

    # Create scrollbars
    vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    hsb = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)

    canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    canvas.pack(side="left", fill="both", expand=True)

    canvas.create_window((3, 3), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))

    # Create the form elements
    makeform(root, frame)

    # Start the Tkinter event loop
    root.mainloop()

# This part runs after the Tkinter window is closed
input("Press Enter to close the program and view the output file (ws which property to buy or rent.txt)...")
f.close()
