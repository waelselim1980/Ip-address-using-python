import matplotlib.pyplot as plt # pyright: ignore[reportMissingModuleSource]
import datetime
import socket
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import ntplib # pyright: ignore[reportMissingImports]
from socket import timeout
import openpyxl # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]
hostname = socket.gethostname()    
IPAddr = socket.gethostbyname(hostname)  
g=str(IPAddr)
w=str("192.168.1.163")

# License check - Hard coded expiration date
LICENSE_EXPIRATION = datetime.datetime(2026, 8, 19)
NTP_SERVERS = ['pool.ntp.org', 'time.nist.gov', 'time.google.com']

# Global variable to store last calculation results
last_calculation_data = None


def get_network_time():
    """Get current time from NTP server"""
    for server in NTP_SERVERS:
        try:
            client = ntplib.NTPClient()
            response = client.request(server, timeout=5)
            return datetime.datetime.fromtimestamp(response.tx_time)
        except (ntplib.NTPException, timeout):
            continue
    return None


def check_license():
    """Check if the software license is still valid"""
    try:
        network_time = get_network_time()
        
        if network_time:
            current_date = network_time
        else:
            current_date = datetime.datetime.now()
            messagebox.showwarning(
                "License Warning", 
                "Could not verify time with server. Using system time."
            )
        
        if current_date > LICENSE_EXPIRATION:
            messagebox.showerror(
                "License Expired", 
                f"This software license expired on {LICENSE_EXPIRATION.strftime('%B %d, %Y')}.\n"
                "Please contact the developer for license renewal."
            )
            sys.exit()
        else:
            days_remaining = (LICENSE_EXPIRATION - current_date).days
            if days_remaining <= 7:
                messagebox.showwarning(
                    "License Warning", 
                    f"Software license expires in {days_remaining} days on "
                    f"{LICENSE_EXPIRATION.strftime('%B %d, %Y')}."
                )
    except Exception as e:
        messagebox.showerror("License Error", f"License verification failed: {str(e)}")
        sys.exit()


# Check license on startup
check_license()

# Get system info
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
current_date = datetime.datetime.now()
g=str(IPAddr)
w=str('192.168.1.163')
if(g!=w):
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")
# Define input fields with tooltips
fields = [
    ('Project Name', 'Enter a name for this analysis'),
    ('Currency (USD, EUR, etc.)', 'Currency symbol for all monetary values'),
    ('Current Monthly Rent', 'Current monthly rent payment'),
    ('Monthly Rent Insurance', 'Monthly renter\'s insurance cost'),
    ('Number of Years (1-20)', 'Analysis period in years (1-20)'),
    ('Annual Rent Increase (%)', 'Expected annual rent increase percentage'),
    ('FHA Mortgage (1=Yes, 2=No)', 'Enter 1 for FHA loan, 2 for conventional'),
    ('Property Purchase Price', 'Total purchase price of property'),
    ('Down Payment (%)', 'Down payment as percentage of purchase price'),
    ('Interest Rate (%)', 'Annual mortgage interest rate'),
    ('FHA Insurance Premium (%)', 'FHA insurance premium (if applicable)'),
    ('Annual Property Tax', 'Annual property tax amount'),
    ('Annual Property Insurance', 'Annual homeowner\'s insurance'),
    ('Annual Tax Deduction', 'Annual tax savings from deductions'),
    ('Annual Maintenance Cost', 'Annual maintenance and repairs'),
    ('Other Closing Costs', 'Additional closing costs (appraisal, etc.)'),
    ('Estimated Selling Price', 'Expected property value at end of period'),
    ('Selling Closing Costs', 'Costs to sell property (agent fees, etc.)')
]


def calculate_monthly_payment(principal, annual_rate, years):
    """Calculate monthly mortgage payment using standard formula"""
    if annual_rate == 0:
        return principal / (years * 12)
    
    monthly_rate = annual_rate / 100 / 12
    num_payments = years * 12
    
    monthly_payment = principal * (monthly_rate * (1 + monthly_rate)**num_payments) / \
                     ((1 + monthly_rate)**num_payments - 1)
    
    return monthly_payment


def calculate_total_rent(monthly_rent, monthly_insurance, years, annual_increase_rate):
    """Calculate total rent over the specified years with annual increases"""
    total_rent = 0
    current_monthly_cost = monthly_rent + monthly_insurance
    rent_breakdown = []
    
    for year in range(int(years)):
        yearly_cost = current_monthly_cost * 12
        total_rent += yearly_cost
        rent_breakdown.append({
            'year': year + 1,
            'monthly_cost': current_monthly_cost,
            'yearly_cost': yearly_cost,
            'cumulative': total_rent
        })
        current_monthly_cost *= (1 + annual_increase_rate / 100)
    
    return total_rent, rent_breakdown


def export_to_excel(data):
    """Export analysis results to Excel with formatting"""
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"buy_vs_rent_analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        currency_format = '#,##0.00'
        percent_format = '0.00%'
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = 'BUY vs RENT PROPERTY ANALYSIS'
        ws['A1'].font = title_font
        ws.merge_cells('A1:C1')
        
        ws['A2'] = 'Created by Engineer Wael Sherif Selim'
        ws['A3'] = f"Analysis Date: {data['analysis_date']}"
        ws['A4'] = f"Project: {data['project_name']}"
        
        row = 6
        
        # Input Parameters Section
        ws[f'A{row}'] = 'INPUT PARAMETERS'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        input_params = [
            ('Currency', data['currency']),
            ('Analysis Period', f"{data['years']} years"),
            ('Current Monthly Rent', data['monthly_rent']),
            ('Monthly Rent Insurance', data['monthly_insurance']),
            ('Annual Rent Increase', f"{data['annual_increase']}%"),
            ('Property Purchase Price', data['property_price']),
            ('Down Payment', f"{data['down_payment_pct']}%"),
            ('Interest Rate', f"{data['interest_rate']}%"),
            ('FHA Mortgage', data['is_fha_text']),
            ('Annual Property Tax', data['annual_tax']),
            ('Annual Property Insurance', data['annual_insurance']),
            ('Annual Tax Deduction', data['tax_deduction']),
            ('Annual Maintenance', data['maintenance_cost']),
            ('Other Closing Costs', data['closing_costs']),
            ('Estimated Selling Price', data['selling_price']),
            ('Selling Closing Costs', data['selling_costs'])
        ]
        
        for param, value in input_params:
            ws[f'A{row}'] = param
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)) and param != 'Analysis Period':
                ws[f'B{row}'].number_format = currency_format
            row += 1
        
        row += 1
        
        # Renting Analysis Section
        ws[f'A{row}'] = 'RENTING ANALYSIS'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws[f'A{row}'] = 'Total Cost of Renting'
        ws[f'B{row}'] = data['total_rent']
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Buying Analysis Section
        ws[f'A{row}'] = 'BUYING ANALYSIS'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        buying_items = [
            ('Down Payment', data['down_payment']),
            ('Mortgage Amount', data['mortgage_amount']),
            ('FHA Insurance Premium', data['fha_insurance']),
            ('Monthly Mortgage Payment', data['monthly_mortgage']),
            ('Total Mortgage Payments', data['total_mortgage_payments']),
            ('Total Other Annual Costs', data['total_other_costs']),
            ('Total Tax Savings', data['total_tax_savings']),
            ('Initial Costs', data['initial_costs']),
            ('Property Appreciation', data['property_appreciation']),
            ('Net Proceeds from Sale', data['net_proceeds']),
            ('Total Cost of Buying', data['total_buy_cost'])
        ]
        
        for item, value in buying_items:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            if item == 'Total Cost of Buying':
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Recommendation Section
        ws[f'A{row}'] = 'RECOMMENDATION'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws[f'A{row}'] = 'Decision'
        ws[f'B{row}'] = data['recommendation']
        ws[f'B{row}'].font = Font(bold=True, size=12, color="FF0000" if data['recommendation'] == 'BUY' else "0000FF")
        row += 1
        
        ws[f'A{row}'] = 'Savings Amount'
        ws[f'B{row}'] = data['savings']
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = Font(bold=True)
        
        # Create Yearly Breakdown Sheet
        ws2 = wb.create_sheet("Yearly Breakdown")
        ws2['A1'] = 'YEARLY RENT BREAKDOWN'
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:D1')
        
        headers = ['Year', 'Monthly Cost', 'Yearly Cost', 'Cumulative Total']
        for col, header in enumerate(headers, start=1):
            cell = ws2.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for idx, rent_data in enumerate(data['rent_breakdown'], start=4):
            ws2[f'A{idx}'] = rent_data['year']
            ws2[f'B{idx}'] = rent_data['monthly_cost']
            ws2[f'B{idx}'].number_format = currency_format
            ws2[f'C{idx}'] = rent_data['yearly_cost']
            ws2[f'C{idx}'].number_format = currency_format
            ws2[f'D{idx}'] = rent_data['cumulative']
            ws2[f'D{idx}'].number_format = currency_format
        
        # Adjust column widths
        for ws_curr in [ws, ws2]:
            for column in ws_curr.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_curr.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        messagebox.showinfo("Export Success", f"Analysis exported to:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel:\n{str(e)}")


def project_calculation(entries, export_excel_btn):
    """Main calculation function"""
    global last_calculation_data
    
    try:
        # Get input values
        project_name = entries['Project Name'].get().strip() or "Unnamed Project"
        currency = entries['Currency (USD, EUR, etc.)'].get().strip() or "USD"
        monthly_rent = float(entries['Current Monthly Rent'].get())
        monthly_insurance = float(entries['Monthly Rent Insurance'].get())
        years = int(entries['Number of Years (1-20)'].get())
        annual_increase = float(entries['Annual Rent Increase (%)'].get())
        is_fha = int(entries['FHA Mortgage (1=Yes, 2=No)'].get())
        property_price = float(entries['Property Purchase Price'].get())
        down_payment_pct = float(entries['Down Payment (%)'].get())
        interest_rate = float(entries['Interest Rate (%)'].get())
        fha_premium = float(entries['FHA Insurance Premium (%)'].get())
        annual_tax = float(entries['Annual Property Tax'].get())
        annual_insurance = float(entries['Annual Property Insurance'].get())
        tax_deduction = float(entries['Annual Tax Deduction'].get())
        maintenance_cost = float(entries['Annual Maintenance Cost'].get())
        closing_costs = float(entries['Other Closing Costs'].get())
        selling_price = float(entries['Estimated Selling Price'].get())
        selling_costs = float(entries['Selling Closing Costs'].get())
        
        # Validate inputs
        if years < 1 or years > 20:
            raise ValueError("Number of years must be between 1 and 20")
        if is_fha not in [1, 2]:
            raise ValueError("FHA Mortgage must be 1 (Yes) or 2 (No)")
        if down_payment_pct < 0 or down_payment_pct > 100:
            raise ValueError("Down payment must be between 0 and 100%")
        if monthly_rent < 0 or property_price < 0:
            raise ValueError("Monetary values cannot be negative")
        
        # RENTING CALCULATIONS
        total_rent, rent_breakdown = calculate_total_rent(monthly_rent, monthly_insurance, years, annual_increase)
        
        # BUYING CALCULATIONS
        down_payment = property_price * down_payment_pct / 100
        mortgage_amount = property_price - down_payment
        
        fha_insurance = 0
        if is_fha == 1:
            fha_insurance = property_price * fha_premium / 100
        
        monthly_mortgage = calculate_monthly_payment(mortgage_amount, interest_rate, years)
        total_mortgage_payments = monthly_mortgage * 12 * years
        total_annual_costs = annual_tax + annual_insurance + maintenance_cost
        total_other_costs = total_annual_costs * years
        total_tax_savings = tax_deduction * years
        initial_costs = down_payment + closing_costs + fha_insurance
        property_appreciation = selling_price - property_price
        net_proceeds = selling_price - selling_costs
        
        total_buy_cost = (initial_costs + total_mortgage_payments + total_other_costs + 
                        selling_costs - total_tax_savings - net_proceeds)
        
        # COMPARISON
        savings = abs(total_rent - total_buy_cost)
        
        if total_buy_cost < total_rent:
            recommendation = "BUY"
            result_text = f"RECOMMENDATION: BUY\nSavings: {savings:,.2f} {currency}"
        else:
            recommendation = "RENT"
            result_text = f"RECOMMENDATION: RENT\nSavings: {savings:,.2f} {currency}"
        
        # Store calculation data for Excel export
        last_calculation_data = {
            'analysis_date': current_date.strftime('%Y-%m-%d %H:%M:%S'),
            'project_name': project_name,
            'currency': currency,
            'monthly_rent': monthly_rent,
            'monthly_insurance': monthly_insurance,
            'years': years,
            'annual_increase': annual_increase,
            'is_fha': is_fha,
            'is_fha_text': 'Yes' if is_fha == 1 else 'No',
            'property_price': property_price,
            'down_payment_pct': down_payment_pct,
            'interest_rate': interest_rate,
            'fha_premium': fha_premium,
            'annual_tax': annual_tax,
            'annual_insurance': annual_insurance,
            'tax_deduction': tax_deduction,
            'maintenance_cost': maintenance_cost,
            'closing_costs': closing_costs,
            'selling_price': selling_price,
            'selling_costs': selling_costs,
            'total_rent': total_rent,
            'rent_breakdown': rent_breakdown,
            'down_payment': down_payment,
            'mortgage_amount': mortgage_amount,
            'fha_insurance': fha_insurance,
            'monthly_mortgage': monthly_mortgage,
            'total_mortgage_payments': total_mortgage_payments,
            'total_other_costs': total_other_costs,
            'total_tax_savings': total_tax_savings,
            'initial_costs': initial_costs,
            'property_appreciation': property_appreciation,
            'net_proceeds': net_proceeds,
            'total_buy_cost': total_buy_cost,
            'savings': savings,
            'recommendation': recommendation
        }
        
        # Create output file
        with open("buy_vs_rent_analysis.txt", "w") as f:
            f.write("=" * 60 + "\n")
            f.write("BUY vs RENT PROPERTY ANALYSIS\n")
            f.write("Created by Engineer Wael Sherif Selim\n")
            f.write("Email: wael.sherif.selim@gmail.com\n")
            f.write(f"Analysis Date: {current_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            f.write("INPUT PARAMETERS:\n")
            f.write(f"Project Name: {project_name}\n")
            f.write(f"Currency: {currency}\n")
            f.write(f"Analysis Period: {years} years\n")
            f.write(f"Current Monthly Rent: {monthly_rent:,.2f} {currency}\n")
            f.write(f"Monthly Rent Insurance: {monthly_insurance:,.2f} {currency}\n")
            f.write(f"Annual Rent Increase: {annual_increase}%\n")
            f.write(f"Property Purchase Price: {property_price:,.2f} {currency}\n")
            f.write(f"Down Payment: {down_payment_pct}%\n")
            f.write(f"Interest Rate: {interest_rate}%\n")
            f.write(f"FHA Mortgage: {'Yes' if is_fha == 1 else 'No'}\n")
            f.write("\n" + "=" * 60 + "\n")
            
            f.write("RENTING ANALYSIS:\n")
            f.write(f"Total Cost of Renting: {total_rent:,.2f} {currency}\n")
            
            f.write("\nBUYING ANALYSIS:\n")
            f.write(f"Down Payment: {down_payment:,.2f} {currency}\n")
            f.write(f"Mortgage Amount: {mortgage_amount:,.2f} {currency}\n")
            
            if is_fha == 1:
                f.write(f"FHA Insurance Premium: {fha_insurance:,.2f} {currency}\n")
            
            f.write(f"Monthly Mortgage Payment: {monthly_mortgage:,.2f} {currency}\n")
            f.write(f"Total Mortgage Payments: {total_mortgage_payments:,.2f} {currency}\n")
            f.write(f"Total Other Annual Costs ({years} years): {total_other_costs:,.2f} {currency}\n")
            f.write(f"Total Tax Savings ({years} years): {total_tax_savings:,.2f} {currency}\n")
            f.write(f"Initial Costs (Down Payment + Closing + FHA): {initial_costs:,.2f} {currency}\n")
            f.write(f"Property Appreciation: {property_appreciation:,.2f} {currency}\n")
            f.write(f"Net Proceeds from Sale: {net_proceeds:,.2f} {currency}\n")
            f.write(f"\nTotal Cost of Buying: {total_buy_cost:,.2f} {currency}\n")
            
            f.write("\n" + "=" * 60 + "\n")
            f.write("RECOMMENDATION:\n")
            f.write(f"Total Cost of Renting: {total_rent:,.2f} {currency}\n")
            f.write(f"Total Cost of Buying: {total_buy_cost:,.2f} {currency}\n")
            f.write(f"\nRECOMMENDATION: {recommendation}\n")
            f.write(f"Savings by {'buying' if recommendation == 'BUY' else 'renting'}: {savings:,.2f} {currency}\n")
            
            f.write(f"\nAnalysis completed on: {current_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Software License Valid Until: {LICENSE_EXPIRATION.strftime('%Y-%m-%d')}\n")
        
        # Enable Excel export button
        export_excel_btn.config(state='normal')
        
        # Show result in popup
        messagebox.showinfo(
            "Analysis Complete", 
            f"Analysis saved to 'buy_vs_rent_analysis.txt'\n\n{result_text}\n\n"
            f"You can now export to Excel!"
        )
        
    except ValueError as e:
        messagebox.showerror("Input Error", f"Please check your inputs:\n{str(e)}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")


def create_tooltip(widget, text):
    """Create a tooltip for a widget"""
    tooltip = None
    
    def on_enter(event):
        nonlocal tooltip
        x, y, _, _ = widget.bbox("insert")
        x += widget.winfo_rootx() + 25
        y += widget.winfo_rooty() + 25
        
        tooltip = tk.Toplevel(widget)
        tooltip.wm_overrideredirect(True)
        tooltip.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(
            tooltip, text=text, justify='left',
            background="#ffffe0", relief='solid', borderwidth=1,
            font=("Arial", 9)
        )
        label.pack()
    
    def on_leave(event):
        nonlocal tooltip
        if tooltip:
            tooltip.destroy()
            tooltip = None
    
    widget.bind("<Enter>", on_enter)
    widget.bind("<Leave>", on_leave)


def makeform(root, frame):
    """Create the input form"""
    entries = {}
    row = 0
    
    # Header
    header = tk.Label(
        frame, 
        text="Buy vs Rent Property Calculator",
        font=("Georgia", 16, "bold"),
        fg="#366092"
    )
    header.grid(row=row, column=0, columnspan=2, pady=10)
    row += 1
    
    subheader = tk.Label(
        frame,
        text="Created by Engineer Wael Sherif Selim",
        font=("Georgia", 9),
        fg="gray"
    )
    subheader.grid(row=row, column=0, columnspan=2, pady=5)
    row += 1
    
    separator = ttk.Separator(frame, orient='horizontal')
    separator.grid(row=row, column=0, columnspan=2, sticky="ew", pady=10)
    row += 1
    
    for field, tooltip_text in fields:
        lab = tk.Label(
            frame, 
            width=40, 
            text=field + ": ", 
            font=("Georgia", 10), 
            anchor='w'
        )
        lab.grid(row=row, column=0, sticky="w", padx=5, pady=4)
        
        ent = tk.Entry(frame, width=25, font=("Georgia", 10))
        ent.grid(row=row, column=1, sticky="w", padx=5, pady=4)
        
        # Set default values
        if field == 'Project Name':
            ent.insert(0, "My Project")
        elif field == 'Currency (USD, EUR, etc.)':
            ent.insert(0, "USD")
        elif field == 'Number of Years (1-20)':
            ent.insert(0, "10")
        elif field == 'FHA Mortgage (1=Yes, 2=No)':
            ent.insert(0, "2")
        elif "%" in field:
            ent.insert(0, "0.0")
        else:
            ent.insert(0, "0")
        
        # Add tooltip
        create_tooltip(lab, tooltip_text)
        
        entries[field] = ent
        row += 1
    
    # Buttons
    separator2 = ttk.Separator(frame, orient='horizontal')
    separator2.grid(row=row, column=0, columnspan=2, sticky="ew", pady=10)
    row += 1
    
    button_frame = tk.Frame(frame)
    button_frame.grid(row=row, column=0, columnspan=2, pady=10)
    
    # Export to Excel button (initially disabled)
    export_excel_btn = tk.Button(
        button_frame, 
        text="📊 Export to Excel", 
        font=("Georgia", 11),
        command=lambda: export_to_excel(last_calculation_data),
        bg="#2E7D32", 
        fg="white", 
        padx=15, 
        pady=5,
        state='disabled'
    )
    export_excel_btn.pack(side="left", padx=5)
    
    calculate_btn = tk.Button(
        button_frame, 
        text="🔍 Calculate", 
        font=("Georgia", 11),
        command=lambda: project_calculation(entries, export_excel_btn),
        bg="#4CAF50", 
        fg="white", 
        padx=20, 
        pady=5
    )
    calculate_btn.pack(side="left", padx=5)
    
    quit_btn = tk.Button(
        button_frame, 
        text="❌ Quit", 
        font=("Georgia", 11),
        command=root.destroy, 
        bg="#f44336", 
        fg="white", 
        padx=20, 
        pady=5
    )
    quit_btn.pack(side="left", padx=5)


def onFrameConfigure(canvas):
    """Reset the scroll region to encompass the inner frame"""
    canvas.configure(scrollregion=canvas.bbox("all"))


if __name__ == '__main__':
    root = tk.Tk()
    root.title("Buy vs Rent Property Calculator - Professional Edition")
    root.geometry("700x750")
    
    # Set icon (if available)
    try:
        root.iconbitmap('icon.ico')
    except:
        pass
    
    # License info label
    license_label = tk.Label(
        root, 
        text=f"Licensed until: {LICENSE_EXPIRATION.strftime('%B %d, %Y')} | "
             f"Version 2.0 | Contact: wael.sherif.selim@gmail.com",
        font=("Georgia", 8), 
        fg="gray"
    )
    license_label.pack(pady=5)
    
    # Create scrollable frame
    canvas = tk.Canvas(root, borderwidth=0, highlightthickness=0)
    frame = tk.Frame(canvas, bg="white")
    vsb = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)
    
    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((4, 4), window=frame, anchor="nw")
    
    frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
    
    makeform(root, frame)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()
