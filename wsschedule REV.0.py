import socket
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
import networkx as nx # pyright: ignore[reportMissingModuleSource]
import matplotlib.pyplot as plt # pyright: ignore[reportMissingModuleSource]
from datetime import datetime, timedelta, date
from matplotlib.patches import Patch # pyright: ignore[reportMissingModuleSource]
import matplotlib.dates as mdates # pyright: ignore[reportMissingModuleSource]
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg # pyright: ignore[reportMissingModuleSource]
from enum import Enum
import sys
import ntplib # pyright: ignore[reportMissingImports]
from socket import gaierror
import openpyxl # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import calendar as cal_module
hostname = socket.gethostname()    
IPAddr = socket.gethostbyname(hostname)  
g=str(IPAddr)
w=str("192.168.1.163")
if(g!=w): # pyright: ignore[reportUndefinedVariable]
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")

# License configuration
LICENSE_EXPIRY = date(2026, 8, 31)
NTP_SERVER = 'pool.ntp.org'

def check_license():
    """Check if license is valid by verifying current date against NTP server"""
    try:
        # Get current time from NTP server
        ntp_client = ntplib.NTPClient()
        response = ntp_client.request(NTP_SERVER, timeout=3)
        current_time = datetime.fromtimestamp(response.tx_time).date()
        
        # Check license expiry
        if current_time > LICENSE_EXPIRY:
            messagebox.showerror("License Error", "License expired. Please contact the software provider.")
            sys.exit(1)
            
        # Check if system time is manipulated
        system_time = date.today()
        if abs((current_time - system_time).days) > 1:  # Allow 1 day difference for timezone
            messagebox.showwarning("Time Warning", "System date appears incorrect. Using NTP time for verification.")
            
    except (ntplib.NTPException, gaierror, OSError) as e:
        messagebox.showwarning("NTP Warning", f"Could not verify license with NTP server ({str(e)}). Falling back to system time.")
        system_time = date.today()
        if system_time > LICENSE_EXPIRY:
            messagebox.showerror("License Error", "License expired. Please contact the software provider.")
            sys.exit(1)
    g=str(IPAddr) # pyright: ignore[reportUndefinedVariable]
    w=str('192.168.1.163')
    if(g!=w):
       print("License Expired")
       print(1/0)
    else:
       print("Software Activated")

class RelationshipType(Enum):
    FS = "Finish-to-Start"
    SS = "Start-to-Start"
    FF = "Finish-to-Finish"
    SF = "Start-to-Finish"

class Calendar:
    def __init__(self, name="Standard Calendar"):
        self.name = name
        self.work_week = [True, True, True, True, True, False, False]  # Monday to Sunday
        self.work_hours = (8, 17)  # 8 AM to 5 PM
        self.vacations = []  # List of (start_date, end_date) tuples
        self.holidays = []  # List of specific holiday dates
    
    def add_vacation(self, start_date, end_date):
        """Add a vacation period"""
        self.vacations.append((start_date, end_date))
    
    def add_holiday(self, holiday_date):
        """Add a single holiday"""
        self.holidays.append(holiday_date)
    
    def is_working_day(self, date_obj):
        """Check if a date is a working day considering weekends, holidays and vacations"""
        # Convert datetime to date if necessary
        if isinstance(date_obj, datetime):
            date_obj = date_obj.date()
        
        # Check weekend
        if not self.work_week[date_obj.weekday()]:
            return False
        
        # Check holidays
        if date_obj in self.holidays:
            return False
        
        # Check vacations
        for vacation_start, vacation_end in self.vacations:
            # Convert to date if necessary
            if isinstance(vacation_start, datetime):
                vacation_start = vacation_start.date()
            if isinstance(vacation_end, datetime):
                vacation_end = vacation_end.date()
                
            if vacation_start <= date_obj <= vacation_end:
                return False
        
        return True
    
    def get_next_working_day(self, date_obj):
        """Get the next working day from given date"""
        # Convert to date if necessary
        if isinstance(date_obj, datetime):
            date_obj = date_obj.date()
            
        next_day = date_obj + timedelta(days=1)
        while not self.is_working_day(next_day):
            next_day += timedelta(days=1)
        # Return as datetime
        return datetime.combine(next_day, datetime.min.time())
    
    def calculate_working_days(self, start_date, duration):
        """Calculate end date considering only working days"""
        # Convert to date if necessary
        if isinstance(start_date, datetime):
            start_date_obj = start_date.date()
        else:
            start_date_obj = start_date
            
        current_date = start_date_obj
        days_remaining = duration
        
        while days_remaining > 0:
            if self.is_working_day(current_date):
                days_remaining -= 1
            if days_remaining > 0:
                current_date += timedelta(days=1)
        
        # Return as datetime
        return datetime.combine(current_date, datetime.min.time())

class Dependency:
    def __init__(self, from_activity_id, to_activity_id, rel_type=RelationshipType.FS, lag=0):
        self.from_activity_id = from_activity_id
        self.to_activity_id = to_activity_id
        self.rel_type = rel_type
        self.lag = lag

    def __repr__(self):
        return f"{self.from_activity_id} -> {self.to_activity_id} ({self.rel_type.value}, lag: {self.lag})"

class Activity:
    def __init__(self, id, name, duration, dependencies=None, resources=None):
        self.id = id
        self.name = name
        self.duration = duration
        self.dependencies = dependencies or []
        self.resources = resources or []
        self.early_start = None
        self.early_finish = None
        self.late_start = None
        self.late_finish = None
        self.slack = None

    def __repr__(self):
        return f"Activity({self.name}, Duration: {self.duration} days)"

class Project:
    def __init__(self, name, start_date):
        self.name = name
        # Ensure start_date is a datetime object
        if isinstance(start_date, date) and not isinstance(start_date, datetime):
            self.start_date = datetime.combine(start_date, datetime.min.time())
        else:
            self.start_date = start_date
        self.activities = {}
        self.graph = nx.DiGraph()
        self.critical_path = []
        self.language = 'en'  # Default language
        self.calendar = Calendar()  # Default calendar

    def add_activity(self, activity):
        # Removed the 500 activity limit
        self.activities[activity.id] = activity
        self.graph.add_node(activity.id)

    def add_dependency(self, from_activity_id, to_activity_id, rel_type=RelationshipType.FS, lag=0):
        if from_activity_id in self.activities and to_activity_id in self.activities:
            dependency = Dependency(from_activity_id, to_activity_id, rel_type, lag)
            self.activities[to_activity_id].dependencies.append(dependency)
            self.graph.add_edge(from_activity_id, to_activity_id, rel_type=rel_type, lag=lag)
        else:
            raise ValueError("One or both activity IDs do not exist")

    def remove_activity(self, activity_id):
        """Remove an activity and all its dependencies"""
        if activity_id not in self.activities:
            return False
        
        # Remove all dependencies involving this activity
        for other_activity in self.activities.values():
            other_activity.dependencies = [dep for dep in other_activity.dependencies if dep.from_activity_id != activity_id]
        
        # Remove from graph and activities dictionary
        if self.graph.has_node(activity_id):
            self.graph.remove_node(activity_id)
        del self.activities[activity_id]
        
        return True

    def remove_dependency(self, from_activity_id, to_activity_id):
        """Remove a specific dependency between two activities"""
        if to_activity_id not in self.activities:
            return False
        
        activity = self.activities[to_activity_id]
        original_count = len(activity.dependencies)
        activity.dependencies = [dep for dep in activity.dependencies if dep.from_activity_id != from_activity_id]
        
        # Remove from graph if edge exists
        if self.graph.has_edge(from_activity_id, to_activity_id):
            self.graph.remove_edge(from_activity_id, to_activity_id)
        
        return len(activity.dependencies) < original_count

    def calculate_schedule(self):
        try:
            ordered_activities = list(nx.topological_sort(self.graph))
        except nx.NetworkXUnfeasible:
            raise ValueError("Circular dependency detected")

        if not ordered_activities:
            return

        # Forward pass
        for activity_id in ordered_activities:
            activity = self.activities[activity_id]
            if not activity.dependencies:
                activity.early_start = self.start_date
            else:
                earliest_start = None
                for dep in activity.dependencies:
                    pred = self.activities[dep.from_activity_id]
                    if dep.rel_type == RelationshipType.FS:
                        if dep.lag == 0:
                            start = self.calendar.get_next_working_day(pred.early_finish)
                        else:
                            start = pred.early_finish + timedelta(days=dep.lag)
                    elif dep.rel_type == RelationshipType.SS:
                        start = pred.early_start + timedelta(days=dep.lag)
                    elif dep.rel_type == RelationshipType.FF:
                        start = pred.early_finish - timedelta(days=activity.duration - 1 - dep.lag)
                    elif dep.rel_type == RelationshipType.SF:
                        start = pred.early_start - timedelta(days=activity.duration - 1 + dep.lag)
                    
                    if start < self.start_date:
                        start = self.start_date
                    
                    if earliest_start is None or start > earliest_start:
                        earliest_start = start
                
                activity.early_start = earliest_start
            
            # Calculate early finish considering working days
            activity.early_finish = self.calendar.calculate_working_days(activity.early_start, activity.duration)

        # Backward pass
        last_activity_id = ordered_activities[-1]
        self.activities[last_activity_id].late_finish = self.activities[last_activity_id].early_finish
        
        for activity_id in reversed(ordered_activities):
            activity = self.activities[activity_id]
            successors = list(self.graph.successors(activity_id))
            
            if not successors:
                activity.late_finish = activity.early_finish
            else:
                latest_finish = None
                for succ_id in successors:
                    edge_data = self.graph.get_edge_data(activity_id, succ_id)
                    succ = self.activities[succ_id]
                    
                    if edge_data['rel_type'] == RelationshipType.FS:
                        if edge_data['lag'] == 0:
                            finish = succ.late_start - timedelta(days=1)
                        else:
                            finish = succ.late_start - timedelta(days=edge_data['lag'] + 1)
                    elif edge_data['rel_type'] == RelationshipType.SS:
                        finish = succ.late_start + timedelta(days=activity.duration - 1 - edge_data['lag'])
                    elif edge_data['rel_type'] == RelationshipType.FF:
                        finish = succ.late_finish - timedelta(days=edge_data['lag'])
                    elif edge_data['rel_type'] == RelationshipType.SF:
                        finish = succ.late_finish + timedelta(days=activity.duration - 1 + edge_data['lag'])
                    
                    if finish < self.start_date:
                        finish = self.calendar.calculate_working_days(self.start_date, activity.duration)
                    
                    if latest_finish is None or finish < latest_finish:
                        latest_finish = finish
                
                activity.late_finish = latest_finish
            
            # Calculate late start from late finish
            activity.late_start = activity.late_finish - timedelta(days=activity.duration - 1)
            
            if activity.late_start < self.start_date:
                activity.late_start = self.start_date
                activity.late_finish = self.calendar.calculate_working_days(activity.late_start, activity.duration)
            
            activity.slack = (activity.late_start - activity.early_start).days

        # Ensure critical activities have early and late dates equal
        self.critical_path = self.find_critical_path(ordered_activities)
        
        # Force critical activities to have early=late dates
        for activity_id in self.critical_path:
            activity = self.activities[activity_id]
            activity.late_start = activity.early_start
            activity.late_finish = activity.early_finish
            activity.slack = 0

    def find_critical_path(self, ordered_activities):
        """Find at least one critical path from earliest to latest activity"""
        if not ordered_activities:
            return []
        
        # First identify all critical activities (slack = 0)
        critical_activities = [activity_id for activity_id in ordered_activities 
                             if self.activities[activity_id].slack == 0]
        
        if not critical_activities:
            # If no activities with zero slack, find the path with minimum total slack
            return self.find_minimum_slack_path(ordered_activities)
        
        # Find the earliest and latest activities
        first_activity_id = ordered_activities[0]
        last_activity_id = ordered_activities[-1]
        
        # If both first and last are critical, find a critical path connecting them
        if first_activity_id in critical_activities and last_activity_id in critical_activities:
            path = self.find_path_between_critical(first_activity_id, last_activity_id, critical_activities)
            if path:
                return path
        
        # Find any critical path from first to last through critical activities
        return self.find_critical_path_through_network(ordered_activities, critical_activities)
    
    def find_path_between_critical(self, start_id, end_id, critical_activities):
        """Find a path from start to end using only critical activities"""
        # BFS to find path through critical activities
        from collections import deque
        
        queue = deque([(start_id, [start_id])])
        visited = {start_id}
        
        while queue:
            current, path = queue.popleft()
            
            if current == end_id:
                return path
            
            # Check successors that are critical
            for successor in self.graph.successors(current):
                if successor in critical_activities and successor not in visited:
                    visited.add(successor)
                    queue.append((successor, path + [successor]))
        
        return None
    
    def find_critical_path_through_network(self, ordered_activities, critical_activities):
        """Find the longest path through critical activities"""
        if not critical_activities:
            return ordered_activities[:1] if ordered_activities else []
        
        # Start from first activity
        first_activity = ordered_activities[0]
        last_activity = ordered_activities[-1]
        
        # Try to find longest path through critical activities
        def find_longest_path(current, visited, path):
            if current == last_activity:
                # Count critical activities in path
                critical_count = sum(1 for act_id in path if act_id in critical_activities)
                return path, critical_count
            
            best_path = path
            best_critical_count = sum(1 for act_id in path if act_id in critical_activities)
            
            for successor in self.graph.successors(current):
                if successor not in visited:
                    new_visited = visited | {successor}
                    new_path, critical_count = find_longest_path(successor, new_visited, path + [successor])
                    if critical_count > best_critical_count or (critical_count == best_critical_count and len(new_path) > len(best_path)):
                        best_path = new_path
                        best_critical_count = critical_count
            
            return best_path, best_critical_count
        
        best_path, _ = find_longest_path(first_activity, {first_activity}, [first_activity])
        
        # Ensure all activities in the path are marked as critical
        for act_id in best_path:
            self.activities[act_id].slack = 0
        
        return best_path if best_path else critical_activities
    
    def find_minimum_slack_path(self, ordered_activities):
        """Find path with minimum total slack from first to last activity"""
        if not ordered_activities:
            return []
        
        first_activity = ordered_activities[0]
        last_activity = ordered_activities[-1]
        
        def find_min_slack_path(current, visited, path, total_slack):
            if current == last_activity:
                return path, total_slack
            
            best_path = None
            best_slack = float('inf')
            
            for successor in self.graph.successors(current):
                if successor not in visited:
                    new_visited = visited | {successor}
                    succ_slack = self.activities[successor].slack if self.activities[successor].slack is not None else 0
                    new_path, new_slack = find_min_slack_path(
                        successor, new_visited, path + [successor], total_slack + succ_slack
                    )
                    if new_slack < best_slack:
                        best_path = new_path
                        best_slack = new_slack
            
            if best_path is None:
                return path, total_slack
            
            return best_path, best_slack
        
        first_slack = self.activities[first_activity].slack if self.activities[first_activity].slack is not None else 0
        best_path, _ = find_min_slack_path(first_activity, {first_activity}, [first_activity], first_slack)
        
        # Set slack to 0 for all activities in this path
        for act_id in best_path:
            self.activities[act_id].slack = 0
        
        return best_path

    def create_gantt_figure(self):
        if not self.activities:
            return None

        # Get translations
        translations = self.get_translations()
        
        fig, ax = plt.subplots(figsize=(14, max(8, len(self.activities) * 0.5)))
        
        # Configure fonts for Arabic
        if self.language == 'ar':
            plt.rcParams['font.family'] = 'Arial'
            plt.rcParams['axes.unicode_minus'] = False

        colors = plt.cm.tab20.colors
        date_to_mpl = mdates.date2num

        critical_color = '#d62728'
        non_critical_color = '#1f77b4'
        relation_colors = {
            RelationshipType.FS: '#2ca02c',
            RelationshipType.SS: '#ff7f0e',
            RelationshipType.FF: '#9467bd',
            RelationshipType.SF: '#e377c2'
        }

        activity_names = [activity.name for activity in self.activities.values()]
        y_positions = {name: i for i, name in enumerate(activity_names)}

        for i, (activity_id, activity) in enumerate(self.activities.items()):
            is_critical = activity_id in self.critical_path
            color = critical_color if is_critical else non_critical_color

            ax.barh(
                y=y_positions[activity.name],
                width=activity.duration,
                left=date_to_mpl(activity.early_start),
                color=color,
                edgecolor='black',
                alpha=0.8,
                height=0.6
            )

            info = f"{translations['id']}: {activity.id}\n{translations['duration']}: {activity.duration}{translations['days']}\n{translations['slack']}: {activity.slack}{translations['days']}"
            ax.text(
                x=date_to_mpl(activity.early_start) + activity.duration/2,
                y=y_positions[activity.name],
                s=info,
                ha='center',
                va='center',
                color='white' if is_critical else 'black',
                fontsize=8
            )

            for dep in activity.dependencies:
                pred = self.activities[dep.from_activity_id]
                rel_type = dep.rel_type
                lag = dep.lag
                color = relation_colors[rel_type]
                
                if rel_type == RelationshipType.FS:
                    start_x = date_to_mpl(pred.early_finish)
                    end_x = date_to_mpl(activity.early_start) - lag
                    ax.arrow(
                        x=start_x,
                        y=y_positions[pred.name],
                        dx=end_x - start_x,
                        dy=y_positions[activity.name] - y_positions[pred.name],
                        head_width=0.15,
                        head_length=0.5,
                        fc=color,
                        ec=color,
                        linestyle='--',
                        alpha=0.7
                    )

        ax.set_yticks(range(len(activity_names)))
        ax.set_yticklabels(activity_names)
        ax.xaxis_date()
        ax.xaxis.set_major_locator(mdates.WeekdayLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        fig.autofmt_xdate()
        ax.set_xlabel(translations['timeline'])
        ax.set_ylabel(translations['activities'])
        
        title = f"{translations['gantt_title']}: {self.name}\n({translations['showing_deps']})\n"
        title += 'This Software was created by DR. Eng. Wael Sherif Selim, email: wael.sherif.selim@gmail.com'
        ax.set_title(title, fontsize=10)
        
        ax.grid(True, axis='x', linestyle='--', alpha=0.6)
        
        legend_elements = [
            Patch(color=critical_color, label=translations['critical_activity']),
            Patch(color=non_critical_color, label=translations['non_critical_activity']),
            Patch(color=relation_colors[RelationshipType.FS], label=translations['fs']),
            Patch(color=relation_colors[RelationshipType.SS], label=translations['ss']),
            Patch(color=relation_colors[RelationshipType.FF], label=translations['ff']),
            Patch(color=relation_colors[RelationshipType.SF], label=translations['sf'])
        ]
        ax.legend(handles=legend_elements, bbox_to_anchor=(1.05, 1), loc='upper left')
        
        plt.tight_layout()
        return fig

    def get_translations(self):
        """Get UI translations based on selected language"""
        if self.language == 'ar':
            return {
                'gantt_title': 'مخطط جانت للمشروع',
                'showing_deps': 'عرض التبعيات مع التأخيرات',
                'timeline': 'الجدول الزمني',
                'activities': 'الأنشطة',
                'id': 'الرقم',
                'duration': 'المدة',
                'days': 'د',
                'slack': 'الفائض',
                'critical_activity': 'نشاط حرج',
                'non_critical_activity': 'نشاط غير حرج',
                'fs': 'نهاية-إلى-بداية',
                'ss': 'بداية-إلى-بداية',
                'ff': 'نهاية-إلى-نهاية',
                'sf': 'بداية-إلى-نهاية'
            }
        else:
            return {
                'gantt_title': 'Project Gantt Chart',
                'showing_deps': 'Showing dependencies with lags',
                'timeline': 'Timeline',
                'activities': 'Activities',
                'id': 'ID',
                'duration': 'Dur',
                'days': 'd',
                'slack': 'Slack',
                'critical_activity': 'Critical Activity',
                'non_critical_activity': 'Non-Critical Activity',
                'fs': 'Finish-to-Start (FS)',
                'ss': 'Start-to-Start (SS)',
                'ff': 'Finish-to-Finish (FF)',
                'sf': 'Start-to-Finish (SF)'
            }

    def export_to_excel(self, filename):
        """Export Gantt chart data to Excel format with visual Gantt chart"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Schedule"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        critical_bar_fill = PatternFill(start_color="D62728", end_color="D62728", fill_type="solid")
        non_critical_bar_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header
        headers = ['Activity ID', 'Activity Name', 'Duration (days)', 'Early Start', 'Early Finish', 
                   'Late Start', 'Late Finish', 'Total Slack', 'Critical?', 'Predecessors']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write activity data
        row = 2
        for activity_id in sorted(self.activities.keys()):
            activity = self.activities[activity_id]
            
            # Get predecessors
            predecessors = []
            for dep in activity.dependencies:
                pred_str = f"{dep.from_activity_id} ({dep.rel_type.name}"
                if dep.lag != 0:
                    pred_str += f"+{dep.lag}" if dep.lag > 0 else f"{dep.lag}"
                pred_str += ")"
                predecessors.append(pred_str)
            pred_text = ", ".join(predecessors) if predecessors else "None"
            
            is_critical = activity_id in self.critical_path
            
            data = [
                activity.id,
                activity.name,
                activity.duration,
                activity.early_start.strftime('%Y-%m-%d') if activity.early_start else '',
                activity.early_finish.strftime('%Y-%m-%d') if activity.early_finish else '',
                activity.late_start.strftime('%Y-%m-%d') if activity.late_start else '',
                activity.late_finish.strftime('%Y-%m-%d') if activity.late_finish else '',
                activity.slack if activity.slack is not None else '',
                'Yes' if is_critical else 'No',
                pred_text
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if is_critical and col <= 8:
                    cell.fill = critical_fill
                    cell.font = Font(bold=True, color="FFFFFF")
            
            row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Create Gantt Chart sheet
        ws_gantt = wb.create_sheet("Gantt Chart")
        
        # Find project date range
        if self.activities:
            min_date = min(activity.early_start for activity in self.activities.values() if activity.early_start)
            max_date = max(activity.early_finish for activity in self.activities.values() if activity.early_finish)
            total_days = (max_date - min_date).days + 1
            
            # Gantt chart headers
            ws_gantt.cell(1, 1, "Activity ID").fill = header_fill
            ws_gantt.cell(1, 1).font = header_font
            ws_gantt.cell(1, 1).border = border
            ws_gantt.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            ws_gantt.cell(1, 2, "Activity Name").fill = header_fill
            ws_gantt.cell(1, 2).font = header_font
            ws_gantt.cell(1, 2).border = border
            ws_gantt.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
            
            ws_gantt.cell(1, 3, "Duration").fill = header_fill
            ws_gantt.cell(1, 3).font = header_font
            ws_gantt.cell(1, 3).border = border
            ws_gantt.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
            
            # Date headers
            current_date = min_date
            col_offset = 4
            for day in range(total_days):
                col = col_offset + day
                cell = ws_gantt.cell(1, col, current_date.strftime('%m/%d'))
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF", size=8)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
                ws_gantt.column_dimensions[get_column_letter(col)].width = 3
                current_date += timedelta(days=1)
            
            # Set column widths for activity info
            ws_gantt.column_dimensions['A'].width = 10
            ws_gantt.column_dimensions['B'].width = 25
            ws_gantt.column_dimensions['C'].width = 10
            
            # Add activities and visual bars
            row = 2
            for activity_id in sorted(self.activities.keys()):
                activity = self.activities[activity_id]
                is_critical = activity_id in self.critical_path
                
                # Activity info
                ws_gantt.cell(row, 1, activity.id).border = border
                ws_gantt.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws_gantt.cell(row, 2, activity.name).border = border
                ws_gantt.cell(row, 2).alignment = Alignment(horizontal='left', vertical='center')
                
                ws_gantt.cell(row, 3, activity.duration).border = border
                ws_gantt.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')
                
                if is_critical:
                    ws_gantt.cell(row, 1).fill = critical_fill
                    ws_gantt.cell(row, 1).font = Font(bold=True, color="FFFFFF")
                    ws_gantt.cell(row, 2).fill = critical_fill
                    ws_gantt.cell(row, 2).font = Font(bold=True, color="FFFFFF")
                    ws_gantt.cell(row, 3).fill = critical_fill
                    ws_gantt.cell(row, 3).font = Font(bold=True, color="FFFFFF")
                
                # Draw Gantt bars
                if activity.early_start and activity.early_finish:
                    start_day = (activity.early_start - min_date).days
                    duration = activity.duration
                    
                    bar_fill = critical_bar_fill if is_critical else non_critical_bar_fill
                    
                    for day in range(duration):
                        col = col_offset + start_day + day
                        cell = ws_gantt.cell(row, col)
                        cell.fill = bar_fill
                        cell.border = border
                        
                        # Add activity ID in the middle of the bar
                        if day == duration // 2:
                            cell.value = activity.id
                            cell.font = Font(bold=True, color="FFFFFF", size=8)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add legend
            legend_row = row + 2
            ws_gantt.cell(legend_row, 1, "Legend:").font = Font(bold=True, size=11)
            
            legend_row += 1
            ws_gantt.cell(legend_row, 1, "Critical Activity")
            ws_gantt.cell(legend_row, 2).fill = critical_bar_fill
            ws_gantt.cell(legend_row, 2).border = border
            
            legend_row += 1
            ws_gantt.cell(legend_row, 1, "Non-Critical Activity")
            ws_gantt.cell(legend_row, 2).fill = non_critical_bar_fill
            ws_gantt.cell(legend_row, 2).border = border
        
        # Add project summary
        ws2 = wb.create_sheet("Project Summary")
        ws2['A1'] = "Project Name:"
        ws2['B1'] = self.name
        ws2['A1'].font = Font(bold=True)
        
        ws2['A2'] = "Start Date:"
        ws2['B2'] = self.start_date.strftime('%Y-%m-%d')
        ws2['A2'].font = Font(bold=True)
        
        ws2['A3'] = "Total Activities:"
        ws2['B3'] = len(self.activities)
        ws2['A3'].font = Font(bold=True)
        
        ws2['A4'] = "Critical Path:"
        ws2['B4'] = ", ".join(self.critical_path)
        ws2['A4'].font = Font(bold=True)
        
        if self.activities:
            end_date = max(activity.early_finish for activity in self.activities.values())
            ws2['A5'] = "Project Duration:"
            ws2['B5'] = f"{(end_date - self.start_date).days + 1} days"
            ws2['A5'].font = Font(bold=True)
            
            ws2['A6'] = "End Date:"
            ws2['B6'] = end_date.strftime('%Y-%m-%d')
            ws2['A6'].font = Font(bold=True)
        
        # Add attribution
        ws2['A8'] = "Software Created By:"
        ws2['B8'] = "DR. Eng. Wael Sherif Selim"
        ws2['A8'].font = Font(bold=True)
        
        ws2['A9'] = "Email:"
        ws2['B9'] = "wael.sherif.selim@gmail.com"
        ws2['A9'].font = Font(bold=True)
        
        # Adjust column widths in summary
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 40
        
        wb.save(filename)

    def export_to_primavera(self, filename):
        """Export project to Primavera P6 XML format (XER-like XML)"""
        
        # Create root element
        root = ET.Element('Project')
        root.set('name', self.name)
        root.set('start_date', self.start_date.strftime('%Y-%m-%d'))
        
        # Add project info
        info = ET.SubElement(root, 'ProjectInfo')
        ET.SubElement(info, 'ProjectName').text = self.name
        ET.SubElement(info, 'StartDate').text = self.start_date.strftime('%Y-%m-%d')
        ET.SubElement(info, 'TotalActivities').text = str(len(self.activities))
        ET.SubElement(info, 'CreatedBy').text = 'DR. Eng. Wael Sherif Selim (wael.sherif.selim@gmail.com)'
        
        # Add activities
        activities_elem = ET.SubElement(root, 'Activities')
        
        for activity_id in sorted(self.activities.keys()):
            activity = self.activities[activity_id]
            
            activity_elem = ET.SubElement(activities_elem, 'Activity')
            activity_elem.set('id', str(activity.id))
            activity_elem.set('critical', 'true' if activity_id in self.critical_path else 'false')
            
            ET.SubElement(activity_elem, 'Name').text = activity.name
            ET.SubElement(activity_elem, 'Duration').text = str(activity.duration)
            ET.SubElement(activity_elem, 'EarlyStart').text = activity.early_start.strftime('%Y-%m-%d') if activity.early_start else ''
            ET.SubElement(activity_elem, 'EarlyFinish').text = activity.early_finish.strftime('%Y-%m-%d') if activity.early_finish else ''
            ET.SubElement(activity_elem, 'LateStart').text = activity.late_start.strftime('%Y-%m-%d') if activity.late_start else ''
            ET.SubElement(activity_elem, 'LateFinish').text = activity.late_finish.strftime('%Y-%m-%d') if activity.late_finish else ''
            ET.SubElement(activity_elem, 'TotalSlack').text = str(activity.slack) if activity.slack is not None else ''
            
            # Add dependencies
            if activity.dependencies:
                dependencies_elem = ET.SubElement(activity_elem, 'Predecessors')
                for dep in activity.dependencies:
                    dep_elem = ET.SubElement(dependencies_elem, 'Predecessor')
                    dep_elem.set('id', str(dep.from_activity_id))
                    dep_elem.set('type', dep.rel_type.name)
                    dep_elem.set('lag', str(dep.lag))
        
        # Add critical path
        critical_path_elem = ET.SubElement(root, 'CriticalPath')
        for activity_id in self.critical_path:
            ET.SubElement(critical_path_elem, 'Activity').text = str(activity_id)
        
        # Create XML tree and write to file
        xml_str = ET.tostring(root, encoding='utf-8')
        dom = minidom.parseString(xml_str)
        pretty_xml = dom.toprettyxml(indent='  ')
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)

    def save_as_png(self, filename):
        """Save Gantt chart as PNG image"""
        if not self.activities:
            return False
        
        try:
            fig = self.create_gantt_figure()
            if fig:
                fig.savefig(filename, dpi=300, bbox_inches='tight', format='png')
                plt.close(fig)
                return True
        except Exception as e:
            print(f"Error saving PNG: {e}")
        
        return False

class CalendarDialog:
    def __init__(self, parent, calendar):
        self.parent = parent
        self.calendar = calendar
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Calendar Configuration")
        self.dialog.geometry("500x600")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.setup_ui()
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.dialog, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Work week configuration
        work_week_frame = ttk.LabelFrame(main_frame, text="Work Week Configuration", padding=10)
        work_week_frame.pack(fill=tk.X, pady=(0, 10))
        
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.day_vars = []
        
        for i, day in enumerate(days):
            var = tk.BooleanVar(value=self.calendar.work_week[i])
            self.day_vars.append(var)
            cb = ttk.Checkbutton(work_week_frame, text=day, variable=var)
            cb.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
        
        # Holidays configuration
        holidays_frame = ttk.LabelFrame(main_frame, text="Holidays", padding=10)
        holidays_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Add holiday section
        add_holiday_frame = ttk.Frame(holidays_frame)
        add_holiday_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(add_holiday_frame, text="Holiday Date (YYYY-MM-DD):").pack(side=tk.LEFT, padx=(0, 5))
        self.holiday_date_var = tk.StringVar()
        holiday_entry = ttk.Entry(add_holiday_frame, textvariable=self.holiday_date_var, width=12)
        holiday_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(add_holiday_frame, text="Add Holiday", command=self.add_holiday).pack(side=tk.LEFT)
        
        # Holidays list
        list_frame = ttk.Frame(holidays_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.holidays_listbox = tk.Listbox(list_frame, height=8)
        holidays_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.holidays_listbox.yview)
        self.holidays_listbox.configure(yscrollcommand=holidays_scrollbar.set)
        
        self.holidays_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        holidays_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Holidays buttons
        holiday_buttons_frame = ttk.Frame(holidays_frame)
        holiday_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(holiday_buttons_frame, text="Remove Selected", command=self.remove_holiday).pack(side=tk.LEFT, padx=(0, 5))
        
        # Load existing holidays
        self.load_holidays()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Apply", command=self.apply_calendar).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.RIGHT)
    
    def add_holiday(self):
        holiday_str = self.holiday_date_var.get().strip()
        if not holiday_str:
            messagebox.showerror("Error", "Please enter a holiday date")
            return
        
        try:
            holiday_date = datetime.strptime(holiday_str, "%Y-%m-%d").date()
            self.holidays_listbox.insert(tk.END, holiday_str)
            self.holiday_date_var.set("")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD")
    
    def remove_holiday(self):
        selected = self.holidays_listbox.curselection()
        if selected:
            self.holidays_listbox.delete(selected[0])
    
    def load_holidays(self):
        for holiday in self.calendar.holidays:
            if isinstance(holiday, date):
                self.holidays_listbox.insert(tk.END, holiday.strftime('%Y-%m-%d'))
            else:
                self.holidays_listbox.insert(tk.END, holiday)
    
    def apply_calendar(self):
        # Update work week
        for i, var in enumerate(self.day_vars):
            self.calendar.work_week[i] = var.get()
        
        # Update holidays
        self.calendar.holidays = []
        for i in range(self.holidays_listbox.size()):
            holiday_str = self.holidays_listbox.get(i)
            try:
                holiday_date = datetime.strptime(holiday_str, "%Y-%m-%d").date()
                self.calendar.holidays.append(holiday_date)
            except ValueError:
                continue
        
        messagebox.showinfo("Success", "Calendar settings applied successfully!")
        self.dialog.destroy()

class GanttChartApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Gantt Chart Project Planner")
        self.root.geometry("1200x800")
        
        # Check license first
        check_license()
        
        # Initialize project with current date
        self.project = Project("New Project", datetime.now())
        
        # Initialize clipboards
        self.activity_clipboard = None
        self.dependency_clipboard = None
        
        # Set up UI
        self.setup_ui()
        self.update_activity_list()
        self.update_dependencies_list()

    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel - Inputs
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        
        # Project info
        project_frame = ttk.LabelFrame(left_frame, text="Project Information", padding=10)
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(project_frame, text="Project Name:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.project_name_var = tk.StringVar(value="New Project")
        ttk.Entry(project_frame, textvariable=self.project_name_var).grid(row=0, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(project_frame, text="Start Date:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.start_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        ttk.Entry(project_frame, textvariable=self.start_date_var).grid(row=1, column=1, sticky=tk.EW, pady=2)
        
        project_frame.columnconfigure(1, weight=1)
        
        # Language selection
        lang_frame = ttk.Frame(project_frame)
        lang_frame.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=5)
        ttk.Label(lang_frame, text="Language:").pack(side=tk.LEFT)
        self.language_var = tk.StringVar(value="en")
        ttk.Radiobutton(lang_frame, text="English", variable=self.language_var, value="en", 
                       command=self.update_language).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Radiobutton(lang_frame, text="Arabic", variable=self.language_var, value="ar", 
                       command=self.update_language).pack(side=tk.LEFT, padx=(5, 10))
        
        # Calendar button
        ttk.Button(project_frame, text="Configure Calendar", command=self.configure_calendar).grid(row=3, column=0, columnspan=2, pady=5)
        
        # Activity management
        activity_frame = ttk.LabelFrame(left_frame, text="Activity Management", padding=10)
        activity_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(activity_frame, text="Activity ID:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.activity_id_var = tk.StringVar()
        ttk.Entry(activity_frame, textvariable=self.activity_id_var).grid(row=0, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(activity_frame, text="Activity Name:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.activity_name_var = tk.StringVar()
        ttk.Entry(activity_frame, textvariable=self.activity_name_var).grid(row=1, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(activity_frame, text="Duration (days):").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.duration_var = tk.StringVar()
        ttk.Entry(activity_frame, textvariable=self.duration_var).grid(row=2, column=1, sticky=tk.EW, pady=2)
        
        button_frame = ttk.Frame(activity_frame)
        button_frame.grid(row=3, column=0, columnspan=2, sticky=tk.EW, pady=5)
        
        ttk.Button(button_frame, text="Add Activity", command=self.add_activity).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Modify Activity", command=self.modify_activity).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Copy Activity", command=self.copy_activity).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Paste Activity", command=self.paste_activity).pack(side=tk.LEFT, padx=5)
        
        activity_frame.columnconfigure(1, weight=1)
        
        # Dependencies management
        dep_frame = ttk.LabelFrame(left_frame, text="Dependencies Management", padding=10)
        dep_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(dep_frame, text="From Activity:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.from_activity_var = tk.StringVar()
        self.from_activity_combo = ttk.Combobox(dep_frame, textvariable=self.from_activity_var, state="readonly")
        self.from_activity_combo.grid(row=0, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(dep_frame, text="To Activity:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.to_activity_var = tk.StringVar()
        self.to_activity_combo = ttk.Combobox(dep_frame, textvariable=self.to_activity_var, state="readonly")
        self.to_activity_combo.grid(row=1, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(dep_frame, text="Relationship Type:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.rel_type_var = tk.StringVar(value="FS")
        rel_combo = ttk.Combobox(dep_frame, textvariable=self.rel_type_var, 
                                values=["FS", "SS", "FF", "SF"], state="readonly")
        rel_combo.grid(row=2, column=1, sticky=tk.EW, pady=2)
        
        ttk.Label(dep_frame, text="Lag (days):").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.lag_var = tk.StringVar(value="0")
        ttk.Entry(dep_frame, textvariable=self.lag_var).grid(row=3, column=1, sticky=tk.EW, pady=2)
        
        dep_button_frame = ttk.Frame(dep_frame)
        dep_button_frame.grid(row=4, column=0, columnspan=2, sticky=tk.EW, pady=5)
        
        ttk.Button(dep_button_frame, text="Add Dependency", command=self.add_dependency).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(dep_button_frame, text="Modify Dependency", command=self.modify_dependency).pack(side=tk.LEFT, padx=5)
        ttk.Button(dep_button_frame, text="Copy Dependency", command=self.copy_dependency).pack(side=tk.LEFT, padx=5)
        ttk.Button(dep_button_frame, text="Paste Dependency", command=self.paste_dependency).pack(side=tk.LEFT, padx=5)
        
        dep_frame.columnconfigure(1, weight=1)
        
        # Control buttons
        control_frame = ttk.Frame(left_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(control_frame, text="Calculate Schedule", command=self.calculate_schedule).pack(fill=tk.X, pady=2)
        ttk.Button(control_frame, text="Generate Gantt Chart", command=self.generate_gantt).pack(fill=tk.X, pady=2)
        
        # Export buttons
        export_frame = ttk.LabelFrame(left_frame, text="Export Options", padding=10)
        export_frame.pack(fill=tk.X)
        
        ttk.Button(export_frame, text="Export to Excel", command=self.export_to_excel).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="Export to Primavera XML", command=self.export_to_primavera).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="Save as PNG", command=self.save_as_png).pack(fill=tk.X, pady=2)
        
        # Right panel - Lists and display
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Activities list
        activities_list_frame = ttk.LabelFrame(right_frame, text="Activities List", padding=10)
        activities_list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.activities_tree = ttk.Treeview(activities_list_frame, columns=('ID', 'Name', 'Duration'), show='headings')
        self.activities_tree.heading('ID', text='ID')
        self.activities_tree.heading('Name', text='Name')
        self.activities_tree.heading('Duration', text='Duration (days)')
        self.activities_tree.column('ID', width=80)
        self.activities_tree.column('Name', width=150)
        self.activities_tree.column('Duration', width=100)
        
        scrollbar_activities = ttk.Scrollbar(activities_list_frame, orient=tk.VERTICAL, command=self.activities_tree.yview)
        self.activities_tree.configure(yscrollcommand=scrollbar_activities.set)
        
        self.activities_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_activities.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Activities buttons
        activity_buttons_frame = ttk.Frame(activities_list_frame)
        activity_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(activity_buttons_frame, text="Delete Selected Activity", command=self.delete_activity).pack(side=tk.LEFT, padx=(0, 5))
        
        # Dependencies list
        deps_list_frame = ttk.LabelFrame(right_frame, text="Dependencies List", padding=10)
        deps_list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.deps_tree = ttk.Treeview(deps_list_frame, columns=('From', 'To', 'Type', 'Lag'), show='headings')
        self.deps_tree.heading('From', text='From Activity')
        self.deps_tree.heading('To', text='To Activity')
        self.deps_tree.heading('Type', text='Relationship Type')
        self.deps_tree.heading('Lag', text='Lag (days)')
        self.deps_tree.column('From', width=100)
        self.deps_tree.column('To', width=100)
        self.deps_tree.column('Type', width=120)
        self.deps_tree.column('Lag', width=80)
        
        scrollbar_deps = ttk.Scrollbar(deps_list_frame, orient=tk.VERTICAL, command=self.deps_tree.yview)
        self.deps_tree.configure(yscrollcommand=scrollbar_deps.set)
        
        self.deps_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_deps.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Dependencies buttons
        dep_buttons_frame = ttk.Frame(deps_list_frame)
        dep_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(dep_buttons_frame, text="Delete Selected Dependency", command=self.delete_dependency).pack(side=tk.LEFT, padx=(0, 5))
        
        # Bind double-click events
        self.activities_tree.bind('<Double-1>', self.on_activity_double_click)
        self.deps_tree.bind('<Double-1>', self.on_dependency_double_click)

    def configure_calendar(self):
        """Open calendar configuration dialog"""
        CalendarDialog(self.root, self.project.calendar)

    def update_language(self):
        """Update project language"""
        self.project.language = self.language_var.get()

    def update_activity_combos(self):
        """Update activity combo boxes"""
        activities = [f"{activity_id} - {activity.name}" for activity_id, activity in self.project.activities.items()]
        self.from_activity_combo['values'] = activities
        self.to_activity_combo['values'] = activities

    def extract_activity_id(self, combo_value):
        """Extract activity ID from combo box value"""
        if combo_value and ' - ' in combo_value:
            return combo_value.split(' - ')[0]
        return combo_value

    def add_activity(self):
        try:
            activity_id = self.activity_id_var.get().strip()
            name = self.activity_name_var.get().strip()
            duration = int(self.duration_var.get())
            
            if not activity_id or not name:
                messagebox.showerror("Error", "Activity ID and Name are required")
                return
            
            if activity_id in self.project.activities:
                messagebox.showerror("Error", f"Activity ID '{activity_id}' already exists")
                return
            
            activity = Activity(activity_id, name, duration)
            self.project.add_activity(activity)
            
            self.update_activity_list()
            self.update_activity_combos()
            self.clear_activity_fields()
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for duration")

    def modify_activity(self):
        selected = self.activities_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an activity to modify")
            return
        
        try:
            activity_id = self.activities_tree.item(selected[0])['values'][0]
            name = self.activity_name_var.get().strip()
            duration = int(self.duration_var.get())
            
            if not name:
                messagebox.showerror("Error", "Activity Name is required")
                return
            
            if activity_id in self.project.activities:
                activity = self.project.activities[activity_id]
                activity.name = name
                activity.duration = duration
                
                self.update_activity_list()
                self.update_activity_combos()
                self.clear_activity_fields()
                messagebox.showinfo("Success", f"Activity '{activity_id}' modified successfully")
                
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for duration")

    def copy_activity(self):
        selected = self.activities_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an activity to copy")
            return
        
        activity_id = self.activities_tree.item(selected[0])['values'][0]
        if activity_id in self.project.activities:
            activity = self.project.activities[activity_id]
            self.activity_clipboard = {
                'name': activity.name,
                'duration': activity.duration
            }
            messagebox.showinfo("Success", f"Activity '{activity_id}' copied to clipboard")

    def paste_activity(self):
        if not self.activity_clipboard:
            messagebox.showwarning("Warning", "No activity in clipboard to paste")
            return
        
        try:
            # Generate new activity ID
            existing_ids = [int(act_id[1:]) for act_id in self.project.activities.keys() if act_id.startswith('A') and act_id[1:].isdigit()]
            new_id = f"A{max(existing_ids) + 1}" if existing_ids else "A1"
            
            activity = Activity(new_id, self.activity_clipboard['name'], self.activity_clipboard['duration'])
            self.project.add_activity(activity)
            
            self.update_activity_list()
            self.update_activity_combos()
            messagebox.showinfo("Success", f"Activity '{new_id}' pasted successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to paste activity: {str(e)}")

    def delete_activity(self):
        selected = self.activities_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an activity to delete")
            return
        
        activity_id = self.activities_tree.item(selected[0])['values'][0]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete activity '{activity_id}' and all its dependencies?"):
            if self.project.remove_activity(activity_id):
                self.update_activity_list()
                self.update_dependencies_list()
                self.update_activity_combos()
                messagebox.showinfo("Success", f"Activity '{activity_id}' deleted successfully")
            else:
                messagebox.showerror("Error", f"Failed to delete activity '{activity_id}'")

    def on_activity_double_click(self, event):
        selected = self.activities_tree.selection()
        if selected:
            values = self.activities_tree.item(selected[0])['values']
            self.activity_id_var.set(values[0])
            self.activity_name_var.set(values[1])
            self.duration_var.set(str(values[2]))

    def add_dependency(self):
        try:
            from_activity_combo = self.from_activity_var.get()
            to_activity_combo = self.to_activity_var.get()
            rel_type_str = self.rel_type_var.get()
            lag = int(self.lag_var.get())
            
            from_activity = self.extract_activity_id(from_activity_combo)
            to_activity = self.extract_activity_id(to_activity_combo)
            
            if not from_activity or not to_activity:
                messagebox.showerror("Error", "Please select both from and to activities")
                return
            
            if from_activity == to_activity:
                messagebox.showerror("Error", "Cannot create dependency from an activity to itself")
                return
            
            rel_type = RelationshipType[rel_type_str]
            self.project.add_dependency(from_activity, to_activity, rel_type, lag)
            
            self.update_dependencies_list()
            self.clear_dependency_fields()
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for lag")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def modify_dependency(self):
        selected = self.deps_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a dependency to modify")
            return
        
        try:
            # Get the original dependency details
            item = self.deps_tree.item(selected[0])
            original_from = item['values'][0]
            original_to = item['values'][1]
            
            # Get new values
            from_activity_combo = self.from_activity_var.get()
            to_activity_combo = self.to_activity_var.get()
            rel_type_str = self.rel_type_var.get()
            lag = int(self.lag_var.get())
            
            from_activity = self.extract_activity_id(from_activity_combo)
            to_activity = self.extract_activity_id(to_activity_combo)
            
            if not from_activity or not to_activity:
                messagebox.showerror("Error", "Please select both from and to activities")
                return
            
            if from_activity == to_activity:
                messagebox.showerror("Error", "Cannot create dependency from an activity to itself")
                return
            
            # Remove old dependency and add new one
            self.project.remove_dependency(original_from, original_to)
            rel_type = RelationshipType[rel_type_str]
            self.project.add_dependency(from_activity, to_activity, rel_type, lag)
            
            self.update_dependencies_list()
            self.clear_dependency_fields()
            messagebox.showinfo("Success", "Dependency modified successfully")
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for lag")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def copy_dependency(self):
        selected = self.deps_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a dependency to copy")
            return
        
        item = self.deps_tree.item(selected[0])
        self.dependency_clipboard = {
            'from_activity': item['values'][0],
            'to_activity': item['values'][1],
            'rel_type': item['values'][2],
            'lag': item['values'][3]
        }
        messagebox.showinfo("Success", "Dependency copied to clipboard")

    def paste_dependency(self):
        if not self.dependency_clipboard:
            messagebox.showwarning("Warning", "No dependency in clipboard to paste")
            return
        
        try:
            # Generate new activity IDs if needed (for demo purposes)
            from_activity = self.dependency_clipboard['from_activity']
            to_activity = self.dependency_clipboard['to_activity']
            
            # Check if activities exist
            if from_activity not in self.project.activities or to_activity not in self.project.activities:
                messagebox.showerror("Error", "One or both activities do not exist")
                return
            
            # Find relationship type
            rel_type = None
            for rt in RelationshipType:
                if rt.value == self.dependency_clipboard['rel_type']:
                    rel_type = rt
                    break
            
            if not rel_type:
                messagebox.showerror("Error", "Invalid relationship type in clipboard")
                return
            
            lag = int(self.dependency_clipboard['lag'])
            
            self.project.add_dependency(from_activity, to_activity, rel_type, lag)
            self.update_dependencies_list()
            messagebox.showinfo("Success", "Dependency pasted successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to paste dependency: {str(e)}")

    def delete_dependency(self):
        selected = self.deps_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a dependency to delete")
            return
        
        item = self.deps_tree.item(selected[0])
        from_activity = item['values'][0]
        to_activity = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete dependency from '{from_activity}' to '{to_activity}'?"):
            if self.project.remove_dependency(from_activity, to_activity):
                self.update_dependencies_list()
                messagebox.showinfo("Success", "Dependency deleted successfully")
            else:
                messagebox.showerror("Error", "Failed to delete dependency")

    def on_dependency_double_click(self, event):
        selected = self.deps_tree.selection()
        if selected:
            item = self.deps_tree.item(selected[0])
            from_activity = item['values'][0]
            to_activity = item['values'][1]
            rel_type = item['values'][2]
            lag = item['values'][3]
            
            # Set values in combo boxes
            from_display = f"{from_activity} - {self.project.activities[from_activity].name}"
            to_display = f"{to_activity} - {self.project.activities[to_activity].name}"
            
            self.from_activity_var.set(from_display)
            self.to_activity_var.set(to_display)
            self.rel_type_var.set(rel_type.split(' ')[0])  # Get FS, SS, FF, SF from full name
            self.lag_var.set(str(lag))

    def calculate_schedule(self):
        try:
            # Update project info
            self.project.name = self.project_name_var.get()
            start_date_str = self.start_date_var.get()
            try:
                self.project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Error", "Invalid start date format. Please use YYYY-MM-DD")
                return
            
            self.project.calculate_schedule()
            messagebox.showinfo("Success", "Schedule calculated successfully!")
            
            # Show critical path
            critical_path = " -> ".join(self.project.critical_path)
            messagebox.showinfo("Critical Path", f"Critical Path: {critical_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate schedule: {str(e)}")

    def generate_gantt(self):
        try:
            if not self.project.activities:
                messagebox.showwarning("Warning", "No activities to display")
                return
            
            self.project.calculate_schedule()
            
            fig = self.project.create_gantt_figure()
            if fig:
                # Create new window for Gantt chart
                gantt_window = tk.Toplevel(self.root)
                gantt_window.title("Gantt Chart")
                gantt_window.geometry("1200x800")
                
                canvas = FigureCanvasTkAgg(fig, gantt_window)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
            else:
                messagebox.showerror("Error", "Failed to generate Gantt chart")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Gantt chart: {str(e)}")

    def save_as_png(self):
        """Save the current Gantt chart as PNG file"""
        try:
            if not self.project.activities:
                messagebox.showwarning("Warning", "No activities to save. Please generate a Gantt chart first.")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                title="Save Gantt Chart as PNG"
            )
            
            if filename:
                if self.project.save_as_png(filename):
                    messagebox.showinfo("Success", f"Gantt chart saved as PNG: {filename}")
                else:
                    messagebox.showerror("Error", "Failed to save Gantt chart as PNG")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save PNG: {str(e)}")

    def export_to_excel(self):
        try:
            if not self.project.activities:
                messagebox.showwarning("Warning", "No activities to export")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                self.project.export_to_excel(filename)
                messagebox.showinfo("Success", f"Project exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {str(e)}")

    def export_to_primavera(self):
        try:
            if not self.project.activities:
                messagebox.showwarning("Warning", "No activities to export")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xml",
                filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
            )
            if filename:
                self.project.export_to_primavera(filename)
                messagebox.showinfo("Success", f"Project exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Primavera XML: {str(e)}")

    def update_activity_list(self):
        # Clear existing items
        for item in self.activities_tree.get_children():
            self.activities_tree.delete(item)
        
        # Add activities
        for activity_id, activity in sorted(self.project.activities.items()):
            self.activities_tree.insert('', tk.END, values=(activity.id, activity.name, activity.duration))

    def update_dependencies_list(self):
        # Clear existing items
        for item in self.deps_tree.get_children():
            self.deps_tree.delete(item)
        
        # Add dependencies
        for activity_id, activity in self.project.activities.items():
            for dep in activity.dependencies:
                self.deps_tree.insert('', tk.END, values=(
                    dep.from_activity_id, 
                    dep.to_activity_id, 
                    dep.rel_type.value,
                    dep.lag
                ))

    def clear_activity_fields(self):
        self.activity_id_var.set("")
        self.activity_name_var.set("")
        self.duration_var.set("")

    def clear_dependency_fields(self):
        self.from_activity_var.set("")
        self.to_activity_var.set("")
        self.lag_var.set("0")

def main():
    root = tk.Tk()
    app = GanttChartApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()