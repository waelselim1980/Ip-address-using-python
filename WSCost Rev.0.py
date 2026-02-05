import datetime
import socket
import tkinter as tk
from tkinter.ttk import *
from tkinter import messagebox
import ntplib # pyright: ignore[reportMissingImports]
import sys
from openpyxl import Workbook # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # pyright: ignore[reportMissingModuleSource]

# Hard License System Configuration
# Set expiry date to August 31, 2026
TARGET_EXPIRY_DATE = datetime.datetime(2026, 10, 31, 23, 59, 59)
NTP_SERVERS = [
    "pool.ntp.org",
    "time.google.com",
    "time.cloudflare.com",
    "time.windows.com",
    "time.apple.com"
]

class HardLicenseManager:
    def __init__(self):
        pass

    def _get_ntp_time(self):
        """
        Get accurate time from NTP servers.
        Tries multiple servers and falls back to system time if all fail.
        """
        for server in NTP_SERVERS:
            try:
                client = ntplib.NTPClient()
                # Request NTP response with a timeout
                response = client.request(server, version=3, timeout=5)
                # Convert NTP timestamp to datetime object
                ntp_time = datetime.datetime.fromtimestamp(response.tx_time)
                print(f"NTP time retrieved from {server}: {ntp_time}")
                return ntp_time
            except Exception as e:
                print(f"Failed to get time from {server}: {e}")
                continue

        # If all NTP servers fail, use system time (less secure but allows offline use)
        print("Warning: Could not connect to NTP servers, using system time.")
        return datetime.datetime.now()

    def verify_license(self):
        """
        Verify if the software license is still valid based on the current date.
        Compares current time (preferably NTP) against a hardcoded expiry date.
        """
        try:
            # Get current time using the NTP time retrieval method
            current_time = self._get_ntp_time()

            # Check if license has expired
            if current_time > TARGET_EXPIRY_DATE:
                print(f"License expired on: {TARGET_EXPIRY_DATE}")
                print(f"Current time: {current_time}")
                return False

            # Calculate days remaining until expiry
            days_remaining = (TARGET_EXPIRY_DATE - current_time).days
            print(f"License is valid. Days remaining: {days_remaining}")
            print(f"License expires on: {TARGET_EXPIRY_DATE}")
            return True

        except Exception as e:
            print(f"License verification failed: {e}")
            return False

# Initialize license manager
license_manager = HardLicenseManager()

# Verify license before running the main program
if not license_manager.verify_license():
    print("\n" + "="*50)
    print("LICENSE VERIFICATION FAILED")
    print("="*50)
    print("This software license has expired.")
    print(f"Software was valid until: {TARGET_EXPIRY_DATE}")
    print("Please contact the developer for a new license.")
    print("="*50)
    input("Press Enter to exit...")
    sys.exit()

# Original code continues here...
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
local_ip_address = str(IPAddr)
print(local_ip_address)
current_datetime = datetime.datetime.now()
g=str(IPAddr)
w=str('192.168.1.163')
if(g!=w):
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")
# creating tkinter window
# Open the output file in append mode
f = open("wscost.txt", "a")
print("This software is created by Engineer Wael Sherif Selim, email:wael.sherif.selim@gmail.com")
print("This software is created by Engineer Wael Sherif Selim, email:wael.sherif.selim@gmail.com", file=f)
print("Wscost")
print("Wscost", file=f)

# Define the fields for the project cost calculation form
fields = (
    'project name',
    'project id',
    'project duration in months',
    'materials cost',
    'workmanship or outsourced subs cost',
    'total monthly salary of all employees in site including bonus and allowances',
    'cost of caravans and temporary buildings for employees',
    'cost of monthly operation of caravans and temporary buildings including water,electicity,internet,printing...etc.',
    'monthly transportation cost of employees and equipment',
    'monthly cost of equipments',
    'monthly accomodation cost of employees',
    'monthly total depreciation cost of material,equipments,cranes,vehicles and engineering tools',
    'total monthly salary of all employees in headoffice including bonus and allowances',
    'cost of monthly expenses of headoffice employees including water,electicity,internet,printing,storage,garage...etc.',
    'total cost of monthly project insurances',
    'total cost of monthly social insurances of all employees',
    'total cost of stamps and syndicates expenses',
    'total banking fees icluding fees of primary,final LC and advance payment...etc.',
    'total fees of consultant per month',
    'design fees',
    'tax%',
    'variation orders',
    'claims value',
    'miscellaneous fees',
    'overhead ,waste,transportation and profit percentage',
    'contingency reserve value',
    'project currency',
    'quantity of material or project area for price analysis , if not needed write 1',
    'unit of material or unit of project area'
)

def export_to_excel(project_data, filename=None):
    """
    Exports project cost data to an Excel file with professional formatting.
    """
    try:
        if filename is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wscost_{project_data['project_name']}_{timestamp}.xlsx"
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Cost Estimate"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        section_font = Font(name="Arial", size=11, bold=True)
        normal_font = Font(name="Arial", size=10)
        bold_font = Font(name="Arial", size=10, bold=True)
        currency_font = Font(name="Arial", size=10, color="0070C0")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "PROJECT COST ESTIMATE"
        cell.font = Font(name="Arial", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        row += 2
        
        # Software credits
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "Software created by Engineer Wael Sherif Selim"
        ws[f'A{row}'].font = Font(name="Arial", size=9, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "Email: wael.sherif.selim@gmail.com"
        ws[f'A{row}'].font = Font(name="Arial", size=9, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        # Project Information Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "PROJECT INFORMATION"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        project_info = [
            ("Project Name:", project_data['project_name']),
            ("Project ID:", project_data['project_id']),
            ("Project Duration:", f"{project_data['duration']} months"),
            ("Project Currency:", project_data['currency']),
            ("Report Date:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for label, value in project_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = normal_font
            row += 1
        
        row += 1
        
        # Direct Costs Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "DIRECT COSTS"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        # Column headers
        ws[f'A{row}'] = "Description"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "Currency"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = bold_font
            ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 1
        
        direct_costs = [
            ("Direct Materials", project_data['direct_materials']),
            ("Direct Labor", project_data['direct_labor']),
            ("Direct Equipment", project_data['direct_equipment']),
            ("Direct Subcontractors", project_data['direct_subs']),
            ("Other Direct Costs", project_data['direct_other'])
        ]
        
        for desc, amount in direct_costs:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = project_data['currency']
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = currency_font
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].font = normal_font
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1
        
        # Total Direct Costs
        ws[f'A{row}'] = "TOTAL DIRECT COSTS"
        ws[f'B{row}'] = project_data['total_direct']
        ws[f'C{row}'] = project_data['currency']
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = Font(name="Arial", size=10, bold=True, color="0070C0")
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].font = bold_font
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 2
        
        # Indirect Costs Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "INDIRECT COSTS"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        # Column headers
        ws[f'A{row}'] = "Description"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "Currency"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = bold_font
            ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 1
        
        indirect_costs = [
            ("Overhead Costs", project_data['overhead']),
            ("Administrative Costs", project_data['admin']),
            ("Other Indirect Costs", project_data['other_indirect'])
        ]
        
        for desc, amount in indirect_costs:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = project_data['currency']
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = currency_font
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].font = normal_font
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1
        
        # Total Indirect Costs
        ws[f'A{row}'] = "TOTAL INDIRECT COSTS"
        ws[f'B{row}'] = project_data['total_indirect']
        ws[f'C{row}'] = project_data['currency']
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = Font(name="Arial", size=10, bold=True, color="0070C0")
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].font = bold_font
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 2
        
        # Margin/Profit Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "MARGIN/PROFIT & OTHER COSTS"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        # Column headers
        ws[f'A{row}'] = "Description"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "Currency"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = bold_font
            ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 1
        
        margin_costs = [
            ("Base Cost (Direct + Indirect)", project_data['base_cost']),
            (f"Profit & Overhead ({project_data['profit_percent']}%)", project_data['profit_margin']),
            (f"Tax ({project_data['tax_percent']}%)", project_data['tax']),
            ("Contingency Reserve", project_data['contingency'])
        ]
        
        for desc, amount in margin_costs:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = project_data['currency']
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = currency_font
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].font = normal_font
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 2
        
        # Total Project Cost
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL PROJECT COST"
        cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        ws[f'A{row}'] = "TOTAL PROJECT COST"
        ws[f'B{row}'] = project_data['total_cost']
        ws[f'C{row}'] = project_data['currency']
        ws[f'A{row}'].font = Font(name="Arial", size=12, bold=True)
        ws[f'B{row}'].font = Font(name="Arial", size=12, bold=True, color="C00000")
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].font = Font(name="Arial", size=12, bold=True)
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[f'{col}{row}'].border = border
        row += 2
        
        # Unit Cost Analysis
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "UNIT COST ANALYSIS"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1
        
        ws[f'A{row}'] = f"Quantity/Area ({project_data['unit']})"
        ws[f'B{row}'] = project_data['quantity']
        ws[f'C{row}'] = project_data['unit']
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = normal_font
            ws[f'{col}{row}'].border = border
        row += 1
        
        ws[f'A{row}'] = f"Price per {project_data['unit']}"
        ws[f'B{row}'] = project_data['price_per_unit']
        ws[f'C{row}'] = project_data['currency']
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = Font(name="Arial", size=10, bold=True, color="0070C0")
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].font = bold_font
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{col}{row}'].border = border
        
        # Save the workbook
        wb.save(filename)
        print(f"\nExcel file saved successfully: {filename}")
        messagebox.showinfo("Export Successful", f"Excel file saved successfully:\n{filename}")
        return True
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        messagebox.showerror("Export Error", f"Failed to export to Excel:\n{str(e)}")
        return False

def project_name(entries):
    """
    Calculates and displays project cost based on user inputs from the Tkinter form.
    Prints results to console and to the 'wscost.txt' file.
    """
    # Retrieve and convert all numerical inputs to float immediately
    pr = entries['project name'].get()
    prid = entries['project id'].get()
    bu = float(entries['project duration in months'].get())
    mat = float(entries['materials cost'].get())
    wrksub = float(entries['workmanship or outsourced subs cost'].get())
    tem = float(entries['total monthly salary of all employees in site including bonus and allowances'].get())
    h = float(entries['cost of caravans and temporary buildings for employees'].get())
    i = float(entries['cost of monthly operation of caravans and temporary buildings including water,electicity,internet,printing...etc.'].get())
    tre = float(entries['monthly transportation cost of employees and equipment'].get())
    etre = float(entries['monthly cost of equipments'].get())
    acm = float(entries['monthly accomodation cost of employees'].get())
    depe = float(entries['monthly total depreciation cost of material,equipments,cranes,vehicles and engineering tools'].get())
    ind = float(entries['total monthly salary of all employees in headoffice including bonus and allowances'].get())
    opc = float(entries['cost of monthly expenses of headoffice employees including water,electicity,internet,printing,storage,garage...etc.'].get())
    pri = float(entries['total cost of monthly project insurances'].get())
    spri = float(entries['total cost of monthly social insurances of all employees'].get())
    stmp = float(entries['total cost of stamps and syndicates expenses'].get())
    comm = float(entries['total banking fees icluding fees of primary,final LC and advance payment...etc.'].get())
    cons = float(entries['total fees of consultant per month'].get())
    desf = float(entries['design fees'].get())
    varo = float(entries['variation orders'].get())
    clmv = float(entries['claims value'].get())
    mscf = float(entries['miscellaneous fees'].get())
    prof = float(entries['overhead ,waste,transportation and profit percentage'].get())
    contg = float(entries['contingency reserve value'].get())
    prc = entries['project currency'].get()
    qmp = float(entries['quantity of material or project area for price analysis , if not needed write 1'].get())
    unma = entries['unit of material or unit of project area'].get()
    taxp = float(entries['tax%'].get())

    # Print project information
    print("\n" + "="*50)
    print("PROJECT INFORMATION")
    print("="*50)
    print(f"Project: {pr}")
    print(f"Project ID: {prid}")
    print(f"Project Duration: {bu} months")
    print(f"Project Currency: {prc}")
    print("="*50 + "\n")

    # Print to file
    print("\n" + "="*50, file=f)
    print("PROJECT INFORMATION", file=f)
    print("="*50, file=f)
    print(f"Project: {pr}", file=f)
    print(f"Project ID: {prid}", file=f)
    print(f"Project Duration: {bu} months", file=f)
    print(f"Project Currency: {prc}", file=f)
    print("="*50 + "\n", file=f)

    # Calculate and print DIRECT COSTS
    print("\n" + "="*50)
    print("DIRECT COSTS")
    print("="*50)
    direct_materials = mat
    direct_labor = tem * bu
    direct_equipment = etre * bu
    direct_subs = wrksub
    direct_other = (h + i + tre + acm + depe) * bu
    
    print(f"Direct Materials: {direct_materials} {prc}")
    print(f"Direct Labor: {direct_labor} {prc}")
    print(f"Direct Equipment: {direct_equipment} {prc}")
    print(f"Direct Subcontractors: {direct_subs} {prc}")
    print(f"Other Direct Costs: {direct_other} {prc}")
    total_direct = direct_materials + direct_labor + direct_equipment + direct_subs + direct_other
    print("-"*50)
    print(f"TOTAL DIRECT COSTS: {total_direct} {prc}")
    print("="*50 + "\n")

    # Print to file
    print("\n" + "="*50, file=f)
    print("DIRECT COSTS", file=f)
    print("="*50, file=f)
    print(f"Direct Materials: {direct_materials} {prc}", file=f)
    print(f"Direct Labor: {direct_labor} {prc}", file=f)
    print(f"Direct Equipment: {direct_equipment} {prc}", file=f)
    print(f"Direct Subcontractors: {direct_subs} {prc}", file=f)
    print(f"Other Direct Costs: {direct_other} {prc}", file=f)
    print("-"*50, file=f)
    print(f"TOTAL DIRECT COSTS: {total_direct} {prc}", file=f)
    print("="*50 + "\n", file=f)

    # Calculate and print INDIRECT COSTS
    print("\n" + "="*50)
    print("INDIRECT COSTS")
    print("="*50)
    overhead = (ind + opc + pri + spri + cons) * bu
    admin = (stmp + comm + desf) 
    other_indirect = (varo + clmv + mscf)
    
    print(f"Overhead Costs: {overhead} {prc}")
    print(f"Administrative Costs: {admin} {prc}")
    print(f"Other Indirect Costs: {other_indirect} {prc}")
    total_indirect = overhead + admin + other_indirect
    print("-"*50)
    print(f"TOTAL INDIRECT COSTS: {total_indirect} {prc}")
    print("="*50 + "\n")

    # Print to file
    print("\n" + "="*50, file=f)
    print("INDIRECT COSTS", file=f)
    print("="*50, file=f)
    print(f"Overhead Costs: {overhead} {prc}", file=f)
    print(f"Administrative Costs: {admin} {prc}", file=f)
    print(f"Other Indirect Costs: {other_indirect} {prc}", file=f)
    print("-"*50, file=f)
    print(f"TOTAL INDIRECT COSTS: {total_indirect} {prc}", file=f)
    print("="*50 + "\n", file=f)

    # Calculate and print MARGIN/PROFIT
    print("\n" + "="*50)
    print("MARGIN/PROFIT & OTHER COSTS")
    print("="*50)
    base_cost = total_direct + total_indirect
    profit_margin = prof / 100 * base_cost
    tax = taxp / 100 * base_cost
    
    print(f"Base Cost (Direct + Indirect): {base_cost} {prc}")
    print(f"Profit & Overhead ({prof}%): {profit_margin} {prc}")
    print(f"Tax ({taxp}%): {tax} {prc}")
    print(f"Contingency Reserve: {contg} {prc}")
    print("="*50 + "\n")

    # Print to file
    print("\n" + "="*50, file=f)
    print("MARGIN/PROFIT & OTHER COSTS", file=f)
    print("="*50, file=f)
    print(f"Base Cost (Direct + Indirect): {base_cost} {prc}", file=f)
    print(f"Profit & Overhead ({prof}%): {profit_margin} {prc}", file=f)
    print(f"Tax ({taxp}%): {tax} {prc}", file=f)
    print(f"Contingency Reserve: {contg} {prc}", file=f)
    print("="*50 + "\n", file=f)

    # Calculate and print TOTAL COST
    print("\n" + "="*50)
    print("TOTAL PROJECT COST")
    print("="*50)
    total_cost = base_cost + profit_margin + tax + contg
    print(f"TOTAL PROJECT COST: {total_cost} {prc}")
    print("="*50 + "\n")

    # Calculate and print COST PER UNIT
    print("\n" + "="*50)
    print("UNIT COST ANALYSIS")
    print("="*50)
    # Use max(1.0, qmp) to prevent division by zero if qmp is 0 or less
    price_per_unit = total_cost / (max(1.0, qmp))
    print(f"Quantity/Area: {qmp} {unma}")
    print(f"Price per {unma}: {price_per_unit} {prc}")
    print("="*50 + "\n")

    # Print to file
    print("\n" + "="*50, file=f)
    print("TOTAL PROJECT COST", file=f)
    print("="*50, file=f)
    print(f"TOTAL PROJECT COST: {total_cost} {prc}", file=f)
    print("="*50 + "\n", file=f)

    print("\n" + "="*50, file=f)
    print("UNIT COST ANALYSIS", file=f)
    print("="*50, file=f)
    print(f"Quantity/Area: {qmp} {unma}", file=f)
    print(f"Price per {unma}: {price_per_unit} {prc}", file=f)
    print("="*50 + "\n", file=f)

    # Prepare data for Excel export
    project_data = {
        'project_name': pr,
        'project_id': prid,
        'duration': bu,
        'currency': prc,
        'direct_materials': direct_materials,
        'direct_labor': direct_labor,
        'direct_equipment': direct_equipment,
        'direct_subs': direct_subs,
        'direct_other': direct_other,
        'total_direct': total_direct,
        'overhead': overhead,
        'admin': admin,
        'other_indirect': other_indirect,
        'total_indirect': total_indirect,
        'base_cost': base_cost,
        'profit_margin': profit_margin,
        'profit_percent': prof,
        'tax': tax,
        'tax_percent': taxp,
        'contingency': contg,
        'total_cost': total_cost,
        'quantity': qmp,
        'unit': unma,
        'price_per_unit': price_per_unit
    }
    
    # Export to Excel
    export_to_excel(project_data)

def makeform(root_window, parent_frame):
    """
    Creates the Tkinter form with input fields and buttons.
    """
    entries = {}
    row = 0
    for field in fields:
        # Create a label for each field
        lab = tk.Label(parent_frame, width=110, text=field + ": ", anchor='w', font=("georgia", 10))
        lab.grid(row=row, column=0)
        # Create an entry widget for each field
        ent = tk.Entry(parent_frame)
        ent.grid(row=row, column=1)
        # Set default value to "1" for numerical inputs
        ent.insert(0, "1")
        entries[field] = ent
        row += 1

    # Create a frame for buttons
    button_frame = tk.Frame(parent_frame)
    # Create "Run" button to trigger calculations
    run_button = tk.Button(button_frame, padx=1, pady=1, text="Run", font=("georgia", 10),
                           command=(lambda e=entries: project_name(e)))
    # Create "Quit" button to close the application
    quit_button = tk.Button(button_frame, padx=1, pady=1, text='Quit', font=("georgia", 10),
                            command=root_window.destroy)

    button_frame.grid(row=row, column=0, columnspan=2, sticky="nsew") # Span across two columns
    run_button.pack(side="left", padx=5, pady=5)
    quit_button.pack(side="left", padx=5, pady=5)

def onFrameConfigure(canvas_widget):
    """
    Resets the scroll region of the canvas to encompass the inner frame's content.
    This ensures the scrollbar works correctly as content is added/resized.
    """
    canvas_widget.configure(scrollregion=canvas_widget.bbox("all"))

if __name__ == '__main__':
    # Initialize the main Tkinter window
    root = tk.Tk()
    root.title("wscost")

    # Create a canvas to hold the scrollable frame
    canvas = tk.Canvas(root, width=950, borderwidth=0)
    # Create a frame inside the canvas to contain the form elements
    frame = tk.Frame(canvas)
    # Create a vertical scrollbar
    vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    # Configure the canvas to use the scrollbar
    canvas.configure(yscrollcommand=vsb.set)

    # Pack the scrollbar and canvas
    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    # Create a window on the canvas to place the frame
    canvas.create_window((6, 6), window=frame, anchor="nw")

    # Bind the frame's configure event to update the scroll region
    frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))

    # Create the form elements within the frame
    makeform(root, frame)

    # Start the Tkinter event loop
    root.mainloop()

# This part runs after the Tkinter window is closed
input("Press Enter to close program and output file is created wscost.txt")
f.close() # Close the output file
