import matplotlib.pyplot as plt # pyright: ignore[reportMissingModuleSource]
import math
import datetime
from datetime import timezone
import socket
import tkinter as tk
from tkinter.ttk import * 
from tkinter import messagebox, filedialog
from time import strftime
import ntplib # pyright: ignore[reportMissingImports]
import sys
import openpyxl # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]

hostname = socket.gethostname()    
IPAddr = socket.gethostbyname(hostname) 
hostname = socket.gethostname()
IPAddr = socket.gethostbyname(hostname)
current_date = datetime.datetime.now()
g=str(IPAddr)
w=str('192.168.1.163')
if(g!=w):
    print("License Expired")
    print(1/0)
else:
    print("Software Activated")
# Global variable to store results for Excel export
calculation_results = {}

def get_international_date():
    """
    Queries an NTP server to get the current UTC time.
    Returns a datetime object with UTC timezone, or None if an error occurs.
    """
    try:
        client = ntplib.NTPClient()
        response = client.request("pool.ntp.org", version=3)
        utc_time = datetime.datetime.fromtimestamp(response.tx_time, tz=timezone.utc)
        return utc_time
    except Exception as e:
        print(f"Error getting international date from NTP server: {e}")
        try:
            local_time = datetime.datetime.now(tz=timezone.utc)
            print("Warning: Using local time as fallback")
            return local_time
        except Exception as e2:
            print(f"Error getting local time: {e2}")
            return None

def check_license_expiry():
    """
    Check if the license has expired by comparing current date with expiry date.
    Returns True if license is valid, False if expired.
    """
    try:
        today_utc_dt = get_international_date()
        
        if today_utc_dt is None:
            print("CRITICAL ERROR: Failed to retrieve current date. Application cannot proceed.")
            print("License validation requires accurate date/time information.")
            return False
        
        print(f"Current Date (UTC): {today_utc_dt.strftime('%Y-%m-%d %H:%M:%S UTC')}")
        
        expiry_date = datetime.datetime(2026, 8, 19, 0, 0, 0, tzinfo=timezone.utc)
        
        print(f"License expires on: {expiry_date.strftime('%Y-%m-%d %H:%M:%S UTC')}")
        
        if today_utc_dt > expiry_date:
            print("=" * 60)
            print("LICENSE EXPIRED")
            print("=" * 60)
            print("This software license has expired and is no longer usable.")
            print(f"Expiry date: {expiry_date.strftime('%Y-%m-%d')}")
            print(f"Current date: {today_utc_dt.strftime('%Y-%m-%d')}")
            print("Please contact the software vendor to renew your license.")
            print("=" * 60)
            return False
        else:
            days_remaining = (expiry_date - today_utc_dt).days
            print(f"License is valid. Days remaining: {days_remaining}")
            return True
    
    except Exception as e:
        print(f"Error during license validation: {e}")
        print("License validation failed. Application cannot proceed.")
        return False

def export_to_excel():
    """
    Export calculation results to an Excel file with formatting
    """
    if not calculation_results:
        messagebox.showwarning("No Data", "Please run calculations before exporting to Excel.")
        return
    
    try:
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Batching_Report_{calculation_results.get('project_id', 'report')}.xlsx"
        )
        
        if not filename:
            return
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batching Properties"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "CONCRETE BATCHING PROPERTIES REPORT - ACI 117"
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Software info
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "Created by Engineer Wael Sherif Selim (wael.sherif.selim@gmail.com)"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Project Information Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "PROJECT INFORMATION"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        project_info = [
            ("Project Name", calculation_results.get('project_name', 'N/A')),
            ("Project ID", calculation_results.get('project_id', 'N/A')),
            ("Building Name", calculation_results.get('building_name', 'N/A')),
            ("Member ID", calculation_results.get('member_id', 'N/A')),
            ("Volume of Concrete (m³)", calculation_results.get('volume', 'N/A')),
        ]
        
        for label, value in project_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Temperature Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "TEMPERATURE ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        ws[f'A{row}'] = "Temperature of Fresh Concrete (°C)"
        ws[f'B{row}'] = calculation_results.get('temperature', 'N/A')
        ws[f'C{row}'] = calculation_results.get('temp_status', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Required Tests (Slump, Air Content, etc.)"
        ws[f'B{row}'] = calculation_results.get('required_tests', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Compressive Strength Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "COMPRESSIVE STRENGTH ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        strength_data = [
            ("Specified Compressive Strength f'c (MPa)", calculation_results.get('fc_specified', 'N/A')),
            ("1st Cylinder Value (MPa)", calculation_results.get('cylinder_1', 'N/A')),
            ("2nd Cylinder Value (MPa)", calculation_results.get('cylinder_2', 'N/A')),
            ("3rd Cylinder Value (MPa)", calculation_results.get('cylinder_3', 'N/A')),
            ("Average Strength (MPa)", calculation_results.get('avg_strength', 'N/A')),
            ("Cylinder Sets Required", calculation_results.get('cylinder_sets', 'N/A')),
            ("Compressive Strength Status", calculation_results.get('fc_status', 'N/A')),
        ]
        
        for label, value in strength_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Chemical Content Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "CHEMICAL CONTENT ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        chemical_data = [
            ("Sulphate % to Cement Weight", calculation_results.get('sulphate_percent', 'N/A'), calculation_results.get('sulphate_status', 'N/A')),
            ("Chloride % to Cement Weight", calculation_results.get('chloride_percent', 'N/A'), calculation_results.get('chloride_status', 'N/A')),
        ]
        
        for label, value, status in chemical_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = status
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Slump Test Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "SLUMP TEST RESULTS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        slump_data = [
            ("Specified Slump (mm)", calculation_results.get('slump_specified', 'N/A')),
            ("Actual Slump (mm)", calculation_results.get('slump_actual', 'N/A')),
            ("Slump Status", calculation_results.get('slump_status', 'N/A')),
        ]
        
        for label, value in slump_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Air Content Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "AIR CONTENT ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        air_data = [
            ("Specified Air Content (%)", calculation_results.get('air_specified', 'N/A')),
            ("Actual Air Content (%)", calculation_results.get('air_actual', 'N/A')),
            ("Air Content Status", calculation_results.get('air_status', 'N/A')),
        ]
        
        for label, value in air_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Material Batching Section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "MATERIAL BATCHING TOLERANCES"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Add headers for batching table
        ws[f'A{row}'] = "Material"
        ws[f'B{row}'] = "Specified"
        ws[f'C{row}'] = "Actual"
        ws[f'D{row}'] = "Tolerance"
        ws[f'E{row}'] = "Status"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        row += 1
        
        batching_data = [
            ("Cementitious Material (ton)", 
             calculation_results.get('cement_specified', 'N/A'),
             calculation_results.get('cement_actual', 'N/A'),
             calculation_results.get('cement_tolerance', 'N/A'),
             calculation_results.get('cement_status', 'N/A')),
            ("Water & Ice (ton)", 
             calculation_results.get('water_specified', 'N/A'),
             calculation_results.get('water_actual', 'N/A'),
             calculation_results.get('water_tolerance', 'N/A'),
             calculation_results.get('water_status', 'N/A')),
            ("Admixture (kg)", 
             calculation_results.get('admix_specified', 'N/A'),
             calculation_results.get('admix_actual', 'N/A'),
             calculation_results.get('admix_tolerance', 'N/A'),
             calculation_results.get('admix_status', 'N/A')),
            ("Aggregate (ton)", 
             calculation_results.get('aggregate_specified', 'N/A'),
             calculation_results.get('aggregate_actual', 'N/A'),
             calculation_results.get('aggregate_tolerance', 'N/A'),
             calculation_results.get('aggregate_status', 'N/A')),
        ]
        
        for material, specified, actual, tolerance, status in batching_data:
            ws[f'A{row}'] = material
            ws[f'B{row}'] = specified
            ws[f'C{row}'] = actual
            ws[f'D{row}'] = tolerance
            ws[f'E{row}'] = status
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Apply borders to all cells with data
        for row_cells in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=5):
            for cell in row_cells:
                cell.border = border
                if not cell.alignment.horizontal:
                    cell.alignment = Alignment(vertical='center')
        
        # Save workbook
        wb.save(filename)
        messagebox.showinfo("Success", f"Report exported successfully to:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Error exporting to Excel:\n{str(e)}")

def run_application():
    """
    Contains the core logic of the application, including the license expiry check
    and the Tkinter GUI for budget calculation.
    """
    if not check_license_expiry():
        input("Press Enter to exit...")
        sys.exit(1)
    
    print("License validation successful. Starting application...")
    
    g = str(IPAddr)
    print(f"IP Address: {g}")
    
    create_gui()

def create_gui():
    """
    Create and run the main GUI application
    """
    global f
    f = open("batching properties ACI 117.txt", "a")
    
    print("This software is created by Engineer Wael Sherif Selim,email:wael.sherif.selim@gmail.com")
    print("This software is created by Engineer Wael Sherif Selim,email:wael.sherif.selim@gmail.com", file=f)
    print("Batching properties tolerance as per ACI 117")
    print("Batching properties tolerance as per ACI 117", file=f)
    
    fields = ('project name', 'project id','building name','member ID','temperature of fresh concrete','specified compressive strength fc in mpa','value of least cylinder compressive value in mpa','2nd value of least cylinder compressive value in mpa','3rd value of least cylinder compressive value in mpa','s04 % to cement weight','chlorides % to cement weight','specified slump in mm','actual slump in mm','volume of concrete in m3','specified air content %','actual air content %','specified cementitous material weight in ton','actual cementitous material in ton','specified water and ice weight in ton','actual water and ice weight in ton','specified admixture weight in kg','actual admixture weight in kg','specified aggregate weight in ton','actual aggregate weight in ton','for scale capacity write "s" or individual batching write "i"','scale capacity in ton')

    def project_name(entries):
        global calculation_results
        calculation_results = {}  # Reset results
        
        if not check_license_expiry():
            print("License expired during application usage!")
            root.destroy()
            return

        try:
            pr = entries['project name'].get()
            print("project:", pr)
            print("project:", pr, file=f)
            calculation_results['project_name'] = pr
            
            prid = entries['project id'].get()
            print("project id:", prid)
            print("project id:", prid, file=f)
            calculation_results['project_id'] = prid
            
            bu = entries['building name'].get()
            print("building:", bu)
            print("building:", bu, file=f)
            calculation_results['building_name'] = bu
            
            id1 = entries['member ID'].get()
            print("member id:", id1)
            print("member id:", id1, file=f)
            calculation_results['member_id'] = id1
            
            vol = float(entries['volume of concrete in m3'].get())
            print(f"volume of concrete= {vol:.2f} m3")
            print(f"volume of concrete= {vol:.2f} m3", file=f)
            calculation_results['volume'] = f"{vol:.2f}"
            
            tem = float(entries['temperature of fresh concrete'].get())
            print(f"temperature of fresh concrete= {tem:.2f} celsius")
            print(f"temperature of fresh concrete= {tem:.2f} celsius", file=f)
            calculation_results['temperature'] = f"{tem:.2f}"
            
            if 10 < tem <= 32:
                temp_msg = "Within tolerance (10-32°C)"
                print("temperature is within tolerance since it is between 10 and 32 celsius")
                print("temperature is within tolerance since it is between 10 and 32 celsius", file=f)
            else:
                temp_msg = "NOT within tolerance"
                print("temperature is not within tolerance since it is not between 10 and 32 celsius")
                print("temperature is not within tolerance since it is not between 10 and 32 celsius", file=f)
            calculation_results['temp_status'] = temp_msg
            
            num_tests = int(vol/40) + 1
            print(f"number of slump tests,number of air content,concrete temperature and unit weight measurements required= {num_tests} tests")
            print(f"number of slump tests,number of air content,concrete temperature and unit weight measurements required= {num_tests} tests", file=f)
            calculation_results['required_tests'] = num_tests
            
            h = float(entries['specified compressive strength fc in mpa'].get())
            print(f"specified compressive strength= {h:.2f} mpa")
            print(f"specified compressive strength= {h:.2f} mpa", file=f)
            calculation_results['fc_specified'] = f"{h:.2f}"
            
            i = float(entries['value of least cylinder compressive value in mpa'].get())
            print(f"value of least cylinder compressive value in mpa= {i:.2f}")
            print(f"value of least cylinder compressive value in mpa= {i:.2f}", file=f)
            calculation_results['cylinder_1'] = f"{i:.2f}"
            
            j = float(entries['2nd value of least cylinder compressive value in mpa'].get())
            print(f"2nd value of least cylinder compressive value in mpa= {j:.2f}")
            print(f"2nd value of least cylinder compressive value in mpa= {j:.2f}", file=f)
            calculation_results['cylinder_2'] = f"{j:.2f}"
            
            k = float(entries['3rd value of least cylinder compressive value in mpa'].get())
            print(f"3rd value of least cylinder compressive value in mpa= {k:.2f}")
            print(f"3rd value of least cylinder compressive value in mpa= {k:.2f}", file=f)
            calculation_results['cylinder_3'] = f"{k:.2f}"
            
            avg_strength = (i + j + k) / 3
            calculation_results['avg_strength'] = f"{avg_strength:.2f}"
            
            sets = int(vol/80) + 1
            print(f"{sets} cylinder sets to be tested")
            print(f"{sets} cylinder sets to be tested", file=f)
            calculation_results['cylinder_sets'] = sets
            print("each set consists of 6 cylinders 2 tested after 7 days and 2 after 28 days and 2 additional tests as per owner request", file=f)
            print("each set consists of 6 cylinders 2 tested after 7 days and 2 after 28 days and 2 additional tests as per owner request")
            
            if avg_strength >= h:
                fc_msg = "Within tolerance"
                print("f'c is within tolerance", file=f)
                print("f'c is within tolerance")
            elif (j + i) / 2 + 3.4 >= h:
                fc_msg = "Within tolerance"
                print("f'c is within tolerance", file=f)
                print("f'c is within tolerance")
            else:
                fc_msg = "NOT within tolerance"
                print("f'c is not within tolerance", file=f)
                print("f'c is not within tolerance")
            calculation_results['fc_status'] = fc_msg
            
            s03c = float(entries['s04 % to cement weight'].get())
            print(f"sulphate % to cement weight= {s03c:.2f} %")
            print(f"sulphate % to cement weight= {s03c:.2f} %", file=f)
            calculation_results['sulphate_percent'] = f"{s03c:.2f}"
            
            if s03c <= 4:
                sulph_msg = "Within tolerance (<4%)"
                print("Sulphates in concrete within tolerance since less than 4% of cement weight", file=f)
                print("Sulphates in concrete within tolerance since less than 4% of cement weight")
            else:
                sulph_msg = "NOT within tolerance (>4%)"
                print("Sulphates in concrete not within tolerance since more than 4% of cement weight", file=f)
                print("Sulphates in concrete not within tolerance since more than 4% of cement weight")
            calculation_results['sulphate_status'] = sulph_msg
            
            chlc3 = float(entries['chlorides % to cement weight'].get())
            print(f"chloride % to cement weight= {chlc3:.2f} %")
            print(f"chloride % to cement weight= {chlc3:.2f} %", file=f)
            calculation_results['chloride_percent'] = f"{chlc3:.2f}"
            
            if chlc3 <= 0.10:
                chl_msg = "Within tolerance (<0.10%)"
                print("Chlorides within tolerance since less than 0.10%", file=f)
                print("Chlorides within tolerance since less than 0.10%")
            else:
                chl_msg = "NOT within tolerance (>0.10%)"
                print("Chlorides not within tolerance since more than 0.10%", file=f)
                print("Chlorides not within tolerance since more than 0.10%")
            calculation_results['chloride_status'] = chl_msg
            
            slf = float(entries['specified slump in mm'].get())
            acslf = float(entries['actual slump in mm'].get())
            calculation_results['slump_specified'] = f"{slf:.2f}"
            calculation_results['slump_actual'] = f"{acslf:.2f}"
            
            if acslf > slf:
                slump_msg = "NOT accepted (actual > specified)"
                print("Slump is not accepted")
                print("Slump is not accepted", file=f)
            elif (slf - acslf <= 12.5) and slf <= 75:
                slump_msg = "Within tolerance"
                print("Slump is within tolerance")
                print("Slump is within tolerance", file=f)
            elif (slf - acslf > 12.5) and slf <= 75:
                slump_msg = "NOT within tolerance"
                print("Slump is not within tolerance")
                print("Slump is not within tolerance", file=f)
            elif (slf - acslf <= 25) and 75 < slf <= 100:
                slump_msg = "Within tolerance"
                print("Slump is within tolerance")
                print("Slump is within tolerance", file=f)
            elif (slf - acslf > 25) and 75 < slf <= 100:
                slump_msg = "NOT within tolerance"
                print("Slump is not within tolerance")
                print("Slump is not within tolerance", file=f)
            elif (slf - acslf <= 35) and slf > 100:
                slump_msg = "Within tolerance"
                print("Slump is within tolerance")
                print("Slump is within tolerance", file=f)
            elif (slf - acslf > 35) and slf > 100:
                slump_msg = "NOT within tolerance"
                print("Slump is not within tolerance")
                print("Slump is not within tolerance", file=f)
            calculation_results['slump_status'] = slump_msg
            
            print("Tolerance of slump should be between 0 and -12.5 mm if specified slump <75 mm,and between 0 and -25 mm if specified slump less than 100mm, and between 0 and -35 mm if specified slump>100mm")
            print("Tolerance of slump should be between 0 and -12.5 mm if specified slump <75 mm,and between 0 and -25 mm if specified slump less than 100mm, and between 0 and -35 mm if specified slump>100mm", file=f)
            
            acof = float(entries['specified air content %'].get())
            print(f"specified air content= {acof:.2f} %")
            print(f"specified air content= {acof:.2f} %", file=f)
            calculation_results['air_specified'] = f"{acof:.2f}"
            
            actacof = float(entries['actual air content %'].get())
            print(f"actual air content= {actacof:.2f} %")
            print(f"actual air content= {actacof:.2f} %", file=f)
            calculation_results['air_actual'] = f"{actacof:.2f}"
            
            if (actacof - acof) < 1 and acof >= 4:
                air_msg = "Within tolerance"
                print("Air content is within tolerance")
                print("Air content is within tolerance", file=f)
            elif (actacof - acof) > 1 and acof >= 4:
                air_msg = "NOT within tolerance"
                print("Air content is not within tolerance")
                print("Air content is not within tolerance", file=f)
            else:
                air_msg = "No tolerance - actual should not exceed specified"
                print("There is no tolerance and actual air content should not exceed specified air content")
                print("There is no tolerance and actual air content should not exceed specified air content", file=f)
            calculation_results['air_status'] = air_msg
            
            print("Tolerance of air content is 0 if air content is specified and 1% if not specified then air content is taken 4% ")
            print("Tolerance of air content is 0 if air content is specified and 1% if not specified then air content is taken 4% ", file=f)
            
            cemmf = float(entries['specified cementitous material weight in ton'].get())
            print(f"specified cementitous material weight in ton= {cemmf:.2f} ton")
            print(f"specified cementitous material weight in ton= {cemmf:.2f} ton", file=f)
            calculation_results['cement_specified'] = f"{cemmf:.2f}"
            
            cemacf = float(entries['actual cementitous material in ton'].get())
            print(f"actual cementitous material weight in ton= {cemacf:.2f} ton")
            print(f"actual cementitous material weight in ton= {cemacf:.2f} ton", file=f)
            calculation_results['cement_actual'] = f"{cemacf:.2f}"
            
            cem_tol = 1/100 * cemmf
            calculation_results['cement_tolerance'] = f"±{cem_tol:.2f} ton"
            
            if cemacf - cemmf > -1/100 * cemmf:
                cem_msg = "Within tolerance"
                print("Cementitous weight material within tolerance")
                print("Cementitous weight material within tolerance", file=f)
            else:
                cem_msg = "NOT within tolerance"
                print("Cementitous weight material not within tolerance")
                print("Cementitous weight material not within tolerance", file=f)
            calculation_results['cement_status'] = cem_msg
            
            print(f"Tolerance of cementitous material is {cem_tol:.2f} Ton")
            print(f"Tolerance of cementitous material is {cem_tol:.2f} Ton", file=f)
            
            wicef = float(entries['specified water and ice weight in ton'].get())
            print(f"specified water and ice weight in ton= {wicef:.2f}")
            print(f"specified water and ice weight in ton= {wicef:.2f}", file=f)
            calculation_results['water_specified'] = f"{wicef:.2f}"
            
            wicacf = float(entries['actual water and ice weight in ton'].get())
            print(f"actual water and ice weight in ton= {wicacf:.2f}")
            print(f"actual water and ice weight in ton= {wicacf:.2f}", file=f)
            calculation_results['water_actual'] = f"{wicacf:.2f}"
            
            water_tol = 3/100 * wicef
            calculation_results['water_tolerance'] = f"±{water_tol:.2f} ton"
            
            if wicacf - wicef < 3/100 * wicef:
                water_msg = "Within tolerance"
                print("water content within tolerance")
                print("water content within tolerance", file=f)
            else:
                water_msg = "NOT within tolerance"
                print("water content not within tolerance")
                print("water content not within tolerance", file=f)
            calculation_results['water_status'] = water_msg
            
            print(f"Tolerance of water and ice content is {water_tol:.2f} ton")   
            print(f"Tolerance of water and ice content is {water_tol:.2f} ton", file=f)
            
            adrf = float(entries['specified admixture weight in kg'].get())
            print(f"specified admixture weight in kg= {adrf:.2f}")
            print(f"specified admixture weight in kg= {adrf:.2f}", file=f)
            calculation_results['admix_specified'] = f"{adrf:.2f}"
            
            adf = float(entries['actual admixture weight in kg'].get())
            print(f"actual admixture weight in kg= {adf:.2f}")
            print(f"actual admixture weight in kg= {adf:.2f}", file=f)
            calculation_results['admix_actual'] = f"{adf:.2f}"
            
            admix_tol = 3/100 * adrf
            calculation_results['admix_tolerance'] = f"±{admix_tol:.2f} kg"
            
            if (adf - adrf) >= -3/100 * adrf:
                admix_msg = "Within tolerance"
                print("Admixture weight within tolerance")
                print("Admixture weight within tolerance", file=f)
            else:
                admix_msg = "NOT within tolerance"
                print("Admixture weight not within tolerance")
                print("Admixture weight not within tolerance", file=f)
            calculation_results['admix_status'] = admix_msg
            
            print(f"Tolerance of admixture weight is {admix_tol:.2f} Kg")
            print(f"Tolerance of admixture weight is {admix_tol:.2f} Kg", file=f)
            
            agrf = float(entries['specified aggregate weight in ton'].get())
            print(f"specified aggregate weight in ton= {agrf:.2f}")
            print(f"specified aggregate weight in ton= {agrf:.2f}", file=f)
            calculation_results['aggregate_specified'] = f"{agrf:.2f}"
            
            agf = float(entries['actual aggregate weight in ton'].get())
            print(f"actual aggregate weight in ton= {agf:.2f}")
            print(f"actual aggregate weight in ton= {agf:.2f}", file=f)
            calculation_results['aggregate_actual'] = f"{agf:.2f}"
            
            sci = entries['for scale capacity write "s" or individual batching write "i"'].get()
            scapf = float(entries['scale capacity in ton'].get())
            
            if sci == "i":
                print("individual batch")
                print("individual batch", file=f)
                tol = 2/100 * agrf
            elif sci == "s":
                tol = min(1/100 * agrf, 0.3/100 * scapf, 3/100 * (agrf + adrf/1000 + wicef + cemmf))
            
            calculation_results['aggregate_tolerance'] = f"±{tol:.2f} ton"
            
            if agrf - agf < tol:
                agg_msg = "Within tolerance"
                print("Aggregate weight within tolerance")
                print("Aggregate weight within tolerance", file=f)
            else:
                agg_msg = "NOT within tolerance"
                print("Aggregate weight not within tolerance", file=f)
                print("Aggregate weight not within tolerance")
            calculation_results['aggregate_status'] = agg_msg
            
            print(f"Tolerance of aggregate weight is {tol:.2f} ton")
            print(f"Tolerance of aggregate weight is {tol:.2f} ton", file=f)
            
            # Show success message
            messagebox.showinfo("Calculation Complete", "Calculations completed successfully!\nYou can now export the results to Excel.")
                
        except Exception as e:
            print(f"Error processing data: {e}")
            print(f"Error processing data: {e}", file=f)
            messagebox.showerror("Error", f"Error processing data:\n{str(e)}")

    def makeform(root, frame):
        entries = {}
        row = 0
        for field in fields:
            lab = tk.Label(frame, width=115, text=field + ": ", anchor='w')
            lab.grid(row=row, column=0)
            ent = tk.Entry(frame)
            ent.grid(row=row, column=1)
            ent.insert(0, "0")
            entries[field] = ent
            row += 1
        f1 = tk.Frame(frame)
        b1 = tk.Button(f1, padx=5, pady=5, text="Run Calculation", command=(lambda e=entries: project_name(e)))
        b2 = tk.Button(f1, padx=5, pady=5, text='Export to Excel', command=export_to_excel, bg='#4CAF50', fg='white')
        b3 = tk.Button(f1, padx=5, pady=5, text='Quit', command=root.destroy)
        f1.grid(row=row, column=0, columnspan=2, sticky="nsew")
        b1.pack(side="left", padx=5)
        b2.pack(side="left", padx=5)
        b3.pack(side="left", padx=5)

    def onFrameConfigure(canvas):
        '''Reset the scroll region to encompass the inner frame'''
        canvas.configure(scrollregion=canvas.bbox("all"))

    # Create the main window
    root = tk.Tk()
    root.title("Batching properties ACI 117 - With Excel Export")
    canvas = tk.Canvas(root, width=950, borderwidth=0)
    frame = tk.Frame(canvas)
    vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((4, 4), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
    makeform(root, frame)
    root.mainloop()
    
    # Close the file when done
    f.close()

if __name__ == '__main__':
    # Run the application with license check
    run_application()
